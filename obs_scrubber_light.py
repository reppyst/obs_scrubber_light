#!/usr/bin/env python3

from __future__ import annotations

import sys, traceback, zipfile, random, string, time, copy, csv, re, os, json

from datetime import datetime

from pathlib import Path

from contextlib import contextmanager

import pandas as pd, math

from dataclasses import dataclass

from typing import Any, Dict, List, Optional, Set, Tuple, Iterable, Callable, Mapping

from collections import defaultdict

from PySide6.QtCore import (
    Qt,
    QObject,
    QThread,
    Signal,
    Slot,
    QUrl,
    QByteArray,
    QRectF,
    QEvent,
    QBuffer,
    QIODevice,
)

from PySide6.QtGui import QIcon, QPixmap, QImage, QPainter, QHelpEvent

try:
    from PySide6.QtQml import QJSEngine
except Exception:
    QJSEngine = None

try:
    from PySide6.QtSvg import QSvgRenderer
except Exception:
    QSvgRenderer = None

from PySide6.QtWidgets import (
    QApplication,
    QWidget,
    QTabWidget,
    QVBoxLayout,
    QHBoxLayout,
    QLabel,
    QPushButton,
    QFileDialog,
    QLineEdit,
    QTextEdit,
    QGroupBox,
    QGridLayout,
    QCheckBox,
    QMessageBox,
    QButtonGroup,
    QRadioButton,
    QComboBox,
    QFormLayout,
    QDialog,
    QTreeWidget,
    QTreeWidgetItem,
    QSplitter,
    QHeaderView,
    QTableWidget,
    QTableWidgetItem,
    QSpinBox,
    QDialogButtonBox,
    QMenu,
    QInputDialog,
    QAbstractItemView,
    QProgressBar,
    QToolTip,
)

try:
    from PySide6.QtWebEngineWidgets import QWebEngineView
except Exception:
    QWebEngineView = None

from obs_license import (
    DEFAULT_LICENSE_FILENAME,
    LicenseError,
    LicensePayload,
    cached_license_path,
    describe_license,
    load_first_valid_license,
    load_license_file,
    remember_license_path,
)

try:

    from lxml import etree as ET

    LXML = True

except Exception:

    import xml.etree.ElementTree as ET  # type: ignore

    LXML = False

__version__ = "1.18b"

JTDS_NS = "https://jtds.jten.mil"

NS = {"j": JTDS_NS}

LICENSE_INFO: Optional[LicensePayload] = None

def _resource_search_paths() -> List[Path]:
    bases: List[Path] = []
    seen: Set[Path] = set()

    def _add(path_like: Any, *, use_parent: bool = False) -> None:
        if not path_like:
            return
        try:
            candidate = Path(path_like)
        except Exception:
            return
        try:
            resolved = candidate.resolve(strict=False)
        except Exception:
            resolved = candidate
        if use_parent:
            resolved = resolved.parent
        if resolved in seen:
            return
        seen.add(resolved)
        bases.append(resolved)

    argv0 = sys.argv[0] if sys.argv else None
    if argv0:
        _add(argv0, use_parent=True)
    meipass = getattr(sys, "_MEIPASS", None)
    if meipass:
        _add(meipass)
    _add(__file__, use_parent=True)
    try:
        _add(Path.cwd())
    except Exception:
        pass
    if not bases:
        bases.append(Path.cwd())
    return bases


def _locate_resource_path(*relative: str) -> Path:
    bases = _resource_search_paths()
    for base in bases:
        candidate = base.joinpath(*relative)
        if candidate.exists():
            return candidate
    if bases:
        return bases[0].joinpath(*relative)
    return Path(*relative)


WARSIM_TYPENAME_MASTER = _locate_resource_path(
    "UnitClass-Masters",
    "WARSIM_UnitClass_TypeName-Master.csv",
)


_SIDC_SCHEME_NAMES = {
    "S": "Warfighting",
    "I": "Intelligence",
    "O": "Stability Operations",
    "E": "Emergency Management",
    "G": "Meteorological",
    "W": "Space",
    "X": "Other",
}

_SIDC_AFFILIATION_NAMES = {
    "P": "Pending",
    "U": "Unknown",
    "A": "Assumed Friendly",
    "F": "Friendly",
    "N": "Neutral",
    "H": "Hostile",
    "S": "Suspect",
    "J": "Joker",
    "K": "Faker",
    "L": "Exercise Friendly",
    "M": "Exercise Neutral",
    "D": "Exercise Hostile",
    "G": "Exercise Pending",
    "O": "Exercise Unknown",
}

_SIDC_DIMENSION_NAMES = {
    "P": "Space",
    "A": "Air",
    "F": "SOF",
    "G": "Ground",
    "S": "Sea Surface",
    "U": "Sea Subsurface",
    "Z": "Other",
}

_SIDC_STATUS_NAMES = {
    "P": "Present / Known",
    "A": "Anticipated / Planned",
    "C": "Present / Fully Capable",
    "D": "Present / Damaged",
    "X": "Present / Destroyed",
    "F": "Present / Full",
    "U": "Present / Unknown",
}


_MILSYMBOL_PREVIEW_HTML = """
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8" />
<title>MIL-STD-2525C Preview</title>
<style>
body {
    margin: 0;
    font-family: Arial, sans-serif;
    background: #f5f5f5;
    color: #222222;
}
#symbol-host {
    min-height: 135px;
    display: flex;
    align-items: center;
    justify-content: center;
    padding: 12px;
}
.message {
    text-align: center;
    opacity: 0.7;
}
svg {
    max-width: 100%;
    height: auto;
}
</style>
<script src="milsymbol.js"></script>
<script>
(function () {
    const hostId = "symbol-host";
    function setMessage(msg) {
        const host = document.getElementById(hostId);
        if (!host) {
            return;
        }
        host.innerHTML = "<div class='message'>" + msg + "</div>";
    }
    function getLibrary() {
        if (window.ms && typeof window.ms.Symbol === "function") {
            return window.ms;
        }
        if (window.milsymbol && typeof window.milsymbol.Symbol === "function") {
            return window.milsymbol;
        }
        return null;
    }
    window.renderSymbol = function (sidc) {
        const host = document.getElementById(hostId);
        if (!host) {
            return;
        }
        if (!sidc) {
            setMessage("Select an entry to preview the SIDC.");
            return;
        }
        const lib = getLibrary();
        if (!lib) {
            setMessage("milsymbol.js failed to load.");
            return;
        }
        try {
            const symbol = new lib.Symbol(sidc, { size: 180 });
            const svg = symbol.asSVG();
            host.innerHTML = "";
            if (typeof svg === "string") {
                host.innerHTML = svg;
            } else if (svg) {
                host.appendChild(svg);
            } else {
                setMessage("Rendered symbol had no output.");
            }
        } catch (err) {
            setMessage("Failed to render symbol: " + err);
        }
    };
    window.renderSymbol("");
})();
</script>
</head>
<body>
<div id="symbol-host"></div>
</body>
</html>
"""


@dataclass(frozen=True)
class _SidcAttributeLookup:

    affiliations: List[Tuple[str, str]]
    statuses: List[Tuple[str, str]]
    sizes: List[Tuple[str, str]]
    countries: List[Tuple[str, str]]
    orders_of_battle: List[Tuple[str, str]]


try:

    from dragon_importer_new import (
        import_dragon_enhanced as cli_import_dragon_enhanced,
        normalize_to_jtds_schema as cli_normalize_to_jtds_schema,
        append_unitclasses_to_master as cli_append_unitclasses_to_master,
        append_entityclasses_to_master as cli_append_entityclasses_to_master,
        DragonImportCancelled,
    )

except Exception:

    cli_import_dragon_enhanced = None
    cli_normalize_to_jtds_schema = None
    cli_append_unitclasses_to_master = None
    cli_append_entityclasses_to_master = None
    class DragonImportCancelled(Exception):
        pass


@dataclass
class ObsModel:

    tree: any
    root: ET._Element
    classdata: ET._Element
    scenario: ET._Element
    unit_list: ET._Element
    entcomp_list: ET._Element


def jfind(elem, path):
    return elem.find(path, NS)


def jfindall(elem, path):
    return elem.findall(path, NS)


def jtag(local: str) -> str:
    return f"{{{JTDS_NS}}}{local}"


def text(elem: Optional[ET._Element]) -> Optional[str]:
    return elem.text if elem is not None else None


def _normalize_user_data(value: Any) -> Any:
    """Ensure tuples stored via Qt item data survive PySide (which returns lists)."""
    if isinstance(value, list):
        return tuple(value)
    return value


def _indent_xml(elem: ET.Element, level: int = 0) -> None:

    indent = "\n" + "  " * level
    if len(elem):
        if not elem.text or not elem.text.strip():
            elem.text = indent + "  "
        for child in list(elem):
            _indent_xml(child, level + 1)
        if not elem.tail or not elem.tail.strip():
            elem.tail = indent
    else:
        if not elem.tail or not elem.tail.strip():
            elem.tail = indent


def _now_iso_seconds_no_tz() -> str:
    """Return 'YYYY-MM-DDTHH:MM:SS' in local time (no timezone suffix)."""
    return datetime.now().strftime("%Y-%m-%dT%H:%M:%S")


def _gen_unit_xml_id(used_ids: Set[str]) -> str:
    """
    Generate a unique Unit @id of the form ID_<epochSeconds>.
    If multiple are created within the same second, increment seconds to preserve uniqueness.
    """
    base = int(time.time())
    seq = 0
    while True:
        cand = f"ID_{base + seq}"
        if cand not in used_ids:
            used_ids.add(cand)
            return cand
        seq += 1


def load_unitlist_map(xml_path: str) -> Dict[str, Dict[str, str]]:
    """
    Build a mapping of MilStd2525CCode -> UnitDisEnumeration attributes
    from a reference unitlist.xml. Works with or without namespaces.
    """
    try:
        root = ET.parse(xml_path).getroot()
    except Exception as e:
        raise ValueError(f"Failed to parse reference XML: {e}")

    def local(t: str) -> str:
        return t.rsplit("}", 1)[-1] if "}" in t else t

    mapping: Dict[str, Dict[str, str]] = {}
    for unit in root.iter():
        if local(unit.tag) != "Unit":
            continue
        code_el = None
        ude_el = None
        for ch in list(unit):
            if local(ch.tag) == "MilStd2525CCode":
                code_el = ch
            elif local(ch.tag) == "UnitDisEnumeration":
                ude_el = ch
        if code_el is not None and code_el.text and ude_el is not None:
            code = code_el.text.strip()
            if code:
                mapping[code] = {k: v for k, v in ude_el.attrib.items()}
    return mapping


def _unitclass_master_search_bases() -> List[Path]:

    return _resource_search_paths()


def _default_unitclass_master_path() -> Optional[Path]:

    for base in _unitclass_master_search_bases():
        candidate = base / "UnitClass-Masters" / "unitclass-master.xlsx"
        if candidate.exists():
            return candidate
    bases = _unitclass_master_search_bases()
    if bases:
        return bases[0] / "UnitClass-Masters" / "unitclass-master.xlsx"
    return None


def _resolve_symbols_path(filename: str) -> Optional[Path]:

    for base in _unitclass_master_search_bases():
        candidate = base / "Symbols" / filename
        if candidate.exists():
            return candidate
    try:
        fallback = Path("Symbols") / filename
    except Exception:
        return None
    if fallback.exists():
        return fallback
    return None


def _is_valid_unit_xml_id(val: Optional[str]) -> bool:

    if not val or not val.startswith("ID_"):
        return False
    digits = val[3:]
    return len(digits) == 10 and digits.isdigit()


def _is_valid_timestamp(val: Optional[str]) -> bool:

    if not val:
        return False
    try:
        datetime.strptime(val.strip(), "%Y-%m-%dT%H:%M:%S")
        return True
    except Exception:
        return False


def ensure_unit_metadata(
    model: ObsModel,
    default_side: Optional[str] = None,
    target_units: Optional[Set[ET._Element]] = None,
) -> Tuple[int, int, int]:
    """Ensure each <Unit> has a unique ID/dateTimeStamp and populate SideSuperior
    when UnitSuperior is absent. Returns (ids_assigned, timestamps_assigned, sides_assigned).
    When target_units is provided, only those units are modified (aside from ID uniqueness checks).
    """
    side_hint = (
        default_side.strip().upper()
        if isinstance(default_side, str) and default_side.strip()
        else None
    )
    units = list(_iter_local(model.unit_list, "Unit"))
    target_set: Set[ET._Element] = (
        set(target_units) if target_units is not None else set(units)
    )
    if side_hint is None:
        known_sides: Set[str] = set()
        for unit in target_set:
            side_el = jfind(unit, "j:SideSuperior")
            if side_el is not None and side_el.text and side_el.text.strip():
                known_sides.add(side_el.text.strip().upper())
        if len(known_sides) == 1:
            side_hint = next(iter(known_sides))
    if side_hint is None:
        side_list = jfind(model.scenario, "j:SideList")
        if side_list is not None:
            scenario_sides: Set[str] = set()
            for side in jfindall(side_list, "j:Side"):
                nm = text(jfind(side, "j:Name"))
                if nm and nm.strip():
                    scenario_sides.add(nm.strip().upper())
            if len(scenario_sides) == 1:
                side_hint = next(iter(scenario_sides))
    used_xml_ids: Set[str] = set()
    for ec in _iter_local(model.entcomp_list, "EntityComposition"):
        eid = ec.get("id")
        if eid:
            used_xml_ids.add(eid)
    ids_assigned = 0
    stamps_assigned = 0
    sides_set = 0
    seen_unit_ids: Set[str] = set()
    for unit in units:
        existing_id = unit.get("id")
        if (
            not _is_valid_unit_xml_id(existing_id)
            or existing_id in used_xml_ids
            or existing_id in seen_unit_ids
        ):
            new_id = _gen_unit_xml_id(used_xml_ids)
            unit.set("id", new_id)
            existing_id = new_id
            if unit in target_set:
                ids_assigned += 1
        used_xml_ids.add(existing_id)
        seen_unit_ids.add(existing_id)
        if unit in target_set:
            stamp = unit.get("dateTimeStamp")
            if not _is_valid_timestamp(stamp):
                unit.set("dateTimeStamp", _now_iso_seconds_no_tz())
                stamps_assigned += 1
            sup_el = jfind(unit, "j:UnitSuperior")
            sup_text = sup_el.text.strip() if sup_el is not None and sup_el.text else ""
            if not sup_text and side_hint:
                side_el = jfind(unit, "j:SideSuperior")
                side_text = (
                    side_el.text.strip() if side_el is not None and side_el.text else ""
                )
                if side_el is None:
                    side_el = ET.SubElement(unit, jtag("SideSuperior"))
                if side_text.upper() != side_hint:
                    side_el.text = side_hint
                    sides_set += 1
    return ids_assigned, stamps_assigned, sides_set


def _resolve_unit_side(
    unit: ET._Element, lvc_to_unit: Dict[str, ET._Element]
) -> Optional[str]:

    seen: Set[int] = set()
    current = unit
    while current is not None and id(current) not in seen:
        seen.add(id(current))
        side_el = jfind(current, "j:SideSuperior")
        if side_el is not None and side_el.text and side_el.text.strip():
            return side_el.text.strip().upper()
        sup = text(jfind(current, "j:UnitSuperior"))
        if not sup:
            break
        current = lvc_to_unit.get(sup.strip())
    return None


def fix_empty_flags(model: ObsModel) -> Dict[str, Any]:

    if model is None:
        return {"created": 0, "per_side": {}, "skipped": 0}
    import re

    id_to_unit, _ = collect_unit_maps(model)
    used_xml_ids: Set[str] = set()
    used_lvc_ids: Set[str] = set()
    existing_superiors: Set[str] = set()
    existing_roles: Set[str] = set()
    unit_names: Dict[str, str] = {}
    for unit in _iter_local(model.unit_list, "Unit"):
        uid = unit.get("id")
        if uid:
            used_xml_ids.add(uid.strip())
        lvc = text(jfind(unit, "j:LvcId"))
        if lvc:
            val = lvc.strip()
            if val:
                used_lvc_ids.add(val)
                unit_names[val] = text(jfind(unit, "j:Name")) or val
    for ec in _iter_local(model.entcomp_list, "EntityComposition"):
        eid = ec.get("id")
        if eid:
            used_xml_ids.add(eid.strip())
        elvc = text(jfind(ec, "j:LvcId"))
        if elvc:
            used_lvc_ids.add(elvc.strip())
        sup = text(jfind(ec, "j:UnitSuperior"))
        if sup:
            existing_superiors.add(sup.strip())
        role = text(jfind(ec, "j:Role"))
        if role:
            existing_roles.add(role.strip())
    side_map = {
        "BLUFOR": "HMMWV ARMORED M1114 M2",
        "NEUTRAL": "URAL 375 PKM",
        "OPFOR": "BMP 3M",
    }

    def _role_token(value: Optional[str]) -> str:
        if not value:
            return "UNKNOWN"
        s = str(value).upper()
        s = re.sub(r"[^A-Z0-9]+", "_", s)
        s = re.sub(r"_+", "_", s).strip("_")
        return s or "UNKNOWN"

    created = 0
    skipped = 0
    per_side = {k: 0 for k in side_map}
    lvc_to_unit = id_to_unit
    next_role_num = 1
    for unit in _iter_local(model.unit_list, "Unit"):
        lvc = text(jfind(unit, "j:LvcId"))
        if not lvc:
            continue
        lvc = lvc.strip()
        if not lvc or lvc in existing_superiors:
            continue
        side = _resolve_unit_side(unit, lvc_to_unit)
        if not side:
            skipped += 1
            continue
        class_name = side_map.get(side)
        if not class_name:
            skipped += 1
            continue
        unit_name = unit_names.get(lvc) or lvc
        base_role = _role_token(class_name)
        unit_token = _role_token(unit_name)
        idx = next_role_num
        while True:
            candidate_role = f"{idx}_{base_role}_{unit_token}_CDR"
            if candidate_role not in existing_roles:
                break
            idx += 1
        role_text = candidate_role
        existing_roles.add(role_text)
        next_role_num = idx + 1
        ec = ET.SubElement(model.entcomp_list, jtag("EntityComposition"))
        ec.set("id", _gen_unit_xml_id(used_xml_ids))
        new_lvc = _gen_unique_lvcid("U", used_lvc_ids)
        ET.SubElement(ec, jtag("LvcId")).text = new_lvc
        used_lvc_ids.add(new_lvc)
        ET.SubElement(ec, jtag("Role")).text = role_text
        ET.SubElement(ec, jtag("UnitSuperior")).text = lvc
        ET.SubElement(ec, jtag("ClassName")).text = class_name
        ow = ET.SubElement(ec, jtag("OwningFederate"))
        ET.SubElement(ow, jtag("Federate")).text = "WARSIM"
        ET.SubElement(ec, jtag("IsInactive")).text = "false"
        ET.SubElement(ec, jtag("IsTransferrable")).text = "true"
        ET.SubElement(ec, jtag("IsInvincible")).text = "false"
        per_side[side] += 1
        created += 1
        existing_superiors.add(lvc)
    return {"created": created, "per_side": per_side, "skipped": skipped}


# ---------------- Core loaders/savers ----------------


def load_obs_xml(path: str) -> ObsModel:

    parser = ET.XMLParser(remove_blank_text=True) if LXML else None
    if path.lower().endswith(".zip"):
        with zipfile.ZipFile(path) as z:
            name = next((n for n in z.namelist() if n.lower().endswith(".xml")), None)
            if not name:
                raise ValueError("No .xml found inside zip")
            data = z.read(name)
            if LXML:
                root = (
                    ET.fromstring(data, parser=parser)
                    if parser is not None
                    else ET.fromstring(data)
                )
            else:
                root = ET.fromstring(data)

            class _Shim:
                def __init__(self, r):
                    self._r = r

                def getroot(self):
                    return self._r

            tree = _Shim(root)
    else:
        if LXML:
            tree_parsed = (
                ET.parse(path, parser=parser) if parser is not None else ET.parse(path)
            )
            root = tree_parsed.getroot()

            class _Shim:
                def __init__(self, r):
                    self._r = r

                def getroot(self):
                    return self._r

            tree = _Shim(root)
        else:
            tree = ET.parse(path)
            root = tree.getroot()
    classdata = jfind(root, "j:ClassData")
    scenario = jfind(root, "j:Scenario")
    if classdata is None or scenario is None:
        raise ValueError(
            "Not a recognized JTDS OBS XML (missing <ClassData> or <Scenario>)"
        )
    unit_list = jfind(scenario, "j:UnitList")
    entcomp_list = jfind(scenario, "j:EntityCompositionList")
    if unit_list is None or entcomp_list is None:
        raise ValueError("Missing <UnitList> or <EntityCompositionList>")
    return ObsModel(tree, root, classdata, scenario, unit_list, entcomp_list)


def _new_blank_model() -> ObsModel:

    if LXML:
        root = ET.Element(jtag("JTDS"), nsmap={None: JTDS_NS})
    else:
        root = ET.Element(jtag("JTDS"))
    classdata = ET.SubElement(root, jtag("ClassData"))
    scenario = ET.SubElement(root, jtag("Scenario"))
    unit_list = ET.SubElement(scenario, jtag("UnitList"))
    entcomp_list = ET.SubElement(scenario, jtag("EntityCompositionList"))

    class _Shim:
        def __init__(self, r):
            self._r = r

        def getroot(self):
            return self._r

    tree = _Shim(root)
    return ObsModel(tree, root, classdata, scenario, unit_list, entcomp_list)


def clone_obs_model(model: ObsModel) -> ObsModel:

    if model is None:
        raise ValueError("Cannot clone a null ObsModel")

    if LXML:
        parser = ET.XMLParser(remove_blank_text=True)
        xml_bytes = ET.tostring(model.root)
        root_copy = (
            ET.fromstring(xml_bytes, parser=parser) if parser is not None else ET.fromstring(xml_bytes)
        )

        class _Shim:
            def __init__(self, r):
                self._r = r

            def getroot(self):
                return self._r

        tree_copy = _Shim(root_copy)
    else:
        xml_bytes = ET.tostring(model.root)
        root_copy = ET.fromstring(xml_bytes)

        class _Shim:
            def __init__(self, r):
                self._r = r

            def getroot(self):
                return self._r

        tree_copy = _Shim(root_copy)

    classdata = jfind(root_copy, "j:ClassData")
    scenario = jfind(root_copy, "j:Scenario")
    if classdata is None or scenario is None:
        raise ValueError("Cloned OBS XML is missing required JTDS sections")
    unit_list = jfind(scenario, "j:UnitList")
    entcomp_list = jfind(scenario, "j:EntityCompositionList")
    if unit_list is None or entcomp_list is None:
        raise ValueError("Cloned OBS XML lacks UnitList/EntityCompositionList")

    cloned = ObsModel(tree_copy, root_copy, classdata, scenario, unit_list, entcomp_list)
    base_fields = {"tree", "root", "classdata", "scenario", "unit_list", "entcomp_list"}
    for key, value in vars(model).items():
        if key in base_fields:
            continue
        try:
            setattr(cloned, key, copy.deepcopy(value))
        except Exception:
            setattr(cloned, key, value)
    return cloned


def save_xml(model: ObsModel, out_path: str):

    ensure_unit_metadata(model)
    if LXML:
        xml_bytes = ET.tostring(
            model.root, pretty_print=True, xml_declaration=True, encoding="UTF-8"
        )
        with open(out_path, "wb") as f:
            f.write(xml_bytes)
    else:
        _indent_xml(model.root)
        ET.ElementTree(model.root).write(
            out_path, encoding="utf-8", xml_declaration=True
        )


# ---------------- Utilities ----------------


def collect_unit_maps(
    model: ObsModel,
) -> Tuple[Dict[str, ET._Element], Dict[str, List[ET._Element]]]:

    id_to_unit, parent_to_children = {}, {}
    for u in list(model.unit_list):
        lid = text(jfind(u, "j:LvcId"))
        if lid:
            id_to_unit[lid] = u
    for u in list(model.unit_list):
        lid = text(jfind(u, "j:LvcId"))
        parent = text(jfind(u, "j:UnitSuperior"))
        if parent:
            parent_to_children.setdefault(parent, []).append(u)
        if lid and lid not in parent_to_children:
            parent_to_children.setdefault(lid, [])
    return id_to_unit, parent_to_children


def clone_element(elem: ET._Element) -> ET._Element:
    """Return a deep copy of *elem* that is safe across XML backends."""
    return copy.deepcopy(elem)


DEFAULT_SIDE_ORDER = ["BLUFOR", "OPFOR", "NEUTRAL", "CIVILIAN", "UNKNOWN"]


def format_unit_label(unit: ET._Element) -> str:

    name = (text(jfind(unit, "j:Name")) or "").strip() or "Unnamed Unit"
    echelon = (text(jfind(unit, "j:Echelon")) or "").strip()
    lvc = (text(jfind(unit, "j:LvcId")) or "").strip()
    parts = [name]
    if echelon:
        parts.append(f"({echelon})")
    if lvc:
        parts.append(f"[{lvc}]")
    return " ".join(parts)


def format_entity_label(
    ec: ET._Element,
    *,
    include_superior: bool = False,
    superior_hint: Optional[str] = None,
) -> str:

    role = (text(jfind(ec, "j:Role")) or "").strip() or "Role"
    qty = (text(jfind(ec, "j:Quantity")) or "").strip()
    lvc = (text(jfind(ec, "j:LvcId")) or "").strip()
    parts = [role]
    if qty:
        parts.append(f"x{qty}")
    if lvc:
        parts.append(f"[{lvc}]")
    if include_superior:
        sup = superior_hint or (text(jfind(ec, "j:UnitSuperior")) or "").strip()
        if sup:
            parts.append(f"(Unit {sup})")
    return " ".join(parts)


def unit_sort_key(unit: ET._Element) -> Tuple[int, str, str]:

    echelon = (text(jfind(unit, "j:Echelon")) or "").strip()
    rank = ECHELON_RANK.get(echelon, len(ECHELON_ORDER))
    name = (text(jfind(unit, "j:Name")) or "").strip()
    lvc = (text(jfind(unit, "j:LvcId")) or "").strip()
    return (rank, name.lower(), lvc)


def entity_sort_key(ec: ET._Element) -> Tuple[str, str]:

    role = (text(jfind(ec, "j:Role")) or "").strip()
    lvc = (text(jfind(ec, "j:LvcId")) or "").strip()
    return (role.lower(), lvc)


def side_sort_key(side: str) -> Tuple[int, str]:

    try:
        idx = DEFAULT_SIDE_ORDER.index(side)
    except ValueError:
        idx = len(DEFAULT_SIDE_ORDER)
    return (idx, side)


def ensure_side_entries(model: ObsModel, sides: Iterable[str]) -> int:
    """Ensure that *sides* exist inside <SideList>; return how many were added."""
    normalized = [s.strip() for s in sides if s and s.strip()]
    if not normalized:
        return 0
    side_list = jfind(model.scenario, "j:SideList")
    if side_list is None:
        side_list = ET.SubElement(model.scenario, jtag("SideList"))
    existing: Set[str] = set()
    for side_elem in list(side_list):
        name = text(jfind(side_elem, "j:Name"))
        if name:
            existing.add(name.strip().upper())
    added = 0
    for side in normalized:
        key = side.upper()
        if key in existing:
            continue
        side_elem = ET.SubElement(side_list, jtag("Side"))
        ET.SubElement(side_elem, jtag("Name")).text = side
        existing.add(key)
        added += 1
    return added


NAME_ECHELON_TOKENS: Set[str] = {
    "TEAM",
    "DETACHMENT",
    "DET",
    "SECTION",
    "SECT",
    "SEC",
    "SQUAD",
    "SQD",
    "PLATOON",
    "PLT",
    "COMPANY",
    "COMP",
    "CO",
    "BATTERY",
    "BTRY",
    "BATTALION",
    "BN",
    "REGIMENT",
    "REGT",
    "BRIGADE",
    "BDE",
    "DIVISION",
    "DIV",
    "CORPS",
    "ARMY",
    "WING",
    "GROUP",
    "SQDN",
    "SQUADRON",
    "FLT",
    "FLIGHT",
    "HQ",
    "HQTRS",
    "COMMAND",
    "CMD",
    "TASKFORCE",
    "TF",
}


def _split_unit_name_components(name: str) -> Tuple[str, str]:

    tokens = [tok for tok in name.split("_") if tok]
    if not tokens:
        return name, ""
    for idx, tok in enumerate(tokens):
        up = tok.upper()
        if up in NAME_ECHELON_TOKENS or tok.isdigit():
            base = "_".join(tokens[:idx])
            suffix = "_".join(tokens[idx:])
            return base or name, suffix
    return name, ""


def _gather_unit_descendants(
    root: ET._Element,
    parent_to_children: Dict[str, List[ET._Element]],
    include_root: bool = False,
) -> List[ET._Element]:

    collected: List[ET._Element] = []
    visited: Set[int] = set()

    def visit(unit: ET._Element) -> None:
        key = id(unit)
        if key in visited:
            return
        visited.add(key)
        if unit is not root or include_root:
            collected.append(unit)
        lvc = (text(jfind(unit, "j:LvcId")) or "").strip()
        children = parent_to_children.get(lvc, []) if lvc else []
        for child in children:
            visit(child)

    visit(root)
    return collected


def _local(tag: str) -> str:
    return tag.rsplit("}", 1)[-1] if "}" in tag else tag


def _iter_local(root: ET._Element, local: str):

    for e in root.iter():
        if _local(e.tag) == local:
            yield e


_DISCODE_ATTRS = (
    "kind",
    "domain",
    "country",
    "category",
    "subcategory",
    "specific",
    "extra",
)

_UNIT_DIS_ENUM_ATTRS = (
    "kind",
    "domain",
    "country",
    "echelon",
    "type",
    "subtype",
    "modifier",
)

# Branch detection (token-based)

BRANCH_TOKENS = {
    "SIG": ["SIG", "SIGNAL"],
    "MNT": ["MNT", "MAINT"],
    "ENG": ["ENG", "ENGINEER", "ENGR"],
    "MED": ["MED", "MEDICAL"],
    "SPT": ["SPT", "SUPPORT", "CSS"],
    "CHEM": ["CHEM", "CBRN", "NBC"],
}


def unit_branch(u: ET._Element) -> Optional[str]:

    name = (text(jfind(u, "j:Name")) or "").upper()
    cls = (text(jfind(u, "j:ClassName")) or "").upper()
    code = (text(jfind(u, "j:MilStd2525CCode")) or "").upper()
    hay = " ".join([name, cls, code])
    for k, toks in BRANCH_TOKENS.items():
        if any(t in hay for t in toks):
            return k
    return None


HQ_KEYWORDS = ("HQ", "HEADQUARTERS")


def is_hq_unit(unit: ET._Element) -> bool:

    name = (text(jfind(unit, "j:Name")) or "").strip().upper()
    return any(keyword in name for keyword in HQ_KEYWORDS)


ECHELON_CANONICAL = {
    "TEAM": "Team",
    "FIRETEAM": "Team",
    "SECTION": "Section",
    "SECT": "Section",
    "SEC": "Section",
    "DETACHMENT": "Section",
    "DET": "Section",
    "SQUAD": "Squad",
    "SQD": "Squad",
    "PLATOON": "Platoon",
    "PLT": "Platoon",
    "COMPANY": "Company/Battery",
    "CO": "Company/Battery",
    "COMPANY/BATTERY": "Company/Battery",
    "BATTERY": "Company/Battery",
    "BTRY": "Company/Battery",
    "BATTALION": "Battalion",
    "BN": "Battalion",
    "REGIMENT": "Brigade",
    "REGT": "Brigade",
    "BRIGADE": "Brigade",
    "BDE": "Brigade",
    "DIVISION": "Division",
    "DIV": "Division",
    "CORPS": "Corps",
    "ARMY": "Army",
}

ECHELON_ORDER = [
    "Team",
    "Section",
    "Squad",
    "Platoon",
    "Company/Battery",
    "Battalion",
    "Brigade",
    "Division",
    "Corps",
    "Army",
]

ECHELON_RANK = {e: i for i, e in enumerate(ECHELON_ORDER)}

CONSOLIDATION_LEVELS = [
    "Section",
    "Platoon",
    "Company/Battery",
    "Battalion",
    "Brigade",
    "Division",
]


def normalize_echelon(value: Optional[str]) -> str:

    if not value:
        return ""
    key = str(value).strip().upper()
    return ECHELON_CANONICAL.get(key, str(value).strip())


def unit_echelon(u: ET._Element) -> str:

    direct = normalize_echelon(text(jfind(u, "j:Echelon")))
    if direct:
        return direct
    name_hint = (text(jfind(u, "j:Name")) or "").upper()
    for guess, token in [
        ("Platoon", " PLT"),
        ("Company/Battery", " CO "),
        ("Company/Battery", " CO_"),
        ("Company/Battery", " CO-"),
        ("Battalion", " BN"),
        ("Brigade", " BDE"),
        ("Division", " DIV"),
        ("Section", " SEC"),
        ("Squad", " SQD"),
        ("Team", " TEAM"),
    ]:
        if token in name_hint:
            return normalize_echelon(guess)
    return ""


def _entity_class_name(node: ET._Element) -> Optional[str]:

    for child in list(node):
        if _local(child.tag) == "ClassName" and child.text:
            return child.text.strip()
    if _local(node.tag) == "ClassName" and node.text:
        return node.text.strip()
    direct = jfind(node, "j:ClassName")
    if direct is not None and direct.text:
        return direct.text.strip()
    return None


def _collect_ecc_class_names(ecc: ET._Element) -> Set[str]:

    labels: Set[str] = set()
    for child in list(ecc):
        if _local(child.tag) in ("Name", "ClassName") and child.text:
            labels.add(child.text.strip())
    for tag in ("Name", "ClassName"):
        node = jfind(ecc, f"j:{tag}")
        if node is not None and node.text:
            labels.add(node.text.strip())
    return {label for label in labels if label}


def _normalize_discode_component(value: str) -> str:

    trimmed = value.strip()
    if trimmed.isdigit():
        return str(int(trimmed))
    return trimmed


def _parse_discode_query(raw: str) -> Dict[str, str]:

    text_val = (raw or "").strip()
    if not text_val:
        raise ValueError("Enter a DisCode value.")
    parts = [part.strip() for part in text_val.split(".") if part.strip()]
    if len(parts) != len(_DISCODE_ATTRS):
        raise ValueError(
            "DisCode must contain 7 dot-separated integers (kind.domain.country.category.subcategory.specific.extra)."
        )
    attrs: Dict[str, str] = {}
    for key, part in zip(_DISCODE_ATTRS, parts):
        if not part.isdigit():
            raise ValueError(f"DisCode component '{part}' for '{key}' must be numeric.")
        attrs[key] = str(int(part))
    return attrs


def _discode_matches(node: ET._Element, expected: Dict[str, str]) -> bool:

    for key, value in expected.items():
        actual = node.get(key)
        if actual is None:
            return False
        if _normalize_discode_component(actual) != value:
            return False
    return True


# Personnel crosswalk via ClassData (kind=3)


def _class_names_for_disc_kind(model: ObsModel, kind: str) -> Set[str]:

    names: Set[str] = set()
    for ecc in _iter_local(model.classdata, "EntityCompositionClass"):
        discode = next((ch for ch in list(ecc) if _local(ch.tag) == "DisCode"), None)
        if discode is None:
            discode = next((ch for ch in _iter_local(ecc, "DisCode")), None)
        if discode is not None and discode.get("kind") == kind:
            cname = None
            for cand in ["Name", "ClassName", "TypeName"]:
                el = next(
                    (x for x in list(ecc) if _local(x.tag) == cand and x.text), None
                )
                if el is not None:
                    cname = el.text.strip()
                    break
            if cname:
                names.add(cname)
    return names


def _personnel_class_names(model: ObsModel) -> Set[str]:

    return _class_names_for_disc_kind(model, "3")


def _munitions_class_names(model: ObsModel) -> Set[str]:

    return _class_names_for_disc_kind(model, "2")


def count_personnel(model: ObsModel) -> int:

    pnames = _personnel_class_names(model)
    if not pnames:
        return 0
    cnt = 0
    for child in list(model.entcomp_list):
        cname = None
        for e in list(child):
            if _local(e.tag) == "ClassName" and e.text:
                cname = e.text.strip()
                break
        if cname is None and _local(child.tag) == "ClassName" and child.text:
            cname = child.text.strip()
        if cname and cname in pnames:
            cnt += 1
    return cnt


def _remove_entities_for_kind(
    model: ObsModel, kind: str, *, removed_lvc_ids: Optional[Set[str]] = None
) -> Tuple[int, int]:

    names = _class_names_for_disc_kind(model, kind)
    if not names:
        return 0, 0
    removed_ec = 0
    for child in list(model.entcomp_list):
        if _local(child.tag) != "EntityComposition":
            continue
        matched = False
        cname = _entity_class_name(child)
        if cname and cname in names:
            matched = True
        else:
            discode = next(
                (ch for ch in list(child) if _local(ch.tag) == "DisCode"), None
            )
            if discode is None:
                discode = next((ch for ch in _iter_local(child, "DisCode")), None)
            if discode is not None and discode.get("kind") == kind:
                matched = True
        if not matched:
            continue
        if removed_lvc_ids is not None:
            lvc = (text(jfind(child, "j:LvcId")) or "").strip()
            if lvc:
                removed_lvc_ids.add(lvc.upper())
        model.entcomp_list.remove(child)
        removed_ec += 1
    removed_ecc = 0
    ecc_list = jfind(model.classdata, jtag("EntityCompositionClassList"))
    if ecc_list is not None:
        for ecc in list(ecc_list):
            discode = next(
                (ch for ch in list(ecc) if _local(ch.tag) == "DisCode"), None
            )
            if discode is None:
                discode = next((ch for ch in _iter_local(ecc, "DisCode")), None)
            if discode is not None and discode.get("kind") == kind:
                ecc_list.remove(ecc)
                removed_ecc += 1
    return removed_ec, removed_ecc


def _prune_entity_crew_lists(model: ObsModel, removed_lvc_ids: Set[str]) -> int:

    if not removed_lvc_ids:
        return 0
    normalized = {lid.strip().upper() for lid in removed_lvc_ids if lid and lid.strip()}
    if not normalized:
        return 0
    item_tags = {"Crew", "Passenger"}
    container_tags = {"CrewList", "PassengerList"}

    def _find_parent(root: ET._Element, target: ET._Element) -> Optional[ET._Element]:
        if hasattr(target, "getparent"):
            parent = target.getparent()  # type: ignore[attr-defined]
            if parent is not None:
                return parent
        for candidate in root.iter():
            for child in list(candidate):
                if child is target:
                    return candidate
        return None

    removed = 0
    for entity in list(model.entcomp_list):
        containers = [
            node for node in list(entity.iter()) if _local(node.tag) in container_tags
        ]
        for container in containers:
            items = [
                child for child in list(container) if _local(child.tag) in item_tags
            ]
            for item in items:
                raw_lvc = (
                    item.get("lvcId") or item.get("lvcID") or item.get("lvcid") or ""
                )
                crew_lvc = raw_lvc.strip()
                if crew_lvc and crew_lvc.upper() in normalized:
                    container.remove(item)
                    removed += 1
            remaining_items = [
                child for child in list(container) if _local(child.tag) in item_tags
            ]
            if not remaining_items:
                parent = _find_parent(entity, container)
                if parent is not None:
                    parent.remove(container)
                    if _local(parent.tag) == "Platform" and len(list(parent)) == 0:
                        grandparent = _find_parent(entity, parent)
                        if grandparent is not None:
                            grandparent.remove(parent)
                else:
                    if container in list(entity):
                        entity.remove(container)
    return removed


def remove_entities_by_discode(model: ObsModel, discode_value: str) -> Tuple[int, int]:

    attrs = _parse_discode_query(discode_value)
    matching_labels: Set[str] = set()
    removed_ecc = 0
    ecc_list = jfind(model.classdata, jtag("EntityCompositionClassList"))
    if ecc_list is not None:
        for ecc in list(ecc_list):
            discode = next(
                (ch for ch in list(ecc) if _local(ch.tag) == "DisCode"), None
            )
            if discode is None:
                discode = next((ch for ch in _iter_local(ecc, "DisCode")), None)
            if discode is not None and _discode_matches(discode, attrs):
                matching_labels.update(_collect_ecc_class_names(ecc))
                ecc_list.remove(ecc)
                removed_ecc += 1
    removed_ec = 0
    for entity in list(model.entcomp_list):
        matched = False
        cname = _entity_class_name(entity)
        if cname and cname in matching_labels:
            matched = True
        else:
            discode = next(
                (ch for ch in list(entity) if _local(ch.tag) == "DisCode"), None
            )
            if discode is None:
                discode = next((ch for ch in _iter_local(entity, "DisCode")), None)
            if discode is not None and _discode_matches(discode, attrs):
                matched = True
        if matched:
            model.entcomp_list.remove(entity)
            removed_ec += 1
    return removed_ec, removed_ecc


def replace_discode_entries(
    model: ObsModel, old_value: str, new_value: str
) -> Tuple[int, int]:

    old_attrs = _parse_discode_query(old_value)
    new_attrs = _parse_discode_query(new_value)
    updated_ecc = 0
    ecc_list = jfind(model.classdata, jtag("EntityCompositionClassList"))
    if ecc_list is not None:
        for ecc in list(ecc_list):
            discode = next(
                (ch for ch in list(ecc) if _local(ch.tag) == "DisCode"), None
            )
            if discode is None:
                discode = next((ch for ch in _iter_local(ecc, "DisCode")), None)
            if discode is not None and _discode_matches(discode, old_attrs):
                for key, value in new_attrs.items():
                    discode.set(key, value)
                updated_ecc += 1
    updated_entities = 0
    for entity in list(model.entcomp_list):
        discode = next((ch for ch in list(entity) if _local(ch.tag) == "DisCode"), None)
        if discode is None:
            discode = next((ch for ch in _iter_local(entity, "DisCode")), None)
        if discode is not None and _discode_matches(discode, old_attrs):
            for key, value in new_attrs.items():
                discode.set(key, value)
            updated_entities += 1
    return updated_ecc, updated_entities


def remove_unused_entity_composition_classes(
    model: ObsModel,
) -> Tuple[int, List[str], int]:

    used_names: Set[str] = set()
    for entity in list(model.entcomp_list):
        if _local(entity.tag) != "EntityComposition":
            continue
        cname = _entity_class_name(entity)
        if cname:
            used_names.add(cname)
    ecc_list = jfind(model.classdata, jtag("EntityCompositionClassList"))
    if ecc_list is None:
        return 0, [], 0
    removed = 0
    removed_labels: Set[str] = set()
    unnamed = 0
    for ecc in list(ecc_list):
        if _local(ecc.tag) != "EntityCompositionClass":
            continue
        labels = _collect_ecc_class_names(ecc)
        if labels & used_names:
            continue
        ecc_list.remove(ecc)
        removed += 1
        if labels:
            removed_labels.update(labels)
        else:
            unnamed += 1
    return removed, sorted(removed_labels), unnamed


def find_missing_entity_composition_classes(model: ObsModel) -> List[str]:

    used_names: Set[str] = set()
    for entity in list(model.entcomp_list):
        if _local(entity.tag) != "EntityComposition":
            continue
        cname = _entity_class_name(entity)
        if cname:
            used_names.add(cname)
    defined_names: Set[str] = set()
    ecc_list = jfind(model.classdata, jtag("EntityCompositionClassList"))
    if ecc_list is not None:
        for ecc in list(ecc_list):
            if _local(ecc.tag) != "EntityCompositionClass":
                continue
            defined_names.update(_collect_ecc_class_names(ecc))
    return sorted(name for name in used_names if name not in defined_names)


def remove_personnel(model: ObsModel) -> Tuple[int, int]:

    removed_lvc: Set[str] = set()
    removed_ec, removed_ecc = _remove_entities_for_kind(
        model, "3", removed_lvc_ids=removed_lvc
    )
    _prune_entity_crew_lists(model, removed_lvc)
    return removed_ec, removed_ecc


def remove_munitions(model: ObsModel) -> Tuple[int, int]:

    return _remove_entities_for_kind(model, "2")


def purge_classdata_sections(
    model: ObsModel, supply=False, emitter=False, munitions=False, flyout=False
) -> Dict[str, int]:

    counts = {}
    for want, tag in [
        (supply, "SupplyClassList"),
        (emitter, "EmitterClassList"),
        (munitions, "MunitionsList"),
        (flyout, "FlyOutList"),
    ]:
        if want:
            node = jfind(model.classdata, f"j:{tag}")
            counts[tag] = len(list(node)) if node is not None else 0
            if node is not None:
                model.classdata.remove(node)
    return counts


def remove_unit_subtree(model: ObsModel, root_lvc: str) -> Tuple[int, int]:

    if not root_lvc:
        return (0, 0)
    id_to_unit, parent_to_children = collect_unit_maps(model)
    to_remove: Set[str] = set()

    def mark(lvc: str) -> None:
        if not lvc or lvc in to_remove:
            return
        to_remove.add(lvc)
        for child in parent_to_children.get(lvc, []):
            child_lvc = (text(jfind(child, "j:LvcId")) or "").strip()
            if child_lvc:
                mark(child_lvc)

    mark(root_lvc)
    if not to_remove:
        return (0, 0)
    removed_ec = 0
    for ec in list(model.entcomp_list):
        sup = (text(jfind(ec, "j:UnitSuperior")) or "").strip()
        if sup and sup in to_remove:
            model.entcomp_list.remove(ec)
            removed_ec += 1
    removed_units = 0
    for unit in list(model.unit_list):
        lvc = (text(jfind(unit, "j:LvcId")) or "").strip()
        if lvc and lvc in to_remove:
            model.unit_list.remove(unit)
            removed_units += 1
    return removed_units, removed_ec


def remove_units_by_branch(model: ObsModel, branches: Set[str]) -> Tuple[int, int]:

    id_to_unit, parent_to_children = collect_unit_maps(model)
    to_remove_units: Set[str] = set()

    def mark_subtree(lvcid: str):
        if lvcid in to_remove_units:
            return
        to_remove_units.add(lvcid)
        for child in parent_to_children.get(lvcid, []):
            clid = text(jfind(child, "j:LvcId"))
            if clid:
                mark_subtree(clid)

    for lid, u in id_to_unit.items():
        br = unit_branch(u)
        if br and br in branches:
            mark_subtree(lid)
    removed_ec = 0
    for ec in list(model.entcomp_list):
        sup = text(jfind(ec, "j:UnitSuperior"))
        if sup and sup in to_remove_units:
            model.entcomp_list.remove(ec)
            removed_ec += 1
    removed_units = 0
    for u in list(model.unit_list):
        lid = text(jfind(u, "j:LvcId"))
        if lid and lid in to_remove_units:
            model.unit_list.remove(u)
            removed_units += 1
    return removed_units, removed_ec


def remove_side(model: ObsModel, side_name: str) -> Tuple[int, int]:

    id_to_unit, parent_to_children = collect_unit_maps(model)
    roots: Set[str] = set()
    for lid, u in id_to_unit.items():
        ss = text(jfind(u, "j:SideSuperior"))
        if ss and ss.strip().upper() == side_name.upper():
            roots.add(lid)
    to_remove: Set[str] = set()

    def mark(lid: str):
        if not lid or lid in to_remove:
            return
        to_remove.add(lid)
        for ch in parent_to_children.get(lid, []):
            clid = text(jfind(ch, "j:LvcId"))
            if clid:
                mark(clid)

    for r in roots:
        mark(r)
    removed_ec = 0
    for ec in list(model.entcomp_list):
        sup = text(jfind(ec, "j:UnitSuperior"))
        if sup and sup in to_remove:
            model.entcomp_list.remove(ec)
            removed_ec += 1
    removed_units = 0
    for u in list(model.unit_list):
        lid = text(jfind(u, "j:LvcId"))
        if lid and lid in to_remove:
            model.unit_list.remove(u)
            removed_units += 1
    return removed_units, removed_ec


# Consolidation + LvcId helpers


def consolidate_units(
    model: ObsModel,
    target_label: str,
    sides: Optional[Set[str]] = None,
    scope_roots: Optional[Set[str]] = None,
) -> Dict[str, int]:

    option = (target_label or "").strip()
    stats: Dict[str, int] = {
        "removed_units": 0,
        "moved_entity_compositions": 0,
        "relinked_units": 0,
        "skipped_units": 0,
    }
    if not option or option == "None":
        return stats
    target_echelon = normalize_echelon(option) or option
    target_rank = ECHELON_RANK.get(target_echelon, len(ECHELON_ORDER))
    protected_rank = ECHELON_RANK.get(
        "Company/Battery", ECHELON_RANK.get("Company", target_rank)
    )
    side_filter = {s.upper() for s in sides} if sides else None
    id_to_unit, parent_to_children = collect_unit_maps(model)
    if not id_to_unit:
        return stats
    root_ids: Set[str] = set()
    if scope_roots:
        for root in scope_roots:
            if root:
                root_ids.add(root.strip())
    scope_ids: Optional[Set[str]] = None
    if root_ids:
        scope_ids = set()
        to_visit = list(root_ids)
        while to_visit:
            current = (to_visit.pop() or "").strip()
            if not current or current in scope_ids:
                continue
            unit_elem = id_to_unit.get(current)
            if unit_elem is None:
                continue
            scope_ids.add(current)
            for child in parent_to_children.get(current, []):
                child_id = (text(jfind(child, "j:LvcId")) or "").strip()
                if child_id and child_id not in scope_ids:
                    to_visit.append(child_id)
        if not scope_ids:
            return stats
        root_ids = {rid for rid in root_ids if rid in scope_ids}
    parent_map: Dict[str, Optional[str]] = {}
    for lvc, unit in id_to_unit.items():
        sup = text(jfind(unit, "j:UnitSuperior"))
        parent_map[lvc] = sup.strip() if sup and sup.strip() else None
    side_cache: Dict[str, str] = {}

    def unit_side_value(lvc: Optional[str]) -> str:
        if not lvc:
            return "UNKNOWN"
        if lvc in side_cache:
            return side_cache[lvc]
        unit = id_to_unit.get(lvc)
        if unit is None:
            side_cache[lvc] = "UNKNOWN"
            return "UNKNOWN"
        side = (text(jfind(unit, "j:SideSuperior")) or "").strip()
        if side:
            side_cache[lvc] = side.upper()
            return side_cache[lvc]
        parent = parent_map.get(lvc)
        resolved = unit_side_value(parent)
        side_cache[lvc] = resolved
        return resolved

    kept_units: Set[str] = set()
    explicit_kept: Set[str] = set()
    to_remove: Set[str] = set()
    for lvc, unit in id_to_unit.items():
        if scope_ids is not None and lvc not in scope_ids:
            continue
        rank = ECHELON_RANK.get(unit_echelon(unit), len(ECHELON_ORDER))
        if rank > target_rank:
            continue
        if root_ids and lvc in root_ids:
            explicit_kept.add(lvc)
            continue
        unit_side = unit_side_value(lvc)
        if side_filter and unit_side not in side_filter:
            continue
        if rank >= protected_rank and is_hq_unit(unit):
            kept_units.add(lvc)
            continue
        to_remove.add(lvc)
    if not to_remove:
        stats["skipped_units"] = len(kept_units) + len(explicit_kept)
        return stats
    hq_by_parent: Dict[Optional[str], List[str]] = defaultdict(list)
    for hq in kept_units:
        parent = parent_map.get(hq)
        hq_by_parent[parent].append(hq)
    for parent, items in hq_by_parent.items():
        items.sort(key=lambda l: (text(jfind(id_to_unit[l], "j:Name")) or "").lower())
    ec_by_superior: Dict[str, List[ET._Element]] = defaultdict(list)
    for ec in list(model.entcomp_list):
        sup = text(jfind(ec, "j:UnitSuperior"))
        if sup and sup.strip():
            ec_by_superior[sup.strip()].append(ec)
    ROOT_PARENT = "__ROOT__"
    used_names: Dict[str, Set[str]] = defaultdict(set)

    def parent_key(pid: Optional[str]) -> str:
        return pid if pid else ROOT_PARENT

    for lvc, unit in id_to_unit.items():
        if lvc in to_remove:
            continue
        parent = parent_map.get(lvc)
        if parent in to_remove:
            continue
        name_val = (text(jfind(unit, "j:Name")) or "").strip()
        if name_val:
            used_names[parent_key(parent)].add(name_val.casefold())
    depth_cache: Dict[str, int] = {}

    def depth_of(lvc: str) -> int:
        if lvc in depth_cache:
            return depth_cache[lvc]
        depth = 0
        cur = parent_map.get(lvc)
        seen: Set[str] = {lvc}
        while cur:
            depth += 1
            if cur in depth_cache:
                depth += depth_cache[cur]
                break
            if cur in seen:
                break
            seen.add(cur)
            cur = parent_map.get(cur)
        depth_cache[lvc] = depth
        return depth

    def preferred_rollup_parent(lvc: str) -> Optional[str]:
        cur = parent_map.get(lvc)
        seen: Set[str] = set()
        while cur:
            if cur in seen:
                break
            if cur not in to_remove:
                hqs = hq_by_parent.get(cur)
                if hqs:
                    return hqs[0]
                return cur
            seen.add(cur)
            cur = parent_map.get(cur)
        return None

    target_parent_for: Dict[str, Optional[str]] = {}
    skipped: Set[str] = set()
    for lvc in list(to_remove):
        target_parent = preferred_rollup_parent(lvc)
        if target_parent is None:
            skipped.add(lvc)
        else:
            target_parent_for[lvc] = target_parent
    if skipped:
        to_remove.difference_update(skipped)
    stats["skipped_units"] = len(kept_units) + len(skipped) + len(explicit_kept)
    if not to_remove:
        return stats

    def ensure_unique_name(
        node: ET._Element, parent_id: Optional[str], source_label: str
    ) -> None:
        key = parent_key(parent_id)
        used = used_names.setdefault(key, set())
        name_elem = jfind(node, "j:Name")
        base = (text(name_elem) or "").strip()
        if not base:
            base = (source_label or "").strip() or "Unit"
        candidate = base
        prefix = (source_label or "").strip()
        if candidate.casefold() in used:
            candidate = f"{prefix} - {base}" if prefix else base
            idx = 2
            while candidate.casefold() in used:
                suffix = f" ({idx})"
                candidate = (
                    f"{prefix} - {base}{suffix}" if prefix else f"{base}{suffix}"
                )
                idx += 1
        if name_elem is None:
            name_elem = ET.Element(jtag("Name"))
            node.insert(0, name_elem)
        name_elem.text = candidate
        used.add(candidate.casefold())

    sorted_targets = sorted(to_remove, key=depth_of, reverse=True)
    for lvc in sorted_targets:
        unit_elem = id_to_unit.get(lvc)
        if unit_elem is None:
            continue
        target_parent = target_parent_for.get(lvc)
        children = list(parent_to_children.get(lvc, []))
        source_label = (text(jfind(unit_elem, "j:Name")) or "").strip()
        for child in children:
            child_id = text(jfind(child, "j:LvcId"))
            if not child_id:
                continue
            existing = parent_to_children.get(lvc)
            if existing and child in existing:
                existing.remove(child)
            if target_parent:
                parent_to_children.setdefault(target_parent, []).append(child)
            sup_elem = jfind(child, "j:UnitSuperior")
            if target_parent:
                if sup_elem is None:
                    sup_elem = ET.Element(jtag("UnitSuperior"))
                    child.insert(0, sup_elem)
                sup_elem.text = target_parent
            elif sup_elem is not None:
                child.remove(sup_elem)
            parent_map[child_id] = target_parent
            if child_id not in to_remove:
                ensure_unique_name(child, target_parent, source_label)
                stats["relinked_units"] += 1
        for ec in ec_by_superior.get(lvc, []):
            sup_elem = jfind(ec, "j:UnitSuperior")
            if sup_elem is None:
                sup_elem = ET.Element(jtag("UnitSuperior"))
                ec.insert(0, sup_elem)
            sup_elem.text = target_parent or ""
            stats["moved_entity_compositions"] += 1
            if target_parent:
                ec_by_superior[target_parent].append(ec)
        parent_to_children.pop(lvc, None)
        try:
            model.unit_list.remove(unit_elem)
        except ValueError:
            pass
        id_to_unit.pop(lvc, None)
        parent_map.pop(lvc, None)
        stats["removed_units"] += 1
    return stats


def autofill_unique_lvcids(model: ObsModel) -> int:

    used: Set[str] = set()

    def record(elem):
        lid_elem = jfind(elem, "j:LvcId")
        if lid_elem is not None and lid_elem.text:
            used.add(lid_elem.text)

    for u in list(model.unit_list):
        record(u)
    for ec in list(model.entcomp_list):
        record(ec)
    counts: Dict[str, int] = {}
    for u in list(model.unit_list):
        lid = text(jfind(u, "j:LvcId"))
        if lid:
            counts[lid] = counts.get(lid, 0) + 1
    for ec in list(model.entcomp_list):
        lid = text(jfind(ec, "j:LvcId"))
        if lid:
            counts[lid] = counts.get(lid, 0) + 1
    used = set([k for k, v in counts.items() if v == 1])

    def new_id(prefix="U", length=10):
        while True:
            cand = prefix + "".join(
                random.choices(string.ascii_uppercase + string.digits, k=length)
            )
            if cand not in used:
                used.add(cand)
                return cand

    changed = 0
    for u in list(model.unit_list):
        lid_elem = jfind(u, "j:LvcId")
        if lid_elem is None or not lid_elem.text or counts.get(lid_elem.text, 0) > 1:
            new = new_id("U")
            if lid_elem is None:
                lid_elem = ET.Element(jtag("LvcId"))
                u.insert(0, lid_elem)
            lid_elem.text = new
            changed += 1
    for ec in list(model.entcomp_list):
        lid_elem = jfind(ec, "j:LvcId")
        if lid_elem is None or not lid_elem.text or counts.get(lid_elem.text, 0) > 1:
            new = new_id("E")
            if lid_elem is None:
                lid_elem = ET.Element(jtag("LvcId"))
                ec.insert(0, lid_elem)
            lid_elem.text = new
            changed += 1
    return changed


# ---------------- DRAGON Import helpers ----------------


def _expand_echelon(e: Optional[str]) -> Optional[str]:

    if not e:
        return e
    m = {
        "BN": "Battalion",
        "BATTALION": "Battalion",
        "CO": "Company",
        "COMPANY": "Company",
        "BDE": "Brigade",
        "BRIGADE": "Brigade",
        "PLT": "Platoon",
        "PLATOON": "Platoon",
        "DIV": "Division",
        "DIVISION": "Division",
        "CORPS": "Corps",
        "ARMY": "Army",
    }
    key = str(e).strip().upper()
    return m.get(key, e)


def _sanitize_unit_name(name: Optional[str]) -> Optional[str]:

    if not name:
        return name
    return str(name).replace(" ", "_")


def _gen_unique_lvcid(prefix: str, used: Set[str], length: int = 10) -> str:

    while True:
        cand = prefix + "".join(
            random.choices(string.ascii_uppercase + string.digits, k=length)
        )
        if cand not in used:
            used.add(cand)
            return cand


def _parse_dis_enum_to_attrs(enum_str: str) -> Dict[str, str]:

    parts = [p.strip() for p in str(enum_str).split(".")]
    keys = ["kind", "domain", "country", "category", "subcategory", "specific", "extra"]
    attrs = {}
    for k, v in zip(keys, parts):
        if v != "":
            attrs[k] = v
    return attrs


def _parse_unit_dis_enum_to_attrs(enum_str: str) -> Dict[str, str]:

    text_val = (enum_str or "").strip()
    if not text_val:
        raise ValueError("Enter a UnitDisEnumeration value.")
    parts = [part.strip() for part in text_val.split(".")]
    if len(parts) != len(_UNIT_DIS_ENUM_ATTRS):
        raise ValueError(
            "UnitDisEnumeration must contain 7 dot-separated integers (kind.domain.country.echelon.type.subtype.modifier)."
        )
    attrs: Dict[str, str] = {}
    for key, value in zip(_UNIT_DIS_ENUM_ATTRS, parts):
        if not value or not value.isdigit():
            raise ValueError(
                f"UnitDisEnumeration component '{value or ''}' for '{key}' must be numeric."
            )
        attrs[key] = str(int(value))
    return attrs


def _format_unit_dis_enum(attrs: Mapping[str, str]) -> str:

    parts: List[str] = []
    for key in _UNIT_DIS_ENUM_ATTRS:
        raw = attrs.get(key, "")
        norm = _normalize_discode_component(raw) if raw else ""
        parts.append(norm)
    while parts and not parts[-1]:
        parts.pop()
    return ".".join(parts)


def import_dragon_xlsx_into_model(
    xlsx_path: str,
    model: Optional[ObsModel],
    unitlist_map: Optional[Dict[str, Dict[str, str]]] = None,
) -> ObsModel:

    import pandas as pd, math

    if model is None:
        model = _new_blank_model()
    # === Units from UNIT INFO ===
    try:
        unit_df = pd.read_excel(xlsx_path, sheet_name="UNIT INFO")
    except Exception as e:
        raise ValueError(f"Failed to read 'UNIT INFO' sheet: {e}")

    def val(x):
        if isinstance(x, float) and math.isnan(x):
            return None
        s = str(x).strip()
        return s if s else None

    used_ids: Set[str] = set()
    used_unit_ids: Set[str] = set()
    for u in list(model.unit_list):
        lid = text(jfind(u, "j:LvcId"))
        if lid:
            used_ids.add(lid)
    cols = {str(c).strip().upper(): c for c in unit_df.columns}
    cname_col = cols.get("NAME") or cols.get("UNIT NAME") or cols.get("TITLE")
    uclass_col = cols.get("UNIT CLASS") or cols.get("CLASS") or cols.get("UNITCLASS")
    echelon_col = cols.get("ECHELON")
    ms2525_col = cols.get("2525C") or cols.get("MILSTD2525C") or cols.get("2525")
    uid_col = cols.get("UID") or cols.get("UNIT UID") or cols.get("UNIT_ID")
    parent_uid_col = (
        cols.get("PARENT UID") or cols.get("PARENT_UID") or cols.get("PARENT")
    )
    side_col = cols.get("SIDE") or cols.get("SIDE SUPERIOR") or cols.get("SIDESUPERIOR")
    uid_to_lvc: Dict[str, str] = {}
    created_units: Dict[str, ET._Element] = {}
    for _, row in unit_df.iterrows():
        uid = val(row.get(uid_col)) if uid_col else None
        name = val(row.get(cname_col)) if cname_col else None
        if not name:
            continue
        lvc = _gen_unique_lvcid("U", used_ids)
        u = ET.SubElement(model.unit_list, jtag("Unit"))
        u.set("id", _gen_unit_xml_id(used_unit_ids))
        u.set("dateTimeStamp", _now_iso_seconds_no_tz())
        ET.SubElement(u, jtag("LvcId")).text = lvc
        ET.SubElement(u, jtag("Name")).text = _sanitize_unit_name(name)
        if side_col:
            sv = val(row.get(side_col))
            if sv:
                ET.SubElement(u, jtag("SideSuperior")).text = sv
        if uclass_col:
            clsname = val(row.get(uclass_col))
            if clsname:
                ET.SubElement(u, jtag("ClassName")).text = clsname
        if echelon_col:
            ech = _expand_echelon(val(row.get(echelon_col)))
            if ech:
                ET.SubElement(u, jtag("Echelon")).text = ech
        if ms2525_col:
            code2525 = val(row.get(ms2525_col))
            if code2525:
                ET.SubElement(u, jtag("MilStd2525CCode")).text = code2525
                # Look up and inject UnitDisEnumeration from reference map
                if unitlist_map is not None:
                    attrs = unitlist_map.get(code2525) or unitlist_map.get(
                        code2525.upper()
                    )
                    if attrs:
                        ude = ET.SubElement(u, jtag("UnitDisEnumeration"))
                        for k, v in attrs.items():
                            ude.set(k, str(v))
        own = ET.SubElement(u, jtag("OwningFederate"))
        ET.SubElement(own, jtag("Federate")).text = "WARSIM"
        ET.SubElement(u, jtag("IsInactive")).text = "false"
        ET.SubElement(u, jtag("Strength")).text = "100"
        ET.SubElement(u, jtag("IsAggCommandable")).text = "false"
        if uid:
            uid_to_lvc[uid] = lvc
        created_units[lvc] = u
    if parent_uid_col and uid_col:
        for _, row in unit_df.iterrows():
            uid = val(row.get(uid_col))
            parent_uid = val(row.get(parent_uid_col))
            if not uid:
                continue
            child_lvc = uid_to_lvc.get(uid)
            if not child_lvc:
                continue
            child_unit = created_units.get(child_lvc)
            if child_unit is None:
                continue
            if parent_uid and parent_uid in uid_to_lvc:
                ET.SubElement(child_unit, jtag("UnitSuperior")).text = uid_to_lvc[
                    parent_uid
                ]
    # === EntityCompositions from other sheets ===
    xls = pd.ExcelFile(xlsx_path)
    for sheet in xls.sheet_names:
        if sheet.strip().upper() == "UNIT INFO":
            continue
        try:
            df = pd.read_excel(xlsx_path, sheet_name=sheet)
        except Exception:
            continue
        cols2 = {str(c).strip().upper(): c for c in df.columns}
        uidc = cols2.get("UID") or cols2.get("UNIT UID") or cols2.get("UNIT_ID")
        if not uidc:
            continue
        for _, row in df.iterrows():
            uid = val(row.get(uidc))
            if not uid:
                continue
            unit_lvc = uid_to_lvc.get(uid)
            if not unit_lvc:
                continue
            # Determine personnel
            is_person = False
            ptyp_col = cols2.get("PERSONNEL TYPE") or cols2.get("PERSONNEL TYPE CODE")
            if ptyp_col:
                ptyp = val(row.get(ptyp_col))
                is_person = bool(ptyp)
            ec = ET.SubElement(model.entcomp_list, jtag("EntityComposition"))
            ET.SubElement(ec, jtag("LvcId")).text = _gen_unique_lvcid("E", used_ids)
            ET.SubElement(ec, jtag("UnitSuperior")).text = unit_lvc
            ecc = ET.SubElement(ec, jtag("EntityCompositionClass"))
            cname = None
            if is_person and ptyp_col:
                cname = val(row.get(ptyp_col))
            else:
                eq_col = (
                    cols2.get("EQUIPMENT TYPE")
                    or cols2.get("EQUIP TYPE")
                    or cols2.get("TYPE")
                )
                cname = val(row.get(eq_col)) if eq_col else None
            if cname:
                ET.SubElement(ecc, jtag("ClassName")).text = cname
            dnode = ET.SubElement(ecc, jtag("DisCode"))
            den_col = cols2.get("DIS ENUMERATION") or cols2.get("DIS ENUM")
            den = val(row.get(den_col)) if den_col else None
            if den:
                attrs = _parse_dis_enum_to_attrs(den)
                for k, v in attrs.items():
                    dnode.set(k, str(v))
                if is_person and "kind" not in attrs:
                    dnode.set("kind", "3")
            else:
                dnode.set("kind", "3" if is_person else "1")
    return model


# ---------------- ClassData builders (from master spreadsheets) ----------------


def build_entitycomposition_classlist_from_xlsx(
    model: ObsModel, ecc_xlsx_path: str
) -> int:

    try:
        df = pd.read_excel(ecc_xlsx_path)
    except Exception as e:
        raise ValueError(f"Failed to read ECC master: {e}")
    cols = {c.strip().upper(): c for c in df.columns}
    plat_col = cols.get("PLATFORM") or list(df.columns)[0]
    dis_col = (
        cols.get("DIS ENUMERATION") or cols.get("DIS_ENUMERATION") or cols.get("DIS")
    )
    lst = jfind(model.classdata, "j:EntityCompositionClassList")
    if lst is None:
        lst = ET.SubElement(model.classdata, jtag("EntityCompositionClassList"))
    existing: Dict[str, List[ET._Element]] = {}
    for child in list(lst):
        name_text = text(jfind(child, "j:Name"))
        if not name_text:
            continue
        existing.setdefault(name_text.strip().upper(), []).append(child)
    new_entries: Dict[str, ET._Element] = {}
    tname_col = cols.get("TYPE NAME") or cols.get("TYPENAME") or None
    for _, row in df.iterrows():
        raw_name = row.get(plat_col)
        if raw_name is None or pd.isna(raw_name):
            continue
        name = str(raw_name).strip()
        if not name:
            continue
        raw_enum = row.get(dis_col)
        if raw_enum is None or pd.isna(raw_enum):
            continue
        enum = str(raw_enum).strip()
        if not enum:
            continue
        key = name.upper()
        prior = existing.pop(key, None)
        if prior:
            for elem in prior:
                lst.remove(elem)
        replacing = new_entries.pop(key, None)
        if replacing is not None:
            lst.remove(replacing)
        ecc = ET.SubElement(lst, jtag("EntityCompositionClass"))
        new_entries[key] = ecc
        ET.SubElement(ecc, jtag("Name")).text = name
        attrs = _parse_dis_enum_to_attrs(enum)
        dc = ET.SubElement(ecc, jtag("DisCode"))
        for k, v in attrs.items():
            dc.set(k, str(v))
        ET.SubElement(ecc, jtag("IsBaseClass")).text = "false"
        if tname_col:
            raw_tname = row.get(tname_col)
            if raw_tname is not None and pd.isna(raw_tname):
                raw_tname = None
            tname_val = str(raw_tname).strip() if raw_tname is not None else ""
            if tname_val:
                alias_list = ET.SubElement(ecc, jtag("AliasList"))
                alias = ET.SubElement(alias_list, jtag("Alias"))
                ET.SubElement(alias, jtag("Federate")).text = "WARSIM"
                ET.SubElement(alias, jtag("TypeName")).text = tname_val
    return len(new_entries)


def build_unitclass_list_from_xlsx(model: ObsModel, ucl_xlsx_path: str) -> int:

    try:
        df = pd.read_excel(ucl_xlsx_path)
    except Exception as e:
        raise ValueError(f"Failed to read UnitClass master: {e}")
    cols = {c.strip().upper(): c for c in df.columns}
    uclass_col = cols.get("UNIT CLASS") or list(df.columns)[0]
    tname_col = cols.get("TYPE NAME")
    code_col = cols.get("2525C") or cols.get("MILSTD2525C") or cols.get("MILSTD 2525C")
    lst = jfind(model.classdata, "j:UnitClassList")
    if lst is None:
        lst = ET.SubElement(model.classdata, jtag("UnitClassList"))
    existing: Dict[str, List[ET._Element]] = {}
    for child in list(lst):
        agg = text(jfind(child, "j:AggregateName"))
        if not agg:
            continue
        existing.setdefault(agg.strip().upper(), []).append(child)
    new_entries: Dict[str, ET._Element] = {}
    den_col = cols.get("DIS ENUMERATION") or cols.get("DIS_ENUMERATION")
    for _, row in df.iterrows():
        raw_name = row.get(uclass_col)
        if raw_name is None or pd.isna(raw_name):
            continue
        cname = str(raw_name).strip()
        if not cname:
            continue
        raw_tname = row.get(tname_col) if tname_col else None
        if raw_tname is not None and pd.isna(raw_tname):
            raw_tname = None
        tname = str(raw_tname).strip() if raw_tname is not None else ""
        raw_code = row.get(code_col) if code_col else None
        if raw_code is not None and pd.isna(raw_code):
            raw_code = None
        code2525 = str(raw_code).strip() if raw_code is not None else ""
        key = cname.upper()
        prior = existing.pop(key, None)
        if prior:
            for elem in prior:
                lst.remove(elem)
        replacing = new_entries.pop(key, None)
        if replacing is not None:
            lst.remove(replacing)
        ucl = ET.SubElement(lst, jtag("UnitClass"))
        new_entries[key] = ucl
        ET.SubElement(ucl, jtag("AggregateName")).text = cname
        if code2525:
            ET.SubElement(ucl, jtag("MilStd2525CCode")).text = code2525
        if den_col:
            raw_enum = row.get(den_col)
            if raw_enum is not None and pd.isna(raw_enum):
                raw_enum = None
            enum_val = str(raw_enum).strip() if raw_enum is not None else ""
            if enum_val:
                attrs = _parse_unit_dis_enum_to_attrs(enum_val)
                ude = ET.SubElement(ucl, jtag("UnitDisEnumeration"))
                for k, v in attrs.items():
                    ude.set(k, str(v))
        if tname:
            alias_list = ET.SubElement(ucl, jtag("AliasList"))
            alias = ET.SubElement(alias_list, jtag("Alias"))
            ET.SubElement(alias, jtag("Federate")).text = "WARSIM"
            ET.SubElement(alias, jtag("TypeName")).text = tname
    return len(new_entries)


class AnalyzerTab(QWidget):

    modelLoaded = Signal(object)

    def __init__(self, parent):
        super().__init__(parent)
        self.model: Optional[ObsModel] = None
        self._build()

    def _build(self):
        lay = QVBoxLayout(self)
        src_box = QHBoxLayout()
        self.path_edit = QLineEdit()
        self.path_edit.setPlaceholderText("Select a JTDS OBS .xml or .zip ...")
        btn_browse = QPushButton("Open...")
        btn_browse.clicked.connect(self.on_browse)
        src_box.addWidget(self.path_edit)
        src_box.addWidget(btn_browse)
        lay.addLayout(src_box)
        self.summary = QTextEdit()
        self.summary.setReadOnly(True)
        self.summary.setMinimumHeight(200)
        lay.addWidget(self.summary)
        self.btn_analyze = QPushButton("Analyze")
        self.btn_analyze.clicked.connect(self.on_analyze)
        self.btn_analyze.setEnabled(False)
        lay.addWidget(self.btn_analyze, alignment=Qt.AlignmentFlag.AlignRight)

    def set_model(self, model: Optional[ObsModel]):
        self.model = model
        self.btn_analyze.setEnabled(model is not None)

    def on_browse(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Open JTDS OBS", "", "OBS XML or ZIP (*.xml *.zip);;All Files (*)"
        )
        if path:
            self.path_edit.setText(path)
            try:
                model = load_obs_xml(path)
                self.set_model(model)
                self.summary.setPlainText(
                    "Loaded. Click Analyze to compute statistics."
                )
                self.modelLoaded.emit(model)
            except Exception as e:
                traceback.print_exc()
                QMessageBox.critical(self, "Open Error", str(e))

    def on_analyze(self):
        if not self.model:
            return
        m = self.model
        units = len(list(m.unit_list))
        total_ecs = len(list(m.entcomp_list))
        pers = count_personnel(m)
        equip = total_ecs - pers
        # Branch presence
        id_to_unit, _ = collect_unit_maps(m)
        found = {b: 0 for b in BRANCH_TOKENS}
        for u in id_to_unit.values():
            b = unit_branch(u)
            if b:
                found[b] += 1
        lines = []
        lines.append(f"Units: {units:,}")
        lines.append(f"Entity Compositions (Equipment): {equip:,}")
        lines.append(f"Personnel (DisCode kind=3): {pers:,}")
        lines.append("")
        lines.append("Parent unit branches present:")
        for key in ["SIG", "MNT", "ENG", "MED", "SPT", "CHEM"]:
            lines.append(f"  {key}: {'Yes' if found[key] else 'No'} ({found[key]})")
        self.summary.setPlainText("\\n".join(lines))


class ScrubberTab(QWidget):

    def __init__(self, parent):
        super().__init__(parent)
        self.model: Optional[ObsModel] = None
        self._build()

    def _build(self):
        lay = QVBoxLayout(self)
        form = QGridLayout()
        self.chk_supply = QCheckBox("Remove <SupplyClassList>")
        self.chk_emitter = QCheckBox("Remove <EmitterClassList>")
        self.chk_munit = QCheckBox("Remove <MunitionsList>")
        self.chk_fly = QCheckBox("Remove <FlyOutList>")
        self.chk_personnel = QCheckBox("Remove Personnel (DisCode kind=3)")
        self.chk_munitions_kind2 = QCheckBox("Remove Munitions (DisCode kind=2)")
        self.chk_remove_unused_ecc = QCheckBox(
            "Remove unused EntityCompositionClass entries"
        )
        self.chk_report_missing_ecc = QCheckBox(
            "Report EntityComposition ClassName mismatches"
        )
        row = 0
        toggles = [
            self.chk_supply,
            self.chk_emitter,
            self.chk_munit,
            self.chk_fly,
            self.chk_personnel,
            self.chk_munitions_kind2,
            self.chk_remove_unused_ecc,
            self.chk_report_missing_ecc,
        ]
        for w in toggles:
            form.addWidget(w, row, 0)
            row += 1
        # Branch group
        grp = QGroupBox("Remove Units by Branch (and their equipment/personnel)")
        grp_l = QHBoxLayout(grp)
        self.chk_SIG = QCheckBox("SIG")
        self.chk_MNT = QCheckBox("MNT")
        self.chk_ENG = QCheckBox("ENG")
        self.chk_MED = QCheckBox("MED")
        self.chk_SPT = QCheckBox("SPT")
        self.chk_CHEM = QCheckBox("CHEM")
        for w in [
            self.chk_SIG,
            self.chk_MNT,
            self.chk_ENG,
            self.chk_MED,
            self.chk_SPT,
            self.chk_CHEM,
        ]:
            grp_l.addWidget(w)
        lay.addLayout(form)
        lay.addWidget(grp)
        # Side removal
        side_box = QHBoxLayout()
        self.chk_side = QCheckBox("Remove side (units + ECs)")
        self.cmb_side = QComboBox()
        self.cmb_side.addItems(["OPFOR", "BLUFOR", "NEUTRAL", "CIVILIAN", "UNKNOWN"])
        side_box.addWidget(self.chk_side)
        side_box.addWidget(self.cmb_side)
        lay.addLayout(side_box)
        discode_box = QHBoxLayout()
        discode_box.addWidget(QLabel("Remove by DisCode:"))
        self.discode_edit = QLineEdit()
        self.discode_edit.setPlaceholderText("e.g. 1.1.225.28.1.0.0")
        self.discode_edit.setEnabled(False)
        discode_box.addWidget(self.discode_edit)
        self.btn_remove_discode = QPushButton("Remove")
        self.btn_remove_discode.setEnabled(False)
        self.btn_remove_discode.clicked.connect(self.on_remove_discode)
        discode_box.addWidget(self.btn_remove_discode)
        lay.addLayout(discode_box)
        replace_box = QHBoxLayout()
        replace_box.addWidget(QLabel("Replace DisCode:"))
        self.discode_replace_from_edit = QLineEdit()
        self.discode_replace_from_edit.setPlaceholderText(
            "Current e.g. 1.1.225.28.1.0.0"
        )
        self.discode_replace_from_edit.setEnabled(False)
        replace_box.addWidget(self.discode_replace_from_edit)
        replace_box.addWidget(QLabel("to"))
        self.discode_replace_to_edit = QLineEdit()
        self.discode_replace_to_edit.setPlaceholderText("New e.g. 1.1.225.28.2.0.0")
        self.discode_replace_to_edit.setEnabled(False)
        replace_box.addWidget(self.discode_replace_to_edit)
        self.btn_replace_discode = QPushButton("Replace")
        self.btn_replace_discode.setEnabled(False)
        self.btn_replace_discode.clicked.connect(self.on_replace_discode)
        replace_box.addWidget(self.btn_replace_discode)
        lay.addLayout(replace_box)
        self.btn_apply = QPushButton("Apply Scrub")
        self.btn_apply.clicked.connect(self.on_apply)
        self.btn_apply.setEnabled(False)
        lay.addWidget(self.btn_apply, alignment=Qt.AlignmentFlag.AlignRight)
        self.out = QTextEdit()
        self.out.setReadOnly(True)
        self.out.setMinimumHeight(180)
        lay.addWidget(self.out)

    def set_model(self, model: Optional[ObsModel]):
        self.model = model
        enabled = model is not None
        self.btn_apply.setEnabled(enabled)
        self.discode_edit.setEnabled(enabled)
        self.btn_remove_discode.setEnabled(enabled)
        self.discode_replace_from_edit.setEnabled(enabled)
        self.discode_replace_to_edit.setEnabled(enabled)
        self.btn_replace_discode.setEnabled(enabled)
        if not enabled:
            self.discode_edit.clear()
            self.discode_replace_from_edit.clear()
            self.discode_replace_to_edit.clear()
            return
        sides = set()
        for u in list(model.unit_list):
            ss = text(jfind(u, "j:SideSuperior"))
            if ss:
                sides.add(ss.strip())
        for s in sorted(sides):
            if self.cmb_side.findText(s) == -1:
                self.cmb_side.addItem(s)

    def on_apply(self):
        if not self.model:
            return
        m = self.model
        report = []
        counts = purge_classdata_sections(
            m,
            supply=self.chk_supply.isChecked(),
            emitter=self.chk_emitter.isChecked(),
            munitions=self.chk_munit.isChecked(),
            flyout=self.chk_fly.isChecked(),
        )
        for k, v in counts.items():
            report.append(f"Removed {k}: {v} entries")
        if self.chk_personnel.isChecked():
            rem_ec, rem_ecc = remove_personnel(m)
            report.append(
                f"Removed Personnel (kind=3): ECs={rem_ec}, Classes={rem_ecc}"
            )
        if (
            getattr(self, "chk_munitions_kind2", None)
            and self.chk_munitions_kind2.isChecked()
        ):
            rem_ec, rem_ecc = remove_munitions(m)
            report.append(
                f"Removed Munitions (kind=2): ECs={rem_ec}, Classes={rem_ecc}"
            )
        branches = set()
        for key, chk in [
            ("SIG", self.chk_SIG),
            ("MNT", self.chk_MNT),
            ("ENG", self.chk_ENG),
            ("MED", self.chk_MED),
            ("SPT", self.chk_SPT),
            ("CHEM", self.chk_CHEM),
        ]:
            if chk.isChecked():
                branches.add(key)
        if branches:
            ru, re = remove_units_by_branch(m, branches)
            report.append(
                f"Removed Units by Branch {sorted(list(branches))}: Units={ru}, EntityCompositions={re}"
            )
        if self.chk_side.isChecked():
            side = self.cmb_side.currentText() or "OPFOR"
            ru2, re2 = remove_side(m, side)
            report.append(f"Removed side {side}: Units={ru2}, EntityCompositions={re2}")
        if (
            getattr(self, "chk_remove_unused_ecc", None)
            and self.chk_remove_unused_ecc.isChecked()
        ):
            removed_ecc, removed_labels, unnamed_removed = (
                remove_unused_entity_composition_classes(m)
            )
            if removed_ecc:
                details: List[str] = []
                if removed_labels:
                    preview = ", ".join(removed_labels[:10])
                    if len(removed_labels) > 10:
                        preview += f", +{len(removed_labels) - 10} more"
                    details.append(f"labels: {preview}")
                if unnamed_removed:
                    details.append(f"{unnamed_removed} unnamed")
                suffix = f" ({', '.join(details)})" if details else ""
                report.append(
                    f"Removed unused EntityCompositionClass entries: {removed_ecc}{suffix}"
                )
            else:
                report.append("No unused EntityCompositionClass entries detected.")
        if (
            getattr(self, "chk_report_missing_ecc", None)
            and self.chk_report_missing_ecc.isChecked()
        ):
            missing = find_missing_entity_composition_classes(m)
            if missing:
                preview = ", ".join(missing[:10])
                if len(missing) > 10:
                    preview += f", +{len(missing) - 10} more"
                report.append(
                    f"Missing {len(missing)} EntityCompositionClass definition(s) for: {preview}"
                )
            else:
                report.append(
                    "All EntityComposition <ClassName> values have matching EntityCompositionClass entries."
                )
        if not report:
            report.append("No changes selected.")
        self.out.setPlainText("\\n".join(report))

    def on_replace_discode(self):
        if not self.model:
            QMessageBox.warning(
                self, "Replace DisCode", "Load a model before replacing entities."
            )
            return
        old = (self.discode_replace_from_edit.text() or "").strip()
        new = (self.discode_replace_to_edit.text() or "").strip()
        if not old:
            QMessageBox.warning(
                self,
                "Replace DisCode",
                "Enter the DisCode value to replace (e.g. 1.1.225.28.1.0.0).",
            )
            self.discode_replace_from_edit.setFocus()
            return
        if not new:
            QMessageBox.warning(self, "Replace DisCode", "Enter the new DisCode value.")
            self.discode_replace_to_edit.setFocus()
            return
        try:
            updated_ecc, updated_entities = replace_discode_entries(
                self.model, old, new
            )
        except ValueError as exc:
            QMessageBox.warning(self, "Replace DisCode", str(exc))
            return
        if updated_ecc == 0 and updated_entities == 0:
            QMessageBox.information(
                self, "Replace DisCode", f"No entries matched DisCode {old}."
            )
            return
        self.out.append(
            f"Replaced DisCode {old} -> {new}: Classes={updated_ecc}, EntityCompositions={updated_entities}"
        )
        QMessageBox.information(
            self,
            "Replace DisCode",
            f"Replaced DisCode {old} -> {new}: Classes={updated_ecc}, EntityCompositions={updated_entities}",
        )
        self.discode_replace_from_edit.selectAll()

    def on_remove_discode(self):
        if not self.model:
            QMessageBox.warning(
                self, "Remove by DisCode", "Load a model before removing entities."
            )
            return
        code = (self.discode_edit.text() or "").strip()
        if not code:
            QMessageBox.warning(
                self,
                "Remove by DisCode",
                "Enter a DisCode value (e.g. 1.1.225.28.1.0.0).",
            )
            self.discode_edit.setFocus()
            return
        try:
            removed_ec, removed_ecc = remove_entities_by_discode(self.model, code)
        except ValueError as exc:
            QMessageBox.warning(self, "Remove by DisCode", str(exc))
            return
        if removed_ec == 0 and removed_ecc == 0:
            QMessageBox.information(
                self, "Remove by DisCode", f"No entities matched DisCode {code}."
            )
            return
        self.out.append(
            f"Removed DisCode {code}: EntityCompositions={removed_ec}, Classes={removed_ecc}"
        )
        QMessageBox.information(
            self,
            "Remove by DisCode",
            f"Removed DisCode {code}: EntityCompositions={removed_ec}, Classes={removed_ecc}",
        )
        self.discode_edit.selectAll()


@dataclass
class _NewUnitData:

    name: str
    class_name: Optional[str]
    milstd_code: Optional[str]
    echelon: Optional[str]
    strength: int
    owning_federate: str
    is_inactive: bool
    is_transferrable: bool
    side_value: Optional[str]
    class_element: Optional[ET._Element]


class CreateUnitDialog(QDialog):

    FEDERATE_OPTIONS = ["WARSIM", "JCATS", "JSAF"]

    def __init__(self, parent: QWidget, model: ObsModel, parent_unit: ET._Element):
        super().__init__(parent)
        self._model = model
        self._parent_unit = parent_unit
        self._result: Optional[_NewUnitData] = None
        self._code_dirty = False
        self._unit_classes = self._collect_unit_classes()
        self._existing_names = self._collect_existing_names()
        self._side_value = self._determine_parent_side()
        self._parent_lvc = (text(jfind(parent_unit, "j:LvcId")) or "").strip()
        self.setWindowTitle("Create Subordinate Unit")
        self.resize(480, 0)
        layout = QVBoxLayout(self)
        context_label = QLabel(
            "Configure the new unit's metadata. LvcId and IDs will be generated on save."
        )
        context_label.setWordWrap(True)
        layout.addWidget(context_label)
        form = QFormLayout()
        parent_label = QLabel(format_unit_label(parent_unit))
        parent_label.setWordWrap(True)
        form.addRow("Parent Unit", parent_label)
        self.unitsuperior_label = QLabel(self._parent_lvc or "(missing LvcId)")
        form.addRow("Unit Superior", self.unitsuperior_label)
        self.side_label = QLabel(self._side_value or "(not set)")
        form.addRow("Side Superior", self.side_label)
        self.name_edit = QLineEdit()
        self.name_edit.setPlaceholderText("Required")
        form.addRow("Name *", self.name_edit)
        self.class_combo = QComboBox()
        self.class_combo.setEditable(True)
        self.class_combo.setInsertPolicy(QComboBox.InsertPolicy.NoInsert)
        line_edit = self.class_combo.lineEdit()
        if line_edit is not None:
            line_edit.setPlaceholderText("Select or type class name")
        self.class_combo.addItem("")
        for entry in self._unit_classes:
            self.class_combo.addItem(entry["name"], entry)
        self.class_combo.currentIndexChanged.connect(self._on_class_changed)
        form.addRow("Class Name", self.class_combo)
        self.code_edit = QLineEdit()
        self.code_edit.setPlaceholderText("MilStd2525C code")
        self.code_edit.textEdited.connect(self._mark_code_dirty)
        form.addRow("MilStd2525C", self.code_edit)
        self.echelon_combo = QComboBox()
        self.echelon_combo.addItem("")
        for echelon in ECHELON_ORDER:
            self.echelon_combo.addItem(echelon)
        parent_echelon = (text(jfind(parent_unit, "j:Echelon")) or "").strip()
        if parent_echelon:
            idx = self.echelon_combo.findText(parent_echelon, Qt.MatchFlag.MatchExactly)
            if idx >= 0:
                self.echelon_combo.setCurrentIndex(idx)
        form.addRow("Echelon", self.echelon_combo)
        self.strength_spin = QSpinBox()
        self.strength_spin.setRange(0, 999999)
        self.strength_spin.setValue(100)
        form.addRow("Strength", self.strength_spin)
        self.federate_combo = QComboBox()
        for entry in self.FEDERATE_OPTIONS:
            self.federate_combo.addItem(entry)
        form.addRow("Owning Federate", self.federate_combo)
        self.inactive_check = QCheckBox("Mark unit as inactive")
        form.addRow("Is Inactive", self.inactive_check)
        self.transfer_check = QCheckBox("Allow transfer between federates")
        self.transfer_check.setChecked(True)
        form.addRow("Is Transferrable", self.transfer_check)
        self.lvc_label = QLabel("Will auto-generate on save")
        form.addRow("LvcId", self.lvc_label)
        layout.addLayout(form)
        buttons = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def _collect_unit_classes(self) -> List[Dict[str, Any]]:
        results: List[Dict[str, Any]] = []
        parent = jfind(self._model.classdata, jtag("UnitClassList"))
        if parent is None:
            return results
        for child in list(parent):
            if _local(child.tag) != "UnitClass":
                continue
            name = (text(jfind(child, "j:AggregateName")) or "").strip()
            if not name:
                continue
            code = (text(jfind(child, "j:MilStd2525CCode")) or "").strip()
            results.append({"name": name, "code": code, "element": child})
        results.sort(key=lambda entry: entry["name"].upper())
        return results

    def _collect_existing_names(self) -> Set[str]:
        names: Set[str] = set()
        for unit in _iter_local(self._model.unit_list, "Unit"):
            nm = (text(jfind(unit, "j:Name")) or "").strip()
            if nm:
                names.add(nm.upper())
        return names

    def _determine_parent_side(self) -> Optional[str]:
        direct = (text(jfind(self._parent_unit, "j:SideSuperior")) or "").strip()
        if direct:
            return direct
        lvc_to_unit, _ = collect_unit_maps(self._model)
        derived = _resolve_unit_side(self._parent_unit, lvc_to_unit)
        return derived

    def _mark_code_dirty(self, _: str) -> None:
        self._code_dirty = True

    def _current_class_entry(self) -> Optional[Dict[str, Any]]:
        index = self.class_combo.currentIndex()
        data = self.class_combo.itemData(index)
        if isinstance(data, dict):
            return data
        text_val = (self.class_combo.currentText() or "").strip().upper()
        for entry in self._unit_classes:
            if entry["name"].upper() == text_val:
                return entry
        return None

    def _on_class_changed(self, index: int) -> None:
        entry = self.class_combo.itemData(index)
        if not isinstance(entry, dict):
            return
        code = entry.get("code")
        if code and (not self._code_dirty or not (self.code_edit.text() or "").strip()):
            self.code_edit.blockSignals(True)
            self.code_edit.setText(code)
            self.code_edit.blockSignals(False)
            self._code_dirty = False

    def result_data(self) -> Optional[_NewUnitData]:
        return self._result

    def accept(self) -> None:
        name = (self.name_edit.text() or "").strip()
        if not name:
            QMessageBox.warning(self, "Create Unit", "Name is required.")
            self.name_edit.setFocus()
            return
        upper_name = name.upper()
        if upper_name in self._existing_names:
            QMessageBox.warning(
                self, "Create Unit", f"A unit named '{name}' already exists."
            )
            self.name_edit.setFocus()
            self.name_edit.selectAll()
            return
        class_text = (self.class_combo.currentText() or "").strip()
        class_entry = self._current_class_entry()
        if not class_text and class_entry:
            class_text = class_entry.get("name", "")
        code = (self.code_edit.text() or "").strip()
        if not code and class_entry and class_entry.get("code"):
            code = class_entry["code"]
        echelon = (self.echelon_combo.currentText() or "").strip()
        if not echelon:
            echelon = None
        owning_federate = (self.federate_combo.currentText() or "WARSIM").strip()
        self._result = _NewUnitData(
            name=name,
            class_name=class_text or None,
            milstd_code=code or None,
            echelon=echelon,
            strength=int(self.strength_spin.value()),
            owning_federate=owning_federate,
            is_inactive=self.inactive_check.isChecked(),
            is_transferrable=self.transfer_check.isChecked(),
            side_value=self._side_value,
            class_element=class_entry.get("element") if class_entry else None,
        )
        super().accept()


@dataclass
class _NewEntityData:

    role: str
    class_name: Optional[str]
    owning_federate: str
    is_inactive: bool
    is_transferrable: bool
    is_invincible: bool
    quantity: int
    batch_count: int


class CreateEntityDialog(QDialog):

    FEDERATE_OPTIONS = ["WARSIM", "JCATS", "JSAF"]

    def __init__(self, parent: QWidget, model: ObsModel, parent_unit: ET._Element):
        super().__init__(parent)
        self._model = model
        self._parent_unit = parent_unit
        self._result: Optional[_NewEntityData] = None
        self._existing_roles = self._collect_existing_roles(parent_unit)
        self._all_class_entries = self._collect_entity_classes()
        self._parent_lvc = (text(jfind(parent_unit, "j:LvcId")) or "").strip()
        self._parent_side = (text(jfind(parent_unit, "j:SideSuperior")) or "").strip()
        self.setWindowTitle("Create Entity Composition")
        layout = QVBoxLayout(self)
        info = QLabel(
            "Define the entity composition metadata. Role must be unique for the parent unit."
        )
        info.setWordWrap(True)
        layout.addWidget(info)
        form = QFormLayout()
        parent_label = QLabel(format_unit_label(parent_unit))
        parent_label.setWordWrap(True)
        form.addRow("Parent Unit", parent_label)
        self.role_edit = QLineEdit()
        self.role_edit.setPlaceholderText("Required and unique within unit")
        form.addRow("Role *", self.role_edit)
        self.class_combo = QComboBox()
        self.class_combo.setEditable(True)
        self.class_combo.setInsertPolicy(QComboBox.InsertPolicy.NoInsert)
        form.addRow("Class Name", self.class_combo)
        self.include_munitions_check = QCheckBox(
            "Include munitions (DisCode kind=2)"
        )
        self.include_munitions_check.setChecked(False)
        self.include_munitions_check.toggled.connect(self._on_include_munitions_toggled)
        form.addRow("Show munitions", self.include_munitions_check)
        self._rebuild_class_combo()
        self.qty_spin = QSpinBox()
        self.qty_spin.setRange(1, 100000)
        self.qty_spin.setValue(1)
        self.qty_spin.setToolTip("Quantity assigned to each new EntityComposition record.")
        form.addRow("Quantity per entity", self.qty_spin)
        self.batch_spin = QSpinBox()
        self.batch_spin.setRange(1, 100)
        self.batch_spin.setValue(1)
        self.batch_spin.setToolTip("Number of EntityComposition records to create.")
        form.addRow("Entities to create", self.batch_spin)
        self.federate_combo = QComboBox()
        for entry in self.FEDERATE_OPTIONS:
            self.federate_combo.addItem(entry)
        form.addRow("Owning Federate", self.federate_combo)
        self.inactive_check = QCheckBox("Mark entity as inactive")
        form.addRow("Is Inactive", self.inactive_check)
        self.transfer_check = QCheckBox("Allow transfer between federates")
        self.transfer_check.setChecked(True)
        form.addRow("Is Transferrable", self.transfer_check)
        self.invincible_check = QCheckBox("Invincible entity")
        form.addRow("Is Invincible", self.invincible_check)
        layout.addLayout(form)
        buttons = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def _collect_entity_classes(self) -> List[Dict[str, Any]]:
        results: List[Dict[str, Any]] = []
        parent = jfind(self._model.classdata, jtag("EntityCompositionClassList"))
        if parent is None:
            return results
        seen: Set[str] = set()
        for child in list(parent):
            if _local(child.tag) != "EntityCompositionClass":
                continue
            label = (text(jfind(child, "j:Name")) or "").strip()
            if not label:
                label = (text(jfind(child, "j:ClassName")) or "").strip()
            if not label:
                continue
            upper = label.upper()
            if upper in seen:
                continue
            seen.add(upper)
            discode = next((node for node in list(child) if _local(node.tag) == "DisCode"), None)
            if discode is None:
                discode = next((node for node in _iter_local(child, "DisCode")), None)
            kind_attr = None
            if discode is not None:
                kind_attr = discode.get("kind") or discode.get("Kind")
            results.append(
                {
                    "label": label,
                    "element": child,
                    "is_munition": kind_attr == "2",
                }
            )
        results.sort(key=lambda entry: entry["label"].upper())
        return results

    def _filtered_class_entries(self) -> List[Dict[str, Any]]:
        entries = list(self._all_class_entries)
        if self.include_munitions_check.isChecked():
            return entries
        return [entry for entry in entries if not entry.get("is_munition")]

    def _rebuild_class_combo(self) -> None:
        current_label = None
        current_entry = self.class_combo.currentData()
        if isinstance(current_entry, dict):
            current_label = current_entry.get("label")
        current_text = (self.class_combo.currentText() or "").strip()
        entries = self._filtered_class_entries()
        self.class_combo.blockSignals(True)
        self.class_combo.clear()
        self.class_combo.addItem("")
        index_to_select = 0
        for entry in entries:
            self.class_combo.addItem(entry["label"], entry)
            if current_label and entry["label"] == current_label:
                index_to_select = self.class_combo.count() - 1
        if index_to_select:
            self.class_combo.setCurrentIndex(index_to_select)
        elif current_text and not current_label:
            self.class_combo.setEditText(current_text)
        self.class_combo.blockSignals(False)

    def _on_include_munitions_toggled(self, _: bool) -> None:
        self._rebuild_class_combo()

    def _collect_existing_roles(self, parent_unit: ET._Element) -> Set[str]:
        roles: Set[str] = set()
        parent_lvc = (text(jfind(parent_unit, "j:LvcId")) or "").strip()
        if not parent_lvc:
            return roles
        for ec in _iter_local(self._model.entcomp_list, "EntityComposition"):
            sup = (text(jfind(ec, "j:UnitSuperior")) or "").strip()
            if sup != parent_lvc:
                continue
            role = (text(jfind(ec, "j:Role")) or "").strip()
            if role:
                roles.add(role.upper())
        return roles

    def result_data(self) -> Optional[_NewEntityData]:
        return self._result

    def accept(self) -> None:
        role = (self.role_edit.text() or "").strip()
        if not role:
            QMessageBox.warning(self, "Create Entity Composition", "Role is required.")
            self.role_edit.setFocus()
            return
        upper_role = role.upper()
        if upper_role in self._existing_roles:
            QMessageBox.warning(
                self,
                "Create Entity Composition",
                f"Role '{role}' already exists for this unit.",
            )
            self.role_edit.setFocus()
            self.role_edit.selectAll()
            return
        class_entry = self.class_combo.currentData()
        class_name = None
        if isinstance(class_entry, dict):
            class_name = class_entry.get("label")
        else:
            raw_text = (self.class_combo.currentText() or "").strip()
            class_name = raw_text or None
        self._result = _NewEntityData(
            role=role,
            class_name=class_name,
            owning_federate=(self.federate_combo.currentText() or "WARSIM").strip(),
            is_inactive=self.inactive_check.isChecked(),
            is_transferrable=self.transfer_check.isChecked(),
            is_invincible=self.invincible_check.isChecked(),
            quantity=int(self.qty_spin.value()),
            batch_count=int(self.batch_spin.value()),
        )
        super().accept()


class _UnitEditDialog(QDialog):

    def __init__(
        self,
        parent: QWidget,
        *,
        name: str,
        class_name: str,
        class_options: List[Dict[str, Any]],
        milstd_code: str,
    ):
        super().__init__(parent)
        self.setWindowTitle("Edit Unit")
        self._class_options = class_options
        self._initial_code = milstd_code or ""
        self._pending_code = self._initial_code
        layout = QVBoxLayout(self)
        form = QFormLayout()
        self.name_edit = QLineEdit(name)
        form.addRow("Name:", self.name_edit)
        self.class_combo = QComboBox()
        self.class_combo.setEditable(True)
        self.class_combo.setInsertPolicy(QComboBox.InsertPolicy.NoInsert)
        line_edit = self.class_combo.lineEdit()
        if line_edit is not None:
            line_edit.setPlaceholderText("Select or type class name")
            line_edit.textEdited.connect(self._handle_class_text_changed)
        self.class_combo.addItem("")
        for entry in class_options:
            self.class_combo.addItem(entry["name"], entry)
        self.class_combo.currentIndexChanged.connect(self._handle_class_changed)
        if class_name:
            idx = self.class_combo.findText(class_name, Qt.MatchFlag.MatchExactly)
            if idx >= 0:
                self.class_combo.setCurrentIndex(idx)
            else:
                self.class_combo.setEditText(class_name)
        form.addRow("Class Name:", self.class_combo)
        self.code_label = QLabel(self._pending_code or "Not set")
        self.code_label.setTextInteractionFlags(Qt.TextInteractionFlag.TextSelectableByMouse)
        form.addRow("MilStd2525C:", self.code_label)
        self._update_pending_code()
        layout.addLayout(form)
        buttons = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def _handle_class_changed(self, _: int) -> None:
        self._update_pending_code()

    def _handle_class_text_changed(self, _: str) -> None:
        self._update_pending_code()

    def _current_class_entry(self) -> Optional[Dict[str, Any]]:
        data = self.class_combo.currentData()
        if isinstance(data, dict):
            return data
        text_val = (self.class_combo.currentText() or "").strip().upper()
        for entry in self._class_options:
            if entry["name"].upper() == text_val:
                return entry
        return None

    def _update_pending_code(self) -> None:
        text = (self.class_combo.currentText() or "").strip()
        entry = self._current_class_entry()
        if not text:
            self._pending_code = ""
        elif entry is not None:
            self._pending_code = entry.get("code") or ""
        else:
            self._pending_code = self._initial_code
        if hasattr(self, "code_label"):
            self.code_label.setText(self._pending_code or "Not set")

    def values(self) -> Tuple[str, str, str, Optional[Dict[str, Any]]]:
        new_name = (self.name_edit.text() or "").strip()
        new_class = (self.class_combo.currentText() or "").strip()
        return new_name, new_class, self._pending_code, self._current_class_entry()


class _EntityEditDialog(QDialog):

    def __init__(
        self,
        parent: QWidget,
        *,
        role: str,
        class_name: str,
        class_options: List[Dict[str, Any]],
    ):
        super().__init__(parent)
        self.setWindowTitle("Edit Entity Composition")
        self._class_options = class_options
        layout = QVBoxLayout(self)
        form = QFormLayout()
        self.role_edit = QLineEdit(role)
        form.addRow("Role:", self.role_edit)
        self.class_combo = QComboBox()
        self.class_combo.setEditable(True)
        self.class_combo.setInsertPolicy(QComboBox.InsertPolicy.NoInsert)
        self.class_combo.addItem("")
        for entry in class_options:
            self.class_combo.addItem(entry["label"], entry)
        line_edit = self.class_combo.lineEdit()
        if line_edit is not None:
            line_edit.setPlaceholderText("Select or type class name")
        if class_name:
            idx = self.class_combo.findText(class_name, Qt.MatchFlag.MatchExactly)
            if idx >= 0:
                self.class_combo.setCurrentIndex(idx)
            else:
                self.class_combo.setEditText(class_name)
        form.addRow("Class Name:", self.class_combo)
        layout.addLayout(form)
        buttons = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def _current_class_entry(self) -> Optional[Dict[str, Any]]:
        data = self.class_combo.currentData()
        if isinstance(data, dict):
            return data
        text_val = (self.class_combo.currentText() or "").strip().upper()
        for option in self._class_options:
            if option["label"].upper() == text_val:
                return option
        return None

    def values(self) -> Tuple[str, str, Optional[Dict[str, Any]]]:
        new_role = (self.role_edit.text() or "").strip()
        new_class = (self.class_combo.currentText() or "").strip()
        return new_role, new_class, self._current_class_entry()


class _SymbolIconFactory:
    """Renders milsymbol SIDCs into cached Qt icons using QJSEngine."""

    def __init__(self, icon_px: int = 36):
        self._default_px = max(16, min(int(icon_px), 96))
        self._cache: Dict[Tuple[str, int], QIcon] = {}
        self._failed: Set[Tuple[str, int]] = set()
        self._engine: Optional[Any] = None
        self._render_func: Optional[Any] = None
        self._ready = False
        self._initialize()

    @property
    def is_ready(self) -> bool:
        return self._ready

    def icon_for(self, sidc: str, size: Optional[int] = None) -> Optional[QIcon]:
        sidc = (sidc or "").strip()
        if not sidc or not self._ready:
            return None
        px = self._normalize_size(size if size is not None else self._default_px)
        key = (sidc, px)
        cached = self._cache.get(key)
        if cached is not None:
            return cached
        if key in self._failed:
            return None
        svg = self._render_svg(sidc, px)
        if not svg:
            self._failed.add(key)
            return None
        icon = self._svg_to_icon(svg, px)
        if icon is None:
            self._failed.add(key)
            return None
        self._cache[key] = icon
        return icon

    def _initialize(self) -> None:
        if QJSEngine is None or QSvgRenderer is None:
            return
        js_path = _resolve_symbols_path("milsymbol.js")
        if js_path is None or not js_path.exists():
            return
        try:
            source = js_path.read_text(encoding="utf-8")
        except Exception:
            return
        try:
            engine = QJSEngine()
        except Exception:
            return
        result = engine.evaluate(source, str(js_path))
        if self._is_js_error(result):
            return
        wrapper = engine.evaluate(
            """
            var __milsymbol_render__ = (function () {
                var lib = null;
                if (typeof ms !== 'undefined' && ms.Symbol) {
                    lib = ms;
                } else if (typeof milsymbol !== 'undefined' && milsymbol.Symbol) {
                    lib = milsymbol;
                }
                return function (sidc, size) {
                    if (!lib || !sidc) {
                        return "";
                    }
                    try {
                        var symbol = new lib.Symbol(sidc, { size: size || 36 });
                        return symbol.asSVG();
                    } catch (err) {
                        return "";
                    }
                };
            })();
            """,
            "milsymbol_render_wrapper.js",
        )
        if self._is_js_error(wrapper):
            return
        func = engine.globalObject().property("__milsymbol_render__")
        if not func.isCallable():
            return
        self._engine = engine
        self._render_func = func
        self._ready = True

    def _render_svg(self, sidc: str, px: int) -> str:
        if self._render_func is None:
            return ""
        px = self._normalize_size(px)
        try:
            result = self._render_func.call([sidc, px])
        except Exception:
            return ""
        if self._is_js_error(result):
            return ""
        if hasattr(result, "isString") and result.isString():
            return result.toString()
        if hasattr(result, "toVariant"):
            value = result.toVariant()
            if isinstance(value, str):
                return value
        return ""

    def _is_js_error(self, value: Any) -> bool:
        try:
            return bool(value is not None and hasattr(value, "isError") and value.isError())
        except Exception:
            return True

    def _svg_to_icon(self, svg: str, px: int) -> Optional[QIcon]:
        if not svg or QSvgRenderer is None:
            return None
        data = QByteArray(svg.encode("utf-8"))
        renderer = QSvgRenderer(data)
        if not renderer.isValid():
            return None
        size = self._normalize_size(px)
        image = QImage(size, size, QImage.Format_ARGB32)
        image.fill(Qt.transparent)
        painter = QPainter(image)
        try:
            renderer.render(painter, QRectF(0, 0, size, size))
        finally:
            painter.end()
        pixmap = QPixmap.fromImage(image)
        if pixmap.isNull():
            return None
        return QIcon(pixmap)

    def _normalize_size(self, value: Optional[int]) -> int:
        if value is None:
            return self._default_px
        try:
            num = int(value)
        except Exception:
            num = self._default_px
        return max(16, min(num, 256))


class ModelPreviewDialog(QDialog):

    SIDE_ORDER = list(DEFAULT_SIDE_ORDER)

    def __init__(self, parent: Optional[QWidget] = None):
        super().__init__(parent)
        self._model: Optional[ObsModel] = None
        self._last_consolidation_label: Optional[str] = None
        self._cached_lvc_to_unit: Optional[Dict[str, ET._Element]] = None
        self._cached_parent_to_children: Optional[Dict[str, List[ET._Element]]] = None
        self._symbol_icon_factory: Optional[_SymbolIconFactory] = None
        self._symbol_icons_disabled = False
        self._icon_base_px = 36
        self._hover_icon_px = self._icon_base_px * 5
        self.setWindowTitle("Model Hierarchy Preview")
        self.resize(720, 560)
        layout = QVBoxLayout(self)
        search_row = QHBoxLayout()
        search_row.addWidget(QLabel("Filter:"))
        self._filter_edit = QLineEdit()
        self._filter_edit.setPlaceholderText(
            "Type to filter by name, role, or LvcId..."
        )
        search_row.addWidget(self._filter_edit)
        self._clear_filter_btn = QPushButton("Clear")
        self._clear_filter_btn.setAutoDefault(False)
        self._clear_filter_btn.clicked.connect(self._clear_filter)
        search_row.addWidget(self._clear_filter_btn)
        search_row.addStretch()
        layout.addLayout(search_row)
        self._filter_edit.textChanged.connect(self._on_filter_changed)
        self.tree = QTreeWidget()
        self.tree.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        self.tree.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.tree.customContextMenuRequested.connect(self._on_tree_context_menu)
        self.tree.setColumnCount(2)
        self.tree.setHeaderLabels(["Name / Role", "2525C"])
        self.tree.setAlternatingRowColors(True)
        header = self.tree.header()
        header.setStretchLastSection(False)
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        self.tree.viewport().installEventFilter(self)
        layout.addWidget(self.tree)
        btn_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Close)
        self._export_btn = btn_box.addButton(
            "Export to Excel...", QDialogButtonBox.ButtonRole.ActionRole
        )
        self._export_btn.setAutoDefault(False)
        self._export_btn.clicked.connect(self._on_export_excel)
        btn_box.rejected.connect(self.reject)
        layout.addWidget(btn_box)

    def _invalidate_unit_cache(self) -> None:
        self._cached_lvc_to_unit = None
        self._cached_parent_to_children = None

    def _ensure_unit_maps(
        self,
    ) -> Tuple[Dict[str, ET._Element], Dict[str, List[ET._Element]]]:
        if self._model is None:
            return {}, {}
        if (
            self._cached_lvc_to_unit is None
            or self._cached_parent_to_children is None
        ):
            self._cached_lvc_to_unit, self._cached_parent_to_children = (
                collect_unit_maps(self._model)
            )
        return self._cached_lvc_to_unit, self._cached_parent_to_children

    def _collect_unit_class_entries(self) -> List[Dict[str, Any]]:
        entries: List[Dict[str, Any]] = []
        if self._model is None:
            return entries
        parent = jfind(self._model.classdata, jtag("UnitClassList"))
        if parent is None:
            return entries
        seen: Set[str] = set()
        for child in list(parent):
            if _local(child.tag) != "UnitClass":
                continue
            name = (text(jfind(child, "j:AggregateName")) or "").strip()
            if not name:
                continue
            key = name.upper()
            if key in seen:
                continue
            seen.add(key)
            code = (text(jfind(child, "j:MilStd2525CCode")) or "").strip()
            entries.append({"name": name, "code": code, "element": child})
        entries.sort(key=lambda entry: entry["name"].upper())
        return entries

    def _collect_entity_class_entries(self) -> List[Dict[str, Any]]:
        entries: List[Dict[str, Any]] = []
        if self._model is None:
            return entries
        parent = jfind(self._model.classdata, jtag("EntityCompositionClassList"))
        if parent is None:
            return entries
        seen: Set[str] = set()
        for child in list(parent):
            if _local(child.tag) != "EntityCompositionClass":
                continue
            label = (text(jfind(child, "j:Name")) or "").strip()
            if not label:
                label = (text(jfind(child, "j:ClassName")) or "").strip()
            if not label:
                continue
            key = label.upper()
            if key in seen:
                continue
            seen.add(key)
            entries.append({"label": label, "element": child})
        entries.sort(key=lambda entry: entry["label"].upper())
        return entries

    def set_model(self, model: Optional[ObsModel]) -> None:
        self._model = model
        self._invalidate_unit_cache()
        self.tree.clear()
        if model is None:
            return
        lvc_to_unit, parent_to_children = collect_unit_maps(model)
        self._cached_lvc_to_unit = dict(lvc_to_unit)
        self._cached_parent_to_children = dict(parent_to_children)
        ec_by_superior: Dict[str, List[ET._Element]] = {}
        orphan_ec: Dict[str, List[Tuple[ET._Element, str]]] = {}
        side_labels: Dict[str, str] = {}
        for ec in _iter_local(model.entcomp_list, "EntityComposition"):
            sup = (text(jfind(ec, "j:UnitSuperior")) or "").strip()
            if sup and sup in lvc_to_unit:
                ec_by_superior.setdefault(sup, []).append(ec)
            else:
                side_name = (text(jfind(ec, "j:SideSuperior")) or "").strip()
                side_key = side_name.upper() if side_name else "UNKNOWN"
                if side_name:
                    side_labels.setdefault(side_key, side_name)
                orphan_ec.setdefault(side_key, []).append((ec, sup))
        units = list(_iter_local(model.unit_list, "Unit"))
        side_roots: Dict[str, List[ET._Element]] = {}

        def unit_side(unit: ET._Element) -> str:
            raw = (text(jfind(unit, "j:SideSuperior")) or "").strip()
            if raw:
                key = raw.upper()
                side_labels.setdefault(key, raw)
                return key
            derived = _resolve_unit_side(unit, lvc_to_unit)
            if derived:
                key = derived.upper()
                side_labels.setdefault(key, derived)
                return key
            side_labels.setdefault("UNKNOWN", "UNKNOWN")
            return "UNKNOWN"

        for unit in units:
            side_key = unit_side(unit)
            parent = (text(jfind(unit, "j:UnitSuperior")) or "").strip()
            if not parent or parent not in lvc_to_unit:
                side_roots.setdefault(side_key, []).append(unit)
        all_sides = set(side_labels) | set(side_roots) | set(orphan_ec)
        if not all_sides:
            all_sides.add("UNKNOWN")
            side_labels.setdefault("UNKNOWN", "UNKNOWN")
        side_items: Dict[str, QTreeWidgetItem] = {}
        for side_key in sorted(all_sides, key=side_sort_key):
            self._ensure_side_item(side_key, side_labels, side_items)
        visited_units: Set[int] = set()
        for side_key, roots in side_roots.items():
            parent_item = self._ensure_side_item(side_key, side_labels, side_items)
            for unit in sorted(roots, key=self._unit_sort_key):
                self._add_unit_recursive(
                    unit,
                    parent_item,
                    lvc_to_unit,
                    parent_to_children,
                    ec_by_superior,
                    visited_units,
                )
        leftovers = [u for u in units if id(u) not in visited_units]
        if leftovers:
            side_item = self._ensure_side_item("UNKNOWN", side_labels, side_items)
            placeholder = QTreeWidgetItem(["Detached Units", ""])
            side_item.addChild(placeholder)
            for unit in sorted(leftovers, key=self._unit_sort_key):
                self._add_unit_recursive(
                    unit,
                    placeholder,
                    lvc_to_unit,
                    parent_to_children,
                    ec_by_superior,
                    visited_units,
                )
        for side_key, entries in orphan_ec.items():
            side_item = self._ensure_side_item(side_key, side_labels, side_items)
            orphan_parent = QTreeWidgetItem(["Entity Compositions (unattached)", ""])
            side_item.addChild(orphan_parent)
            for ec, sup in sorted(
                entries, key=lambda pair: self._entity_sort_key(pair[0])
            ):
                label = self._format_entity_label(
                    ec, include_superior=True, superior_hint=sup
                )
                code = (text(jfind(ec, "j:MilStd2525CCode")) or "").strip()
                item = QTreeWidgetItem([label, code])
                item.setData(0, Qt.ItemDataRole.UserRole, ("entity", ec))
                orphan_parent.addChild(item)
        current_filter = (self._filter_edit.text() or "").strip().lower()
        if current_filter:
            self._apply_filter(current_filter)
        else:
            self.tree.expandToDepth(0)

    def eventFilter(self, obj: QObject, event: QEvent) -> bool:
        if obj is self.tree.viewport() and isinstance(event, QHelpEvent):
            if self._handle_icon_tooltip(event):
                return True
        return super().eventFilter(obj, event)

    def _clear_filter(self) -> None:
        if hasattr(self, "_filter_edit"):
            self._filter_edit.blockSignals(True)
            self._filter_edit.clear()
            self._filter_edit.blockSignals(False)
        self._apply_filter("")

    def _on_filter_changed(self, text: str) -> None:
        term = (text or "").strip().lower()
        self._apply_filter(term)

    def _apply_filter(self, term: str) -> None:
        if not term:
            for i in range(self.tree.topLevelItemCount()):
                item = self.tree.topLevelItem(i)
                self._set_hidden_recursive(item, False)
            self.tree.collapseAll()
            self.tree.expandToDepth(0)
            return
        for i in range(self.tree.topLevelItemCount()):
            item = self.tree.topLevelItem(i)
            visible = self._filter_item_recursive(item, term)
            item.setHidden(not visible)
        self.tree.expandAll()

    def _on_export_excel(self) -> None:
        if self._model is None:
            QMessageBox.information(
                self, "Export Preview", "Load a model before exporting the preview."
            )
            return
        rows = self._collect_tree_rows(include_hidden=True)
        if not rows:
            QMessageBox.information(
                self, "Export Preview", "The preview tree is empty."
            )
            return
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        suggested = f"ModelPreview_{timestamp}"
        path, _ = QFileDialog.getSaveFileName(
            self, "Export Preview to Excel", suggested, "Excel Workbook (*.xlsx)"
        )
        if not path:
            return
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"
        try:
            self._write_preview_excel(path, rows)
        except ImportError:
            QMessageBox.critical(
                self,
                "Export Preview",
                "openpyxl is required to export Excel files. Install openpyxl and try again.",
            )
        except Exception as exc:
            traceback.print_exc()
            QMessageBox.critical(
                self, "Export Preview", f"Failed to export preview: {exc}"
            )
        else:
            QMessageBox.information(
                self, "Export Preview", f"Preview exported to:\n{path}"
            )

    def _on_tree_context_menu(self, pos) -> None:
        item = self.tree.itemAt(pos)
        if item is None:
            return
        if not item.isSelected():
            self.tree.clearSelection()
            item.setSelected(True)
        selected_items: List[Tuple[str, ET._Element]] = []
        for candidate in self.tree.selectedItems():
            data = _normalize_user_data(candidate.data(0, Qt.ItemDataRole.UserRole))
            if isinstance(data, tuple) and data:
                kind = data[0]
                if kind in ("unit", "entity"):
                    selected_items.append(data)
        if not selected_items:
            return
        units = [data for data in selected_items if data[0] == "unit"]
        entities = [data for data in selected_items if data[0] == "entity"]
        menu = QMenu(self)
        rename_action = None
        move_action = None
        create_action = None
        create_entity_action = None
        mass_rename_action = None
        duplicate_action = None
        if len(selected_items) == 1:
            kind, _ = selected_items[0]
            if kind == "unit":
                rename_action = menu.addAction("Edit unit...")
                move_action = menu.addAction("Move unit...")
                create_action = menu.addAction("Add subordinate unit...")
                create_entity_action = menu.addAction("Add entity composition...")
                mass_rename_action = menu.addAction("Mass rename subtree...")
                duplicate_action = menu.addAction("Duplicate unit subtree...")
            elif kind == "entity":
                rename_action = menu.addAction("Edit entity composition...")
                move_action = menu.addAction("Move entity to unit...")
            if rename_action or move_action or create_action or duplicate_action:
                menu.addSeparator()
        delete_action = None
        consolidate_action = None
        if units or entities:
            if units and entities:
                delete_text = f"Delete {len(units)} unit(s) and {len(entities)} entity composition(s)"
            elif units:
                delete_text = (
                    "Delete selected unit"
                    if len(units) == 1
                    else f"Delete {len(units)} units and subordinates"
                )
            else:
                delete_text = (
                    "Delete selected entity composition"
                    if len(entities) == 1
                    else f"Delete {len(entities)} entity compositions"
                )
            delete_action = menu.addAction(delete_text)
        if units:
            if len(units) == 1:
                consolidate_text = "Consolidate subtree into this unit..."
            else:
                consolidate_text = (
                    f"Consolidate {len(units)} units into selected level..."
                )
            consolidate_action = menu.addAction(consolidate_text)
            if entities:
                consolidate_action.setEnabled(False)
        chosen = menu.exec(self.tree.viewport().mapToGlobal(pos))
        if chosen is None:
            return
        if create_action is not None and chosen == create_action:
            self._handle_create_unit(selected_items[0][1])
            return
        if create_entity_action is not None and chosen == create_entity_action:
            self._handle_create_entity(selected_items[0][1])
            return
        if rename_action is not None and chosen == rename_action:
            self._handle_rename_request(selected_items[0])
            return
        if mass_rename_action is not None and chosen == mass_rename_action:
            self._handle_mass_rename_subtree(selected_items[0][1])
            return
        if duplicate_action is not None and chosen == duplicate_action:
            self._handle_duplicate_subtree(selected_items[0][1])
            return
        if move_action is not None and chosen == move_action:
            self._handle_move_request(selected_items[0])
            return
        if delete_action is not None and chosen == delete_action:
            self._handle_delete_request(selected_items)
            return
        if (
            consolidate_action is not None
            and chosen == consolidate_action
            and (
                consolidate_action.isEnabled()
                if hasattr(consolidate_action, "isEnabled")
                else not entities
            )
        ):
            self._handle_consolidate_subtree([data[1] for data in units])

    def _handle_delete_request(self, payload) -> None:
        if self._model is None:
            return
        if isinstance(payload, list):
            raw_items = payload
        elif isinstance(payload, tuple):
            raw_items = [payload]
        elif payload is None:
            raw_items = []
        else:
            raw_items = list(payload)
        unique_payloads = []
        seen = set()
        for entry in raw_items:
            if not isinstance(entry, tuple) or len(entry) < 2:
                continue
            kind, element = entry[0], entry[1]
            if kind not in ("unit", "entity"):
                continue
            key = (kind, id(element))
            if key in seen:
                continue
            seen.add(key)
            unique_payloads.append((kind, element))
        if not unique_payloads:
            return
        changed = False
        unit_targets = []
        skipped_units = []
        for kind, element in unique_payloads:
            if kind != "unit":
                continue
            lvc = (text(jfind(element, "j:LvcId")) or "").strip()
            if not lvc:
                skipped_units.append(self._format_unit_label(element))
                continue
            unit_targets.append((element, lvc))
        if unit_targets:
            labels = [self._format_unit_label(elem) for elem, _ in unit_targets]
            if len(labels) == 1:
                prompt = f"Delete unit '{labels[0]}' and all subordinate units and entity compositions?"
            else:
                preview = ", ".join(labels[:5])
                if len(labels) > 5:
                    preview += ", ..."
                prompt = f"Delete {len(labels)} units and all subordinate units and entity compositions?\n\nExamples: {preview}"
            confirm = QMessageBox.question(
                self,
                "Confirm Deletion",
                prompt,
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No,
            )
            if confirm == QMessageBox.StandardButton.Yes:
                removed_units = 0
                removed_ecs = 0
                for _, lvc in unit_targets:
                    ru, re = remove_unit_subtree(self._model, lvc)
                    removed_units += ru
                    removed_ecs += re
                if removed_units == 0 and removed_ecs == 0:
                    QMessageBox.information(
                        self, "Delete Unit", "No units were removed."
                    )
                else:
                    changed = True
        if skipped_units:
            message = (
                "The following units do not have an LvcId and were skipped:\n"
                + "\n".join(skipped_units[:10])
            )
            if len(skipped_units) > 10:
                message += "\n..."
            QMessageBox.warning(self, "Delete Unit", message)
        entity_targets = [
            element for kind, element in unique_payloads if kind == "entity"
        ]
        if entity_targets:
            labels = [self._format_entity_label(elem) for elem in entity_targets]
            if len(labels) == 1:
                prompt = f"Delete entity composition '{labels[0]}'?"
            else:
                preview = ", ".join(labels[:5])
                if len(labels) > 5:
                    preview += ", ..."
                prompt = (
                    f"Delete {len(labels)} entity compositions?\n\nExamples: {preview}"
                )
            confirm = QMessageBox.question(
                self,
                "Confirm Deletion",
                prompt,
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No,
            )
            if confirm == QMessageBox.StandardButton.Yes:
                removed = 0
                missing = []
                for elem in entity_targets:
                    try:
                        self._model.entcomp_list.remove(elem)
                        removed += 1
                    except ValueError:
                        missing.append(self._format_entity_label(elem))
                if removed:
                    changed = True
                if missing:
                    detail = "\n".join(missing[:10])
                    if len(missing) > 10:
                        detail += "\n..."
                    QMessageBox.information(
                        self,
                        "Delete Entity Composition",
                        "Some entity compositions were not found and could not be removed:\n"
                        + detail,
                    )
        if changed:
            self.set_model(self._model)

    def _handle_rename_request(self, payload: Tuple[str, ET._Element]) -> None:
        if self._model is None:
            return
        kind, element = payload
        changed = False
        if kind == "unit":
            changed = self._rename_unit(element)
        elif kind == "entity":
            changed = self._rename_entity(element)
        if changed:
            self.set_model(self._model)

    def _collect_unit_name_set(
        self, *, exclude_ids: Optional[Set[int]] = None
    ) -> Set[str]:
        names: Set[str] = set()
        if self._model is None:
            return names
        for unit in _iter_local(self._model.unit_list, "Unit"):
            if exclude_ids and id(unit) in exclude_ids:
                continue
            value = (text(jfind(unit, "j:Name")) or "").strip()
            if value:
                names.add(value.upper())
        return names

    def _open_mass_rename_dialog(
        self,
        unit_elem: ET._Element,
        *,
        existing_names: Set[str],
        context_text: Optional[str] = None,
    ) -> Optional[_MassRenameResult]:
        if self._model is None:
            return None
        unit_maps = self._ensure_unit_maps()
        dialog = MassRenameDialog(
            self,
            self._model,
            unit_elem,
            existing_unit_names=existing_names,
            context_text=context_text,
            unit_maps=unit_maps,
        )
        if dialog.exec() != QDialog.DialogCode.Accepted:
            return None
        return dialog.result()

    def _handle_mass_rename_subtree(self, unit_elem: ET._Element) -> None:
        if self._model is None or unit_elem is None:
            return
        _, parent_to_children = self._ensure_unit_maps()
        scope_units = _gather_unit_descendants(
            unit_elem, parent_to_children, include_root=True
        )
        scope_ids = {id(u) for u in scope_units}
        existing_names = self._collect_unit_name_set(exclude_ids=scope_ids)
        result = self._open_mass_rename_dialog(
            unit_elem, existing_names=existing_names
        )
        if result is None:
            return
        units_changed = 0
        for target, new_name in result.unit_names.items():
            name_el = jfind(target, "j:Name")
            if name_el is None:
                name_el = ET.SubElement(target, jtag("Name"))
            if (name_el.text or "").strip() != new_name:
                name_el.text = new_name
                units_changed += 1
        entities_changed = 0
        for entity, new_role in result.entity_names.items():
            role_el = jfind(entity, "j:Role")
            if role_el is None:
                role_el = ET.SubElement(entity, jtag("Role"))
            if (role_el.text or "").strip() != new_role:
                role_el.text = new_role
                entities_changed += 1
        if units_changed == 0 and entities_changed == 0:
            QMessageBox.information(
                self,
                "Mass Rename",
                "No unit or entity names were changed.",
            )
            return
        summary_parts = []
        if units_changed:
            summary_parts.append(f"units: {units_changed}")
        if entities_changed:
            summary_parts.append(f"entities: {entities_changed}")
        self.set_model(self._model)
        self._select_tree_item_for_unit(unit_elem)
        QMessageBox.information(
                self,
                "Mass Rename",
                "Updated " + ", ".join(summary_parts) + ".",
            )

    def _handle_duplicate_subtree(self, unit_elem: ET._Element) -> None:
        if self._model is None or unit_elem is None:
            return
        lvc_map, parent_to_children = self._ensure_unit_maps()
        subtree_units = _gather_unit_descendants(
            unit_elem, parent_to_children, include_root=True
        )
        if not subtree_units:
            QMessageBox.information(
                self, "Duplicate Unit", "No units were found under the selection."
            )
            return
        subtree_root = subtree_units[0]
        duplicates_allowed = self._collect_unit_name_set(exclude_ids=None)
        context_text = (
            "Configure the names for the duplicated units and their entity compositions. "
            "The original units remain unchanged."
        )
        rename_plan = self._open_mass_rename_dialog(
            unit_elem, existing_names=duplicates_allowed, context_text=context_text
        )
        if rename_plan is None:
            return
        subtree_lvc: Set[str] = set()
        for unit in subtree_units:
            lvc = (text(jfind(unit, "j:LvcId")) or "").strip()
            if lvc:
                subtree_lvc.add(lvc)
        subtree_entities: List[ET._Element] = []
        if subtree_lvc and self._model.entcomp_list is not None:
            for ec in _iter_local(self._model.entcomp_list, "EntityComposition"):
                sup = (text(jfind(ec, "j:UnitSuperior")) or "").strip()
                if sup and sup in subtree_lvc:
                    subtree_entities.append(ec)
        root_clone, unit_count, entity_count = self._duplicate_subtree_elements(
            subtree_units,
            subtree_entities,
            rename_plan.all_unit_names,
            rename_plan.all_entity_names,
        )
        if root_clone is None or unit_count == 0:
            QMessageBox.warning(
                self, "Duplicate Unit", "Failed to duplicate the selected unit."
            )
            return
        self.set_model(self._model)
        self._select_tree_item_for_unit(root_clone)
        QMessageBox.information(
            self,
            "Duplicate Unit",
            f"Duplicated {unit_count} unit(s) and {entity_count} entity composition(s).",
        )

    def _rename_unit(self, unit_elem: ET._Element) -> bool:
        name_el = jfind(unit_elem, "j:Name")
        old_name = (
            (name_el.text or "").strip() if name_el is not None and name_el.text else ""
        )
        class_el = jfind(unit_elem, "j:ClassName")
        old_class = (
            (class_el.text or "").strip()
            if class_el is not None and class_el.text
            else ""
        )
        code_el = jfind(unit_elem, "j:MilStd2525CCode")
        old_code = (
            (code_el.text or "").strip()
            if code_el is not None and code_el.text
            else ""
        )
        class_entries = self._collect_unit_class_entries()
        dialog = _UnitEditDialog(
            self,
            name=old_name,
            class_name=old_class,
            class_options=class_entries,
            milstd_code=old_code,
        )
        if dialog.exec() != QDialog.DialogCode.Accepted:
            return False
        new_name, new_class, target_code, selected_entry = dialog.values()
        if not new_name:
            QMessageBox.information(self, "Edit Unit", "Name cannot be empty.")
            return False
        name_changed = new_name != old_name
        if name_el is None:
            name_el = ET.SubElement(unit_elem, jtag("Name"))
        updated_children = 0
        if name_changed:
            name_el.text = new_name
        old_base = ""
        if name_changed and old_name:
            old_base, _ = _split_unit_name_components(old_name)
        new_base, _ = _split_unit_name_components(new_name)
        if name_changed and old_base and new_base and old_base.upper() != new_base.upper():
            _, parent_to_children = self._ensure_unit_maps()
            for descendant in _gather_unit_descendants(unit_elem, parent_to_children):
                name_node = jfind(descendant, "j:Name")
                if name_node is None or not name_node.text:
                    continue
                current_name = name_node.text.strip()
                child_base, child_suffix = _split_unit_name_components(current_name)
                if child_base and child_base.upper() == old_base.upper():
                    new_child = new_base
                    if child_suffix:
                        new_child = f"{new_base}_{child_suffix}"
                    name_node.text = new_child
                    updated_children += 1
        class_changed = False
        if new_class:
            if class_el is None:
                class_el = ET.SubElement(unit_elem, jtag("ClassName"))
            if (class_el.text or "").strip() != new_class:
                class_el.text = new_class
                class_changed = True
        else:
            if self._safe_remove_element(unit_elem, class_el):
                class_changed = True
                class_el = None
        target_code = (target_code or "").strip()
        code_changed = False
        if target_code:
            if code_el is None:
                code_el = ET.SubElement(unit_elem, jtag("MilStd2525CCode"))
            if (code_el.text or "").strip() != target_code:
                code_el.text = target_code
                code_changed = True
        else:
            if self._safe_remove_element(unit_elem, code_el):
                code_changed = True
                code_el = None
        enumeration_added = False
        enumeration_removed = False
        class_source = selected_entry.get("element") if isinstance(selected_entry, dict) else None
        if class_source is not None:
            if self._remove_unit_dis_enumeration(unit_elem):
                enumeration_removed = True
            self._copy_unitclass_enumeration(unit_elem, class_source)
            enumeration_added = True
        elif class_changed and not new_class:
            if self._remove_unit_dis_enumeration(unit_elem):
                enumeration_removed = True
        if (
            not name_changed
            and not class_changed
            and not code_changed
            and not enumeration_added
            and not enumeration_removed
        ):
            QMessageBox.information(self, "Edit Unit", "No changes were made.")
            return False
        details = []
        if name_changed:
            msg = f"Name set to '{new_name}'."
            if updated_children:
                msg += f" Updated {updated_children} subordinate name(s)."
            details.append(msg)
        if class_changed:
            if new_class:
                details.append(f"Class name set to '{new_class}'.")
            else:
                details.append("Class name cleared.")
        if code_changed:
            if target_code:
                details.append(f"2525C code set to '{target_code}'.")
            else:
                details.append("2525C code cleared.")
        if enumeration_added and enumeration_removed:
            details.append("Unit DIS enumeration refreshed from selected class.")
        elif enumeration_added:
            details.append("Unit DIS enumeration added from selected class.")
        elif enumeration_removed:
            details.append("Unit DIS enumeration cleared.")
        QMessageBox.information(self, "Edit Unit", " ".join(details))
        return True

    def _rename_entity(self, entity_elem: ET._Element) -> bool:
        role_el = jfind(entity_elem, "j:Role")
        old_role = (
            (role_el.text or "").strip() if role_el is not None and role_el.text else ""
        )
        class_el = jfind(entity_elem, "j:ClassName")
        old_class = (
            (class_el.text or "").strip()
            if class_el is not None and class_el.text
            else ""
        )
        class_entries = self._collect_entity_class_entries()
        dialog = _EntityEditDialog(
            self, role=old_role, class_name=old_class, class_options=class_entries
        )
        if dialog.exec() != QDialog.DialogCode.Accepted:
            return False
        new_role, new_class, _ = dialog.values()
        if not new_role:
            QMessageBox.information(
                self, "Edit Entity Composition", "Role cannot be empty."
            )
            return False
        role_changed = new_role != old_role
        if role_el is None:
            role_el = ET.SubElement(entity_elem, jtag("Role"))
        if role_changed:
            role_el.text = new_role
        name_el = jfind(entity_elem, "j:Name")
        if role_changed and name_el is not None:
            name_el.text = new_role
        class_changed = False
        if new_class:
            if class_el is None:
                class_el = ET.SubElement(entity_elem, jtag("ClassName"))
            if (class_el.text or "").strip() != new_class:
                class_el.text = new_class
                class_changed = True
        else:
            if class_el is not None:
                existing = (class_el.text or "").strip()
                if existing:
                    class_changed = True
                try:
                    entity_elem.remove(class_el)
                except Exception:
                    class_el.text = ""
        if not role_changed and not class_changed:
            QMessageBox.information(
                self, "Edit Entity Composition", "No changes were made."
            )
            return False
        details = []
        if role_changed:
            details.append(f"Role set to '{new_role}'.")
        if class_changed:
            if new_class:
                details.append(f"Class name set to '{new_class}'.")
            else:
                details.append("Class name cleared.")
        QMessageBox.information(
            self, "Edit Entity Composition", " ".join(details)
        )
        return True

    def _duplicate_subtree_elements(
        self,
        subtree_units: List[ET._Element],
        subtree_entities: List[ET._Element],
        name_plan: Dict[ET._Element, str],
        entity_plan: Dict[ET._Element, str],
    ) -> Tuple[Optional[ET._Element], int, int]:
        if self._model is None:
            return None, 0, 0
        ent_list = self._model.entcomp_list
        if ent_list is None:
            ent_list = ET.SubElement(self._model.scenario, jtag("EntityCompositionList"))
            self._model.entcomp_list = ent_list
        used_ids, used_lvc = self._collect_used_identifiers()
        orig_to_clone: Dict[ET._Element, ET._Element] = {}
        orig_lvc_to_new: Dict[str, str] = {}
        new_root: Optional[ET._Element] = None
        units_created = 0
        for unit in subtree_units:
            clone = clone_element(unit)
            target_name = name_plan.get(unit)
            if not target_name:
                target_name = (text(jfind(unit, "j:Name")) or "").strip() or "Unit"
            name_el = jfind(clone, "j:Name")
            if name_el is None:
                name_el = ET.SubElement(clone, jtag("Name"))
            name_el.text = target_name
            unit_id = clone.get("id")
            if not unit_id or unit_id in used_ids:
                unit_id = _gen_unit_xml_id(used_ids)
                clone.set("id", unit_id)
            used_ids.add(unit_id)
            clone.set("dateTimeStamp", _now_iso_seconds_no_tz())
            lvc_el = jfind(clone, "j:LvcId")
            new_lvc = _gen_unique_lvcid("U", used_lvc)
            if lvc_el is None:
                lvc_el = ET.SubElement(clone, jtag("LvcId"))
            lvc_el.text = new_lvc
            used_lvc.add(new_lvc)
            orig_lvc = (text(jfind(unit, "j:LvcId")) or "").strip()
            if orig_lvc:
                orig_lvc_to_new[orig_lvc] = new_lvc
            orig_to_clone[unit] = clone
            if new_root is None:
                new_root = clone
            sup_el = jfind(clone, "j:UnitSuperior")
            sup_val = (
                (sup_el.text or "").strip() if sup_el is not None and sup_el.text else ""
            )
            if sup_val and sup_val in orig_lvc_to_new:
                if sup_el is None:
                    sup_el = ET.SubElement(clone, jtag("UnitSuperior"))
                sup_el.text = orig_lvc_to_new[sup_val]
            units_created += 1
            self._model.unit_list.append(clone)
        entities_created = 0
        if subtree_entities:
            for entity in subtree_entities:
                sup = (text(jfind(entity, "j:UnitSuperior")) or "").strip()
                if not sup or sup not in orig_lvc_to_new:
                    continue
                clone = clone_element(entity)
                target_role = entity_plan.get(entity)
                if not target_role:
                    target_role = (text(jfind(entity, "j:Role")) or "").strip()
                if not target_role:
                    target_role = "ENTITY"
                role_el = jfind(clone, "j:Role")
                if role_el is None:
                    role_el = ET.SubElement(clone, jtag("Role"))
                role_el.text = target_role
                entity_id = clone.get("id")
                if not entity_id or entity_id in used_ids:
                    entity_id = _gen_unit_xml_id(used_ids)
                    clone.set("id", entity_id)
                used_ids.add(entity_id)
                lvc_el = jfind(clone, "j:LvcId")
                new_lvc = _gen_unique_lvcid("E", used_lvc)
                if lvc_el is None:
                    lvc_el = ET.SubElement(clone, jtag("LvcId"))
                lvc_el.text = new_lvc
                used_lvc.add(new_lvc)
                sup_el = jfind(clone, "j:UnitSuperior")
                if sup_el is None:
                    sup_el = ET.SubElement(clone, jtag("UnitSuperior"))
                sup_el.text = orig_lvc_to_new.get(sup, sup)
                ent_list.append(clone)
                entities_created += 1
        return new_root, units_created, entities_created

    def _handle_move_request(self, payload: Tuple[str, ET._Element]) -> None:
        if self._model is None:
            return
        kind, element = payload
        if kind == "unit":
            dialog = ObsMoveUnitDialog(
                self, self._model, blocked_units={element}, allow_side_selection=True
            )
        else:
            dialog = ObsMoveUnitDialog(
                self, self._model, blocked_units=set(), allow_side_selection=False
            )
        if dialog.exec() != QDialog.DialogCode.Accepted:
            return
        target = dialog.selected_target()
        if target is None:
            return
        if kind == "unit":
            message = self._move_unit(element, target)
        else:
            message = self._move_entity(element, target)
        if message:
            QMessageBox.information(self, "Move", message)
        self.set_model(self._model)

    def _handle_create_unit(self, parent_unit: ET._Element) -> None:
        if self._model is None:
            return
        dialog = CreateUnitDialog(self, self._model, parent_unit)
        if dialog.exec() != QDialog.DialogCode.Accepted:
            return
        data = dialog.result_data()
        if data is None:
            return
        new_unit = self._create_unit_under_parent(parent_unit, data)
        if new_unit is None:
            QMessageBox.warning(
                self,
                "Create Unit",
                "Failed to create the new unit. Verify the model is loaded correctly.",
            )
            return
        self.set_model(self._model)
        self._select_tree_item_for_unit(new_unit)
        QMessageBox.information(
            self,
            "Create Unit",
            f"Created new unit '{self._format_unit_label(new_unit)}'.",
        )

    def _handle_create_entity(self, parent_unit: ET._Element) -> None:
        if self._model is None:
            return
        dialog = CreateEntityDialog(self, self._model, parent_unit)
        if dialog.exec() != QDialog.DialogCode.Accepted:
            return
        data = dialog.result_data()
        if data is None:
            return
        created = self._create_entities_under_unit(parent_unit, data)
        if not created:
            QMessageBox.warning(
                self,
                "Create Entity Composition",
                "Failed to create the entity composition(s). Ensure the unit has an LvcId.",
            )
            return
        self.set_model(self._model)
        QMessageBox.information(
            self,
            "Create Entity Composition",
            f"Created {created} entity composition(s) under '{self._format_unit_label(parent_unit)}'.",
        )

    def _move_unit(self, unit_elem: ET._Element, target: Tuple[str, Any]) -> str:
        kind = target[0]
        label = self._format_unit_label(unit_elem)
        if kind == "unit":
            parent_unit = target[1]
            parent_lvc = (text(jfind(parent_unit, "j:LvcId")) or "").strip()
            if not parent_lvc:
                QMessageBox.warning(
                    self, "Move Unit", "Selected parent unit does not have an LvcId."
                )
                return ""
            current_sup = (text(jfind(unit_elem, "j:UnitSuperior")) or "").strip()
            if current_sup == parent_lvc:
                return "Unit is already assigned to the selected parent."
            sup_el = jfind(unit_elem, "j:UnitSuperior")
            if sup_el is None:
                sup_el = ET.SubElement(unit_elem, jtag("UnitSuperior"))
            sup_el.text = parent_lvc
            parent_label = self._format_unit_label(parent_unit)
            parent_side = (text(jfind(parent_unit, "j:SideSuperior")) or "").strip()
            updated = 0
            if parent_side:
                updated = self._apply_side_to_subtree(unit_elem, parent_side)
            message = f"Moved unit '{label}' under '{parent_label}'."
            if updated:
                message += f" Updated side '{parent_side}' on {updated} unit(s)."
            return message
        if kind == "side":
            side_key = target[1]
            side_display = target[2]
            sup_el = jfind(unit_elem, "j:UnitSuperior")
            if sup_el is not None:
                try:
                    unit_elem.remove(sup_el)
                except Exception:
                    sup_el.text = ""
            side_value = side_display or side_key
            updated = self._apply_side_to_subtree(unit_elem, side_value)
            return f"Moved unit '{label}' to side '{side_value}'. Updated side on {updated} unit(s)."
        QMessageBox.warning(self, "Move Unit", "Invalid destination selected.")
        return ""

    def _move_entity(self, entity_elem: ET._Element, target: Tuple[str, Any]) -> str:
        if target[0] != "unit":
            QMessageBox.warning(
                self, "Move Entity Composition", "Select a destination unit."
            )
            return ""
        parent_unit = target[1]
        parent_lvc = (text(jfind(parent_unit, "j:LvcId")) or "").strip()
        if not parent_lvc:
            QMessageBox.warning(
                self,
                "Move Entity Composition",
                "Selected destination unit does not have an LvcId.",
            )
            return ""
        sup_el = jfind(entity_elem, "j:UnitSuperior")
        if sup_el is None:
            sup_el = ET.SubElement(entity_elem, jtag("UnitSuperior"))
        sup_el.text = parent_lvc
        parent_side = (text(jfind(parent_unit, "j:SideSuperior")) or "").strip()
        if parent_side:
            side_el = jfind(entity_elem, "j:SideSuperior")
            if side_el is None:
                side_el = ET.SubElement(entity_elem, jtag("SideSuperior"))
            side_el.text = parent_side
        return (
            f"Moved entity composition under '{self._format_unit_label(parent_unit)}'."
        )

    def _create_unit_under_parent(
        self, parent_unit: ET._Element, data: _NewUnitData
    ) -> Optional[ET._Element]:
        used_ids, used_lvc = self._collect_used_identifiers()
        parent_lvc = self._ensure_unit_lvc(parent_unit, used_lvc)
        side_value = data.side_value
        if not side_value:
            lvc_to_unit, _ = self._ensure_unit_maps()
            derived = _resolve_unit_side(parent_unit, lvc_to_unit)
            side_value = derived
        if self._model is None:
            return None
        new_unit = ET.SubElement(self._model.unit_list, jtag("Unit"))
        new_unit.set("id", _gen_unit_xml_id(used_ids))
        new_unit.set("dateTimeStamp", _now_iso_seconds_no_tz())
        ET.SubElement(new_unit, jtag("Name")).text = data.name
        if data.class_name:
            ET.SubElement(new_unit, jtag("ClassName")).text = data.class_name
        if data.echelon:
            ET.SubElement(new_unit, jtag("Echelon")).text = data.echelon
        code_value = data.milstd_code
        if not code_value and data.class_element is not None:
            code_value = (text(jfind(data.class_element, "j:MilStd2525CCode")) or "").strip()
        if code_value:
            ET.SubElement(new_unit, jtag("MilStd2525CCode")).text = code_value
        if parent_lvc:
            ET.SubElement(new_unit, jtag("UnitSuperior")).text = parent_lvc
        if side_value:
            ET.SubElement(new_unit, jtag("SideSuperior")).text = side_value
        lvc_value = _gen_unique_lvcid("U", used_lvc)
        ET.SubElement(new_unit, jtag("LvcId")).text = lvc_value
        ET.SubElement(new_unit, jtag("Strength")).text = str(data.strength)
        own = ET.SubElement(new_unit, jtag("OwningFederate"))
        ET.SubElement(own, jtag("Federate")).text = data.owning_federate
        ET.SubElement(new_unit, jtag("IsInactive")).text = (
            "true" if data.is_inactive else "false"
        )
        ET.SubElement(new_unit, jtag("IsTransferrable")).text = (
            "true" if data.is_transferrable else "false"
        )
        ET.SubElement(new_unit, jtag("IsAggCommandable")).text = "false"
        if data.class_element is not None:
            self._copy_unitclass_enumeration(new_unit, data.class_element)
        return new_unit

    def _copy_unitclass_enumeration(
        self, unit_elem: ET._Element, class_elem: ET._Element
    ) -> None:
        ude = jfind(class_elem, "j:UnitDisEnumeration")
        if ude is None:
            return
        dest = ET.SubElement(unit_elem, jtag("UnitDisEnumeration"))
        for key, value in ude.attrib.items():
            if value is not None:
                dest.set(key, str(value))

    def _remove_unit_dis_enumeration(self, unit_elem: ET._Element) -> bool:
        ude = jfind(unit_elem, "j:UnitDisEnumeration")
        return self._safe_remove_element(unit_elem, ude)

    def _safe_remove_element(
        self, parent: ET._Element, child: Optional[ET._Element]
    ) -> bool:
        if child is None:
            return False
        try:
            parent.remove(child)
        except Exception:
            child.text = ""
            for sub in list(child):
                try:
                    child.remove(sub)
                except Exception:
                    continue
        return True

    def _existing_roles_for_unit(self, parent_unit: ET._Element) -> Set[str]:
        roles: Set[str] = set()
        if self._model is None:
            return roles
        parent_lvc = (text(jfind(parent_unit, "j:LvcId")) or "").strip()
        if not parent_lvc:
            return roles
        for ec in _iter_local(self._model.entcomp_list, "EntityComposition"):
            sup = (text(jfind(ec, "j:UnitSuperior")) or "").strip()
            if sup != parent_lvc:
                continue
            role = (text(jfind(ec, "j:Role")) or "").strip()
            if role:
                roles.add(role.upper())
        return roles

    def _collect_used_identifiers(self) -> Tuple[Set[str], Set[str]]:
        used_ids: Set[str] = set()
        used_lvc: Set[str] = set()
        if self._model is None:
            return used_ids, used_lvc
        for unit in _iter_local(self._model.unit_list, "Unit"):
            unit_id = unit.get("id")
            if unit_id:
                used_ids.add(unit_id.strip())
            lvc = (text(jfind(unit, "j:LvcId")) or "").strip()
            if lvc:
                used_lvc.add(lvc)
        for ec in _iter_local(self._model.entcomp_list, "EntityComposition"):
            eid = ec.get("id")
            if eid:
                used_ids.add(eid.strip())
            lvc = (text(jfind(ec, "j:LvcId")) or "").strip()
            if lvc:
                used_lvc.add(lvc)
        return used_ids, used_lvc

    def _ensure_unit_lvc(self, unit_elem: ET._Element, used_lvc: Set[str]) -> str:
        existing = (text(jfind(unit_elem, "j:LvcId")) or "").strip()
        if existing:
            used_lvc.add(existing)
            return existing
        new_value = _gen_unique_lvcid("U", used_lvc)
        ET.SubElement(unit_elem, jtag("LvcId")).text = new_value
        return new_value

    def _select_tree_item_for_unit(self, target: ET._Element) -> None:
        if target is None:
            return

        def visit(item: QTreeWidgetItem) -> bool:
            data = _normalize_user_data(item.data(0, Qt.ItemDataRole.UserRole))
            if isinstance(data, tuple) and data[0] == "unit" and data[1] is target:
                self.tree.setCurrentItem(item)
                self.tree.scrollToItem(item)
                return True
            for idx in range(item.childCount()):
                child = item.child(idx)
                if visit(child):
                    item.setExpanded(True)
                    return True
            return False

        for i in range(self.tree.topLevelItemCount()):
            if visit(self.tree.topLevelItem(i)):
                break

    def _create_entities_under_unit(
        self, parent_unit: ET._Element, data: _NewEntityData
    ) -> int:
        if self._model is None:
            return 0
        ent_list = self._model.entcomp_list
        if ent_list is None:
            ent_list = ET.SubElement(self._model.scenario, jtag("EntityCompositionList"))
            self._model.entcomp_list = ent_list
        used_ids, used_lvc = self._collect_used_identifiers()
        parent_lvc = self._ensure_unit_lvc(parent_unit, used_lvc)
        if not parent_lvc:
            return 0
        parent_side = (text(jfind(parent_unit, "j:SideSuperior")) or "").strip()
        created = 0
        existing_roles = self._existing_roles_for_unit(parent_unit)
        roles_used: Set[str] = set()
        class_name = data.class_name
        for _ in range(data.batch_count):
            unique_role = data.role
            if created > 0:
                base = f"{data.role}_{created + 1}"
                candidate = base
                suffix = 1
                while (
                    candidate.upper() in existing_roles
                    or candidate.upper() in roles_used
                ):
                    suffix += 1
                    candidate = f"{base}_{suffix}"
                unique_role = candidate
            unique_upper = unique_role.upper()
            while unique_upper in existing_roles or unique_upper in roles_used:
                unique_role = f"{unique_role}_{created+1}"
                unique_upper = unique_role.upper()
            ec = ET.SubElement(ent_list, jtag("EntityComposition"))
            ec.set("id", _gen_unit_xml_id(used_ids))
            lvc_value = _gen_unique_lvcid("U", used_lvc)
            ET.SubElement(ec, jtag("LvcId")).text = lvc_value
            ET.SubElement(ec, jtag("Role")).text = unique_role
            if data.quantity != 1:
                ET.SubElement(ec, jtag("Quantity")).text = str(data.quantity)
            ET.SubElement(ec, jtag("UnitSuperior")).text = parent_lvc
            if class_name:
                ET.SubElement(ec, jtag("ClassName")).text = class_name
            if parent_side:
                ET.SubElement(ec, jtag("SideSuperior")).text = parent_side
            ow = ET.SubElement(ec, jtag("OwningFederate"))
            ET.SubElement(ow, jtag("Federate")).text = data.owning_federate
            ET.SubElement(ec, jtag("IsInactive")).text = (
                "true" if data.is_inactive else "false"
            )
            ET.SubElement(ec, jtag("IsTransferrable")).text = (
                "true" if data.is_transferrable else "false"
            )
            ET.SubElement(ec, jtag("IsInvincible")).text = (
                "true" if data.is_invincible else "false"
            )
            created += 1
            roles_used.add(unique_upper)
            existing_roles.add(unique_upper)
        return created

    def _apply_side_to_subtree(self, unit_elem: ET._Element, side_value: str) -> int:
        _, parent_to_children = self._ensure_unit_maps()
        nodes = _gather_unit_descendants(
            unit_elem, parent_to_children, include_root=True
        )
        updated = 0
        for node in nodes:
            side_el = jfind(node, "j:SideSuperior")
            if not side_value:
                if side_el is not None:
                    try:
                        node.remove(side_el)
                    except Exception:
                        side_el.text = ""
                    updated += 1
                continue
            if side_el is None:
                side_el = ET.SubElement(node, jtag("SideSuperior"))
            if (side_el.text or "").strip() != side_value:
                side_el.text = side_value
                updated += 1
        return updated

    def _handle_consolidate_subtree(self, unit_elems) -> None:
        if self._model is None:
            return
        if isinstance(unit_elems, list):
            candidates = unit_elems
        elif isinstance(unit_elems, tuple):
            candidates = list(unit_elems)
        elif unit_elems is None:
            candidates = []
        else:
            candidates = [unit_elems]
        unique_units = []
        seen_ids = set()
        for unit in candidates:
            if unit is None:
                continue
            key = id(unit)
            if key in seen_ids:
                continue
            seen_ids.add(key)
            unique_units.append(unit)
        if not unique_units:
            return
        unit_targets = []
        skipped = []
        for unit in unique_units:
            lvc = (text(jfind(unit, "j:LvcId")) or "").strip()
            if not lvc:
                skipped.append(self._format_unit_label(unit))
                continue
            unit_targets.append((unit, lvc))
        if not unit_targets:
            if skipped:
                message = (
                    "The selected unit does not have an LvcId."
                    if len(skipped) == 1
                    else "The selected units do not have LvcIds."
                )
                QMessageBox.warning(self, "Consolidate Subtree", message)
            return
        if skipped:
            detail = "\n".join(skipped[:10])
            if len(skipped) > 10:
                detail += "\n..."
            QMessageBox.warning(
                self,
                "Consolidate Subtree",
                "The following units are missing an LvcId and will be skipped:\n"
                + detail,
            )
        levels = CONSOLIDATION_LEVELS
        if not levels:
            QMessageBox.warning(
                self, "Consolidate Subtree", "No consolidation levels are available."
            )
            return
        default_label = self._last_consolidation_label or levels[0]
        try:
            default_index = levels.index(default_label)
        except ValueError:
            default_index = 0
        selection, ok = QInputDialog.getItem(
            self,
            "Consolidate Subtree",
            "Consolidate lower echelons up to:",
            levels,
            default_index,
            False,
        )
        if not ok:
            return
        target_label = (selection or "").strip()
        if not target_label:
            return
        self._last_consolidation_label = target_label
        scope_roots = {lvc for _, lvc in unit_targets}
        stats = consolidate_units(self._model, target_label, scope_roots=scope_roots)
        total_changes = (
            stats["removed_units"]
            + stats["moved_entity_compositions"]
            + stats["relinked_units"]
        )
        if total_changes == 0:
            message = (
                "No changes were made for the selected unit."
                if len(scope_roots) == 1
                else "No changes were made for the selected units."
            )
            QMessageBox.information(self, "Consolidate Subtree", message)
            return
        skipped_units = stats.get("skipped_units", 0)
        parts = [
            f"removed units {stats['removed_units']}",
            f"moved entity compositions {stats['moved_entity_compositions']}",
            f"relinked units {stats['relinked_units']}",
        ]
        if skipped_units:
            parts.append(f"skipped {skipped_units}")
        if len(unit_targets) == 1:
            label = self._format_unit_label(unit_targets[0][0])
            summary = f"Consolidated '{label}': " + ", ".join(parts)
        else:
            preview = ", ".join(
                self._format_unit_label(unit) for unit, _ in unit_targets[:5]
            )
            if len(unit_targets) > 5:
                preview += ", ..."
            summary = (
                f"Consolidated {len(unit_targets)} units: {preview}\n\nChanges: "
                + ", ".join(parts)
            )
        QMessageBox.information(self, "Consolidate Subtree", summary)
        self.set_model(self._model)

    def _set_hidden_recursive(self, item: QTreeWidgetItem, hidden: bool) -> None:
        item.setHidden(hidden)
        for i in range(item.childCount()):
            self._set_hidden_recursive(item.child(i), hidden)

    def _filter_item_recursive(self, item: QTreeWidgetItem, term: str) -> bool:
        current_match = any(
            term in (item.text(col) or "").lower() for col in range(item.columnCount())
        )
        child_match = False
        for i in range(item.childCount()):
            child = item.child(i)
            if self._filter_item_recursive(child, term):
                child_match = True
            else:
                child.setHidden(True)
        visible = current_match or child_match
        item.setHidden(not visible)
        return visible

    def _handle_icon_tooltip(self, event: QHelpEvent) -> bool:
        item = self.tree.itemAt(event.pos())
        if item is None:
            QToolTip.hideText()
            return False
        column = self.tree.columnAt(event.pos().x())
        if column != 0:
            QToolTip.hideText()
            return False
        icon = item.icon(0)
        if icon.isNull():
            QToolTip.hideText()
            return False
        code = (item.text(1) or "").strip()
        large_icon = self._get_symbol_icon(code, size=self._hover_icon_px) or icon
        html = self._icon_to_tooltip_html(large_icon, self._hover_icon_px)
        if not html:
            QToolTip.hideText()
            return False
        rect = self.tree.visualItemRect(item)
        QToolTip.showText(event.globalPos(), html, self.tree, rect)
        event.accept()
        return True

    def _icon_to_tooltip_html(self, icon: QIcon, size: int) -> Optional[str]:
        pixmap = icon.pixmap(size, size)
        if pixmap.isNull():
            return None
        buffer = QBuffer()
        if not buffer.open(QIODevice.OpenModeFlag.WriteOnly):
            return None
        try:
            if not pixmap.save(buffer, "PNG"):
                return None
            encoded = bytes(buffer.data().toBase64()).decode("ascii")
        finally:
            buffer.close()
        return f"<img src='data:image/png;base64,{encoded}' />"

    def _ensure_side_item(
        self,
        side_key: str,
        side_labels: Dict[str, str],
        side_items: Dict[str, QTreeWidgetItem],
    ) -> QTreeWidgetItem:
        item = side_items.get(side_key)
        if item is not None:
            return item
        display = side_labels.get(side_key, side_key)
        if side_key == "UNKNOWN" and display == "UNKNOWN":
            display = "UNKNOWN / Unassigned"
        item = QTreeWidgetItem([display, ""])
        item.setData(0, Qt.ItemDataRole.UserRole, side_key)
        self.tree.addTopLevelItem(item)
        side_items[side_key] = item
        return item

    def _add_unit_recursive(
        self,
        unit: ET._Element,
        parent_item: QTreeWidgetItem,
        lvc_to_unit: Dict[str, ET._Element],
        parent_to_children: Dict[str, List[ET._Element]],
        ec_by_superior: Dict[str, List[ET._Element]],
        visited_units: Set[int],
    ) -> None:
        if id(unit) in visited_units:
            return
        visited_units.add(id(unit))
        label = self._format_unit_label(unit)
        code = (text(jfind(unit, "j:MilStd2525CCode")) or "").strip()
        item = QTreeWidgetItem([label, code])
        icon = self._get_symbol_icon(code)
        if icon is not None:
            item.setIcon(0, icon)
        item.setData(0, Qt.ItemDataRole.UserRole, ("unit", unit))
        parent_item.addChild(item)
        lvc = (text(jfind(unit, "j:LvcId")) or "").strip()
        for ec in sorted(ec_by_superior.get(lvc, []), key=self._entity_sort_key):
            self._add_entity_item(ec, item)
        for child in sorted(parent_to_children.get(lvc, []), key=self._unit_sort_key):
            self._add_unit_recursive(
                child,
                item,
                lvc_to_unit,
                parent_to_children,
                ec_by_superior,
                visited_units,
            )

    def _add_entity_item(self, ec: ET._Element, parent_item: QTreeWidgetItem) -> None:
        label = self._format_entity_label(ec)
        code = (text(jfind(ec, "j:MilStd2525CCode")) or "").strip()
        item = QTreeWidgetItem([label, code])
        item.setData(0, Qt.ItemDataRole.UserRole, ("entity", ec))
        parent_item.addChild(item)

    def _format_unit_label(self, unit: ET._Element) -> str:
        return format_unit_label(unit)

    def _format_entity_label(
        self,
        ec: ET._Element,
        *,
        include_superior: bool = False,
        superior_hint: Optional[str] = None,
    ) -> str:
        return format_entity_label(
            ec, include_superior=include_superior, superior_hint=superior_hint
        )

    def _get_symbol_icon(self, code: str, size: Optional[int] = None) -> Optional[QIcon]:
        code = (code or "").strip()
        if not code or self._symbol_icons_disabled:
            return None
        px = size if size is not None else self._icon_base_px
        try:
            px = max(16, int(px))
        except Exception:
            px = self._icon_base_px
        factory = self._symbol_icon_factory
        if factory is None:
            candidate = _SymbolIconFactory(icon_px=self._icon_base_px)
            if not candidate.is_ready:
                self._symbol_icons_disabled = True
                return None
            self._symbol_icon_factory = candidate
            factory = candidate
        return factory.icon_for(code, size=px)

    def _unit_sort_key(self, unit: ET._Element) -> Tuple[int, str, str]:
        return unit_sort_key(unit)

    def _entity_sort_key(self, ec: ET._Element) -> Tuple[str, str]:
        return entity_sort_key(ec)

    def _side_sort_key(self, side: str) -> Tuple[int, str]:
        return side_sort_key(side)

    def _collect_tree_rows(
        self, *, include_hidden: bool
    ) -> List[Tuple[int, List[str], str]]:
        columns = self.tree.columnCount()
        rows: List[Tuple[int, List[str], str]] = []
        root = self.tree.invisibleRootItem()

        def visit(parent: QTreeWidgetItem, level: int) -> None:
            for idx in range(parent.childCount()):
                child = parent.child(idx)
                if not include_hidden and child.isHidden():
                    continue
                texts = [child.text(col) for col in range(columns)]
                data = _normalize_user_data(child.data(0, Qt.ItemDataRole.UserRole))
                if isinstance(data, str):
                    kind = "side"
                elif isinstance(data, tuple) and data:
                    if data[0] == "unit":
                        kind = "unit"
                    elif data[0] == "entity":
                        kind = "entity"
                    else:
                        kind = "node"
                else:
                    kind = "group"
                rows.append((level, texts, kind))
                visit(child, level + 1)

        visit(root, 0)
        return rows

    def _write_preview_excel(
        self, path: str, rows: List[Tuple[int, List[str], str]]
    ) -> None:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font

        wb = Workbook()
        ws = wb.active
        ws.title = "Preview"
        headers = ["Hierarchy", "2525C"]
        ws.append(headers)
        header_font = Font(bold=True)
        for col_index, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_index)
            cell.font = header_font
            cell.alignment = Alignment(horizontal="left", vertical="center")
        row_count = len(rows)
        for idx, (level, texts, kind) in enumerate(rows):
            row_index = idx + 2
            label = texts[0] if texts else ""
            code = texts[1] if len(texts) > 1 else ""
            space_prefix = " " * max(level * 2, 0)
            display_label = f"{space_prefix}{label}" if label else ""
            cell_label = ws.cell(row=row_index, column=1, value=display_label)
            cell_code = ws.cell(row=row_index, column=2, value=code or None)
            indent = min(level, 15)
            cell_label.alignment = Alignment(indent=indent, vertical="center")
            cell_code.alignment = Alignment(vertical="center")
            next_level = rows[idx + 1][0] if idx + 1 < row_count else None
            has_children = next_level is not None and next_level > level
            font = None
            if kind == "side":
                font = Font(bold=True)
            elif kind == "unit" and has_children:
                font = Font(bold=True)
            elif kind == "entity":
                font = Font(italic=True)
            elif kind == "group":
                font = Font(italic=True)
            if font is not None:
                cell_label.font = font
            if level > 0:
                ws.row_dimensions[row_index].outlineLevel = min(level, 7)
        ws.column_dimensions["A"].width = 64
        ws.column_dimensions["B"].width = 18
        ws.freeze_panes = "A2"
        ws.sheet_view.showOutlineSymbols = True
        ws.sheet_properties.outlinePr.summaryBelow = False
        ws.sheet_properties.outlinePr.summaryRight = False
        ws.column_dimensions.group("B", "B", outline_level=1, hidden=False)
        wb.save(path)


@dataclass
class _MassRenameUnitRow:

    unit: ET._Element
    parent_index: Optional[int]
    depth: int
    prefix1: str
    prefix2: str
    prefix3: str
    prefix4: str
    original_name: str
    echelon: str
    side: str
    lvcid: str
    preview_name: str = ""


@dataclass
class _MassRenameEntityRow:

    entity: ET._Element
    unit_index: int
    original_role: str
    class_name: str
    quantity: str
    preview_name: str = ""


@dataclass
class _MassRenameResult:

    unit_names: Dict[ET._Element, str]
    entity_names: Dict[ET._Element, str]
    all_unit_names: Dict[ET._Element, str]
    all_entity_names: Dict[ET._Element, str]


class MassRenameDialog(QDialog):

    UNIT_TEMPLATE_DEFAULT = (
        "{prefix1}_{prefix2}_{prefix3}_{prefix4}_{suffix1}_{suffix2}_{suffix3}_{suffix4}"
    )
    ENTITY_TEMPLATE_DEFAULT = (
        "{class_slug}_{unit_prefix1}_{unit_prefix2}_{unit_prefix3}_{unit_suffix1}_{extra}"
    )

    def __init__(
        self,
        parent: Optional[QWidget],
        model: ObsModel,
        root_unit: ET._Element,
        *,
        existing_unit_names: Set[str],
        context_text: Optional[str] = None,
        unit_maps: Optional[
            Tuple[Dict[str, ET._Element], Dict[str, List[ET._Element]]]
        ] = None,
    ):
        super().__init__(parent)
        self._model = model
        self._root_unit = root_unit
        self._existing_unit_names = {name.upper() for name in existing_unit_names}
        self._unit_maps = unit_maps
        self._rows: List[_MassRenameUnitRow] = self._build_unit_rows()
        (
            self._entity_rows,
            self._entities_by_unit,
        ) = self._build_entity_rows()
        self._result: Optional[_MassRenameResult] = None
        self._block_table = False
        self._current_error: Optional[str] = None
        self._context_text = context_text
        self.setWindowTitle("Mass Rename Units and Entities")
        self.resize(920, 640)
        self._build_ui()
        self._refresh_previews()

    def result(self) -> Optional[_MassRenameResult]:
        return self._result

    def _build_ui(self) -> None:
        layout = QVBoxLayout(self)
        intro = self._context_text or (
            "Provide prefix values for each unit, adjust the template, and preview the "
            "result before applying the rename. Blank prefix or suffix tokens automatically "
            "drop extra underscores so names stay tidy."
        )
        help_label = QLabel(intro)
        help_label.setWordWrap(True)
        layout.addWidget(help_label)
        template_group = QGroupBox("Templates")
        form = QFormLayout(template_group)
        self.unit_template_edit = QLineEdit(self.UNIT_TEMPLATE_DEFAULT)
        self.unit_template_edit.textChanged.connect(self._refresh_previews)
        form.addRow("Unit template", self.unit_template_edit)
        self.unit_numbering_check = QCheckBox(
            "Always number duplicates (include {extra} starting at 1)"
        )
        self.unit_numbering_check.setChecked(False)
        self.unit_numbering_check.toggled.connect(self._refresh_previews)
        form.addRow("", self.unit_numbering_check)
        self.rename_entities_check = QCheckBox("Rename entity compositions")
        self.rename_entities_check.setChecked(bool(self._entity_rows))
        self.rename_entities_check.toggled.connect(self._refresh_previews)
        form.addRow("", self.rename_entities_check)
        self.entity_template_edit = QLineEdit(self.ENTITY_TEMPLATE_DEFAULT)
        self.entity_template_edit.textChanged.connect(self._refresh_previews)
        form.addRow("Entity template", self.entity_template_edit)
        self.entity_numbering_check = QCheckBox(
            "Always number duplicate entity names"
        )
        self.entity_numbering_check.setChecked(True)
        self.entity_numbering_check.toggled.connect(self._refresh_previews)
        form.addRow("", self.entity_numbering_check)
        layout.addWidget(template_group)
        token_info = QLabel(
            "Unit tokens: {prefix1}, {prefix2}, {prefix3}, {prefix4}, "
            "{suffix1}, {suffix2}, {suffix3}, {suffix4}, {echelon}, {side}, {lvcid}, "
            "{original_name}, {depth}, {index}, {extra}. "
            "Entity tokens: {role}, {class_name}, {class_slug}, {quantity}, {unit_name}, "
            "{unit_prefix1}, {unit_prefix2}, {unit_prefix3}, {unit_prefix4}, "
            "{unit_suffix1}, {unit_suffix2}, {unit_suffix3}, {unit_suffix4}, "
            "{unit_echelon}, {unit_original_name}, {unit_depth}, "
            "{unit_entity_index}, {extra}."
        )
        token_info.setWordWrap(True)
        layout.addWidget(token_info)
        self.table = QTableWidget()
        self.table.setColumnCount(8)
        self.table.setHorizontalHeaderLabels(
            [
                "Level",
                "Current Name",
                "Echelon",
                "Prefix 1",
                "Prefix 2",
                "Prefix 3",
                "Prefix 4",
                "Preview",
            ]
        )
        header = self.table.horizontalHeader()
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(7, QHeaderView.ResizeMode.Stretch)
        self.table.verticalHeader().setVisible(False)
        self.table.itemChanged.connect(self._handle_table_change)
        layout.addWidget(self.table)
        self._populate_unit_table()
        self.entity_preview = QTextEdit()
        self.entity_preview.setReadOnly(True)
        self.entity_preview.setMinimumHeight(120)
        layout.addWidget(self.entity_preview)
        self._status_label = QLabel("")
        self._status_label.setStyleSheet("color: red;")
        layout.addWidget(self._status_label)
        buttons = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)
        self._ok_button = buttons.button(QDialogButtonBox.StandardButton.Ok)

    def _populate_unit_table(self) -> None:
        self._block_table = True
        self.table.setRowCount(len(self._rows))
        for idx, row in enumerate(self._rows):
            level_item = QTableWidgetItem(str(row.depth))
            level_item.setFlags(level_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.table.setItem(idx, 0, level_item)
            indent = "  " * row.depth
            display_name = row.original_name or "Unnamed Unit"
            name_item = QTableWidgetItem(f"{indent}{display_name}")
            name_item.setFlags(name_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.table.setItem(idx, 1, name_item)
            echelon_item = QTableWidgetItem(row.echelon)
            echelon_item.setFlags(echelon_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.table.setItem(idx, 2, echelon_item)
            for col, value in zip(
                (3, 4, 5, 6),
                (row.prefix1, row.prefix2, row.prefix3, row.prefix4),
            ):
                item = QTableWidgetItem(value)
                self.table.setItem(idx, col, item)
            preview_item = QTableWidgetItem("")
            preview_item.setFlags(preview_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.table.setItem(idx, 7, preview_item)
        self._block_table = False

    def _handle_table_change(self, item: QTableWidgetItem) -> None:
        if self._block_table:
            return
        row_idx = item.row()
        col = item.column()
        if row_idx < 0 or row_idx >= len(self._rows):
            return
        if col not in (3, 4, 5, 6):
            return
        value = (item.text() or "").strip()
        target = self._rows[row_idx]
        if col == 3:
            target.prefix1 = value
        elif col == 4:
            target.prefix2 = value
        elif col == 5:
            target.prefix3 = value
        else:
            target.prefix4 = value
        self._refresh_previews()

    def _refresh_previews(self) -> None:
        self._current_error = None
        unit_template = (self.unit_template_edit.text() or "").strip()
        if not unit_template:
            unit_template = self.UNIT_TEMPLATE_DEFAULT
        try:
            contexts = []
            base_names = []
            for idx, row in enumerate(self._rows):
                context = self._unit_context(idx)
                contexts.append(context)
                base_value = self._format_template(unit_template, {**context, "extra": ""})
                if not base_value.strip():
                    raise ValueError("Unit template produced an empty name.")
                base_names.append(base_value.strip())
        except ValueError as exc:
            self._set_error(str(exc))
            return
        extras = self._assign_extras(
            base_names, force_numbering=self.unit_numbering_check.isChecked()
        )
        for idx, row in enumerate(self._rows):
            context = dict(contexts[idx])
            context["extra"] = extras[idx]
            try:
                preview = self._format_template(unit_template, context).strip()
            except ValueError as exc:
                self._set_error(str(exc))
                return
            row.preview_name = preview
            preview_item = self.table.item(idx, 7)
            if preview_item is not None:
                preview_item.setText(preview)
        if not self._update_entity_preview():
            return
        self._set_error(None)

    def _set_error(self, message: Optional[str]) -> None:
        self._current_error = message
        if message:
            self._status_label.setText(message)
            if hasattr(self, "_ok_button") and self._ok_button is not None:
                self._ok_button.setEnabled(False)
        else:
            self._status_label.setText("")
            if hasattr(self, "_ok_button") and self._ok_button is not None:
                self._ok_button.setEnabled(True)

    def _update_entity_preview(self) -> bool:
        if not self._entity_rows or not self.rename_entities_check.isChecked():
            for row in self._entity_rows:
                fallback = (
                    row.original_role
                    or self._slugify(row.class_name)
                    or row.class_name
                    or "ENTITY"
                )
                row.preview_name = fallback
            self.entity_preview.setPlainText("Entity rename disabled.")
            return True
        template = (self.entity_template_edit.text() or "").strip()
        if not template:
            template = self.ENTITY_TEMPLATE_DEFAULT
        lines: List[str] = []
        for unit_index, entity_indices in self._entities_by_unit.items():
            base_names: List[str] = []
            contexts = []
            rows = []
            for ordinal, entity_idx in enumerate(entity_indices, start=1):
                row = self._entity_rows[entity_idx]
                rows.append(row)
                context = self._entity_context(row, unit_index, ordinal)
                contexts.append(context)
                try:
                    candidate = self._format_template(
                        template, {**context, "extra": ""}
                    )
                except ValueError as exc:
                    self._set_error(str(exc))
                    return False
                fallback = candidate or (
                    self._slugify(row.class_name)
                    or self._slugify(row.original_role)
                    or row.original_role
                    or "ENTITY"
                )
                base_names.append(fallback)
            extras = self._assign_extras(
                base_names, force_numbering=self.entity_numbering_check.isChecked()
            )
            for ctx, extra, row, base_name in zip(contexts, extras, rows, base_names):
                ctx = dict(ctx)
                ctx["extra"] = extra
                try:
                    preview = self._format_template(template, ctx).strip()
                except ValueError as exc:
                    self._set_error(str(exc))
                    return False
                if not preview:
                    preview = base_name
                row.preview_name = preview
                unit_display = self._rows[unit_index].preview_name or (
                    self._rows[unit_index].original_name or "Unit"
                )
                lines.append(f"{unit_display} -> {row.original_role} => {preview}")
        self.entity_preview.setPlainText("\n".join(lines) if lines else "No entities.")
        return True

    def accept(self) -> None:
        if self._current_error:
            QMessageBox.warning(self, "Mass Rename", self._current_error)
            return
        proposed_units: Dict[ET._Element, str] = {}
        all_units: Dict[ET._Element, str] = {}
        used = set(self._existing_unit_names)
        for row in self._rows:
            new_name = (row.preview_name or "").strip()
            if not new_name:
                QMessageBox.warning(
                    self, "Mass Rename", "All units must have a non-empty name."
                )
                return
            upper = new_name.upper()
            if upper in used:
                QMessageBox.warning(
                    self,
                    "Mass Rename",
                    f"Name '{new_name}' conflicts with another unit in the model.",
                )
                return
            used.add(upper)
            all_units[row.unit] = new_name
            if new_name != (row.original_name or ""):
                proposed_units[row.unit] = new_name
        proposed_entities: Dict[ET._Element, str] = {}
        all_entities: Dict[ET._Element, str] = {}
        if self._entity_rows and self.rename_entities_check.isChecked():
            for unit_index, entity_indices in self._entities_by_unit.items():
                seen: Set[str] = set()
                for entity_idx in entity_indices:
                    row = self._entity_rows[entity_idx]
                    new_role = (row.preview_name or "").strip()
                    if not new_role:
                        QMessageBox.warning(
                            self,
                            "Mass Rename",
                            "All renamed entity compositions must have a non-empty role.",
                        )
                        return
                    upper = new_role.upper()
                    if upper in seen:
                        QMessageBox.warning(
                            self,
                            "Mass Rename",
                            "Duplicate entity role detected under the same unit. "
                            "Adjust the template or prefixes to ensure uniqueness.",
                        )
                        return
                    seen.add(upper)
                    all_entities[row.entity] = new_role
                    if new_role != (row.original_role or ""):
                        proposed_entities[row.entity] = new_role
        else:
            for row in self._entity_rows:
                fallback = (
                    (row.preview_name or "").strip()
                    or row.original_role
                    or self._slugify(row.class_name)
                    or row.class_name
                    or "ENTITY"
                )
                all_entities[row.entity] = fallback
        self._result = _MassRenameResult(
            proposed_units, proposed_entities, all_units, all_entities
        )
        super().accept()

    def _unit_context(self, index: int) -> Dict[str, Any]:
        row = self._rows[index]
        parent_prefix = ("", "", "", "")
        if row.parent_index is not None and 0 <= row.parent_index < len(self._rows):
            parent = self._rows[row.parent_index]
            parent_prefix = (
                parent.prefix1,
                parent.prefix2,
                parent.prefix3,
                parent.prefix4,
            )
        return {
            "prefix1": row.prefix1,
            "prefix2": row.prefix2,
            "prefix3": row.prefix3,
            "prefix4": row.prefix4,
            "suffix1": parent_prefix[0],
            "suffix2": parent_prefix[1],
            "suffix3": parent_prefix[2],
            "suffix4": parent_prefix[3],
            "original_name": row.original_name,
            "echelon": row.echelon,
            "side": row.side,
            "lvcid": row.lvcid,
            "depth": row.depth,
            "index": index + 1,
            "unit_name": row.preview_name or row.original_name,
        }

    def _entity_context(
        self, row: _MassRenameEntityRow, unit_index: int, ordinal: int
    ) -> Dict[str, Any]:
        unit_row = self._rows[unit_index]
        parent_prefix = ("", "", "", "")
        if (
            unit_row.parent_index is not None
            and 0 <= unit_row.parent_index < len(self._rows)
        ):
            parent = self._rows[unit_row.parent_index]
            parent_prefix = (
                parent.prefix1,
                parent.prefix2,
                parent.prefix3,
                parent.prefix4,
            )
        return {
            "role": row.original_role,
            "class_name": row.class_name,
            "class_slug": self._slugify(row.class_name),
            "quantity": row.quantity,
            "unit_name": unit_row.preview_name or unit_row.original_name,
            "unit_prefix1": unit_row.prefix1,
            "unit_prefix2": unit_row.prefix2,
            "unit_prefix3": unit_row.prefix3,
            "unit_prefix4": unit_row.prefix4,
            "unit_suffix1": parent_prefix[0],
            "unit_suffix2": parent_prefix[1],
            "unit_suffix3": parent_prefix[2],
            "unit_suffix4": parent_prefix[3],
            "unit_echelon": unit_row.echelon,
            "unit_original_name": unit_row.original_name,
            "unit_depth": unit_row.depth,
            "unit_entity_index": ordinal,
        }

    def _build_unit_rows(self) -> List[_MassRenameUnitRow]:
        rows: List[_MassRenameUnitRow] = []
        if self._model is None or self._model.unit_list is None:
            return rows
        if self._unit_maps is not None:
            _, parent_to_children = self._unit_maps
        else:
            _, parent_to_children = collect_unit_maps(self._model)
        queue: List[Tuple[ET._Element, Optional[int], int]] = [
            (self._root_unit, None, 0)
        ]
        seen: Set[int] = set()
        while queue:
            unit, parent_index, depth = queue.pop(0)
            key = id(unit)
            if key in seen:
                continue
            seen.add(key)
            name = (text(jfind(unit, "j:Name")) or "").strip()
            echelon = (text(jfind(unit, "j:Echelon")) or "").strip()
            side = (text(jfind(unit, "j:SideSuperior")) or "").strip()
            lvcid = (text(jfind(unit, "j:LvcId")) or "").strip()
            prefix = self._guess_prefix_tokens(name)
            row = _MassRenameUnitRow(
                unit=unit,
                parent_index=parent_index,
                depth=depth,
                prefix1=prefix[0],
                prefix2=prefix[1],
                prefix3=prefix[2],
                prefix4=prefix[3],
                original_name=name,
                echelon=echelon,
                side=side,
                lvcid=lvcid,
            )
            rows.append(row)
            current_index = len(rows) - 1
            children = []
            if lvcid:
                children = parent_to_children.get(lvcid, [])
            for child in sorted(children, key=unit_sort_key):
                queue.append((child, current_index, depth + 1))
        return rows

    def _build_entity_rows(
        self,
    ) -> Tuple[List[_MassRenameEntityRow], Dict[int, List[int]]]:
        if self._model is None or self._model.entcomp_list is None:
            return [], {}
        lvc_to_index: Dict[str, int] = {}
        for idx, row in enumerate(self._rows):
            if row.lvcid:
                lvc_to_index[row.lvcid] = idx
        buckets: Dict[int, List[ET._Element]] = defaultdict(list)
        for ec in _iter_local(self._model.entcomp_list, "EntityComposition"):
            sup = (text(jfind(ec, "j:UnitSuperior")) or "").strip()
            if not sup:
                continue
            owner_index = lvc_to_index.get(sup)
            if owner_index is None:
                continue
            buckets[owner_index].append(ec)
        rows: List[_MassRenameEntityRow] = []
        mapping: Dict[int, List[int]] = defaultdict(list)
        for owner_index in range(len(self._rows)):
            entries = buckets.get(owner_index)
            if not entries:
                continue
            for ec in sorted(entries, key=self._entity_sort_key_for_row):
                role = (text(jfind(ec, "j:Role")) or "").strip()
                class_name = (text(jfind(ec, "j:ClassName")) or "").strip()
                quantity = (text(jfind(ec, "j:Quantity")) or "").strip()
                row = _MassRenameEntityRow(
                    entity=ec,
                    unit_index=owner_index,
                    original_role=role,
                    class_name=class_name,
                    quantity=quantity,
                )
                mapping[owner_index].append(len(rows))
                rows.append(row)
        return rows, mapping

    def _entity_sort_key_for_row(self, ec: ET._Element) -> Tuple[str, str]:
        role = (text(jfind(ec, "j:Role")) or "").strip().lower()
        lvc = (text(jfind(ec, "j:LvcId")) or "").strip()
        return (role, lvc)

    @staticmethod
    def _assign_extras(
        names: List[str], *, force_numbering: bool
    ) -> List[str]:
        totals: Dict[str, int] = defaultdict(int)
        for name in names:
            totals[name] += 1
        counters: Dict[str, int] = defaultdict(int)
        extras: List[str] = []
        for name in names:
            counters[name] += 1
            total = totals[name]
            order = counters[name]
            if total == 1 and not force_numbering:
                extras.append("")
            elif order == 1 and not force_numbering:
                extras.append("")
            else:
                extras.append(str(order))
        return extras

    @staticmethod
    def _guess_prefix_tokens(name: str) -> Tuple[str, str, str, str]:
        cleaned = name.replace("-", "_") if name else ""
        tokens = [tok for tok in cleaned.split("_") if tok]
        p1 = tokens[0] if len(tokens) > 0 else ""
        p2 = tokens[1] if len(tokens) > 1 else ""
        p3 = tokens[2] if len(tokens) > 2 else ""
        p4 = tokens[3] if len(tokens) > 3 else ""
        return (p1, p2, p3, p4)

    @staticmethod
    def _format_template(template: str, context: Dict[str, Any]) -> str:
        formatter = string.Formatter()
        safe = defaultdict(str, context)  # type: ignore[arg-type]
        try:
            rendered = formatter.vformat(template, (), safe)
        except Exception as exc:
            raise ValueError(str(exc)) from exc
        return MassRenameDialog._cleanup_value(rendered)

    @staticmethod
    def _cleanup_value(value: str) -> str:
        if not value:
            return ""
        text = value.strip()
        text = re.sub(r"\s+", " ", text)
        text = re.sub(r"_+", "_", text)
        text = text.strip("_ ").strip()
        return text

    @staticmethod
    def _slugify(value: Optional[str]) -> str:
        if not value:
            return ""
        text = value.strip()
        if not text:
            return ""
        text = re.sub(r"[^0-9A-Za-z]+", "_", text)
        return text.strip("_")

class ObsUnitSelectionDialog(QDialog):

    def __init__(self, parent: Optional[QWidget], model: ObsModel):
        super().__init__(parent)
        self._model = model
        self._selected_roots: Set[ET._Element] = set()
        self.setWindowTitle("Select Units to Import")
        self.resize(680, 540)
        layout = QVBoxLayout(self)
        layout.addWidget(
            QLabel("Choose the root units to merge into the current model.")
        )
        self.tree = QTreeWidget()
        self.tree.setAlternatingRowColors(True)
        self.tree.setColumnCount(3)
        self.tree.setHeaderLabels(["Unit", "LvcId", "2525C"])
        self.tree.setRootIsDecorated(True)
        self.tree.setUniformRowHeights(True)
        self.tree.itemChanged.connect(self._on_item_changed)
        layout.addWidget(self.tree)
        ctrl_row = QHBoxLayout()
        btn_all = QPushButton("Select All")
        btn_none = QPushButton("Select None")
        btn_expand = QPushButton("Expand All")
        btn_collapse = QPushButton("Collapse All")
        btn_all.clicked.connect(lambda: self._set_all(Qt.CheckState.Checked))
        btn_none.clicked.connect(lambda: self._set_all(Qt.CheckState.Unchecked))
        btn_expand.clicked.connect(self.tree.expandAll)
        btn_collapse.clicked.connect(self.tree.collapseAll)
        for btn in (btn_all, btn_none, btn_expand, btn_collapse):
            ctrl_row.addWidget(btn)
        ctrl_row.addStretch()
        layout.addLayout(ctrl_row)
        buttons = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        buttons.accepted.connect(self._on_accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)
        self._populate_tree()
        self.tree.expandToDepth(0)

    def _populate_tree(self) -> None:
        model = self._model
        if model is None:
            return
        lvc_to_unit, parent_to_children = collect_unit_maps(model)
        self._parent_to_children = parent_to_children
        self._visited_units: Set[int] = set()
        units = list(_iter_local(model.unit_list, "Unit"))
        side_labels: Dict[str, str] = {}

        def unit_side(unit: ET._Element) -> str:
            raw = (text(jfind(unit, "j:SideSuperior")) or "").strip()
            if raw:
                side_labels.setdefault(raw.upper(), raw)
                return raw.upper()
            derived = _resolve_unit_side(unit, lvc_to_unit)
            if derived:
                side_labels.setdefault(derived.upper(), derived)
                return derived.upper()
            side_labels.setdefault("UNKNOWN", "UNKNOWN")
            return "UNKNOWN"

        side_roots: Dict[str, List[ET._Element]] = {}
        for unit in units:
            superior = (text(jfind(unit, "j:UnitSuperior")) or "").strip()
            if superior and superior in lvc_to_unit:
                continue
            key = unit_side(unit)
            side_roots.setdefault(key, []).append(unit)
        side_items: Dict[str, QTreeWidgetItem] = {}
        self.tree.blockSignals(True)
        for side_key in sorted(side_roots.keys(), key=side_sort_key):
            display = side_labels.get(side_key, side_key)
            side_item = QTreeWidgetItem([display, "", ""])
            side_item.setFlags(side_item.flags() & ~Qt.ItemFlag.ItemIsUserCheckable)
            self.tree.addTopLevelItem(side_item)
            side_items[side_key] = side_item
            roots = sorted(side_roots.get(side_key, []), key=unit_sort_key)
            for unit in roots:
                self._add_unit_item(side_item, unit)
        # Catch any units that did not appear due to missing superior references
        remaining = [u for u in units if id(u) not in self._visited_units]
        if remaining:
            side_item = side_items.get("UNKNOWN")
            if side_item is None:
                side_item = QTreeWidgetItem(["UNKNOWN / Unassigned", "", ""])
                side_item.setFlags(side_item.flags() & ~Qt.ItemFlag.ItemIsUserCheckable)
                self.tree.addTopLevelItem(side_item)
            placeholder = QTreeWidgetItem(["Detached Units", "", ""])
            placeholder.setFlags(placeholder.flags() & ~Qt.ItemFlag.ItemIsUserCheckable)
            side_item.addChild(placeholder)
            for unit in sorted(remaining, key=unit_sort_key):
                self._add_unit_item(placeholder, unit)
        self.tree.blockSignals(False)
        self._refresh_parent_states()

    def _add_unit_item(self, parent_item: QTreeWidgetItem, unit: ET._Element) -> None:
        if id(unit) in self._visited_units:
            return
        self._visited_units.add(id(unit))
        label = format_unit_label(unit)
        lvc = (text(jfind(unit, "j:LvcId")) or "").strip()
        code = (text(jfind(unit, "j:MilStd2525CCode")) or "").strip()
        item = QTreeWidgetItem([label, lvc, code])
        flags = (
            item.flags()
            | Qt.ItemFlag.ItemIsUserCheckable
            | Qt.ItemFlag.ItemIsSelectable
            | Qt.ItemFlag.ItemIsAutoTristate
        )
        item.setFlags(flags)
        item.setCheckState(0, Qt.CheckState.Checked)
        item.setData(0, Qt.ItemDataRole.UserRole, unit)
        parent_item.addChild(item)
        children = []
        if lvc:
            children = sorted(self._parent_to_children.get(lvc, []), key=unit_sort_key)
        for child in children:
            self._add_unit_item(item, child)

    def _set_all(self, state: Qt.CheckState) -> None:
        self.tree.blockSignals(True)

        def set_item(item: QTreeWidgetItem) -> None:
            if item.flags() & Qt.ItemFlag.ItemIsUserCheckable:
                item.setCheckState(0, state)
            for idx in range(item.childCount()):
                set_item(item.child(idx))

        for i in range(self.tree.topLevelItemCount()):
            set_item(self.tree.topLevelItem(i))
        self.tree.blockSignals(False)
        self._refresh_parent_states()

    def _on_item_changed(self, item: QTreeWidgetItem, column: int) -> None:
        if column != 0:
            return
        state = item.checkState(0)
        self.tree.blockSignals(True)
        for idx in range(item.childCount()):
            item.child(idx).setCheckState(0, state)
        self.tree.blockSignals(False)
        self._update_parent_state(item)

    def _update_parent_state(self, item: QTreeWidgetItem) -> None:
        parent = item.parent()
        while parent is not None:
            checked = unchecked = partial = 0
            for idx in range(parent.childCount()):
                child_state = parent.child(idx).checkState(0)
                if child_state == Qt.CheckState.Checked:
                    checked += 1
                elif child_state == Qt.CheckState.Unchecked:
                    unchecked += 1
                else:
                    partial += 1
            self.tree.blockSignals(True)
            if partial > 0 or (checked > 0 and unchecked > 0):
                parent.setCheckState(0, Qt.CheckState.PartiallyChecked)
            elif checked == parent.childCount():
                parent.setCheckState(0, Qt.CheckState.Checked)
            else:
                parent.setCheckState(0, Qt.CheckState.Unchecked)
            self.tree.blockSignals(False)
            parent = parent.parent()

    def _refresh_parent_states(self) -> None:
        def refresh(item: QTreeWidgetItem) -> Qt.CheckState:
            child_states = []
            for idx in range(item.childCount()):
                child_states.append(refresh(item.child(idx)))
            if not child_states:
                return item.checkState(0)
            if all(state == Qt.CheckState.Checked for state in child_states):
                state = Qt.CheckState.Checked
            elif all(state == Qt.CheckState.Unchecked for state in child_states):
                state = Qt.CheckState.Unchecked
            else:
                state = Qt.CheckState.PartiallyChecked
            self.tree.blockSignals(True)
            item.setCheckState(0, state)
            self.tree.blockSignals(False)
            return state

        for i in range(self.tree.topLevelItemCount()):
            refresh(self.tree.topLevelItem(i))

    def _collect_roots(self, item: QTreeWidgetItem, out: Set[ET._Element]) -> None:
        data = _normalize_user_data(item.data(0, Qt.ItemDataRole.UserRole))
        state = item.checkState(0)
        parent = item.parent()
        parent_checked = (
            parent.checkState(0) if parent is not None else Qt.CheckState.Unchecked
        )
        parent_data = (
            _normalize_user_data(parent.data(0, Qt.ItemDataRole.UserRole))
            if parent is not None
            else None
        )
        if isinstance(data, ET._Element) and state == Qt.CheckState.Checked:
            if (
                parent is None
                or parent_data is None
                or parent_checked != Qt.CheckState.Checked
            ):
                out.add(data)
                return
        for idx in range(item.childCount()):
            self._collect_roots(item.child(idx), out)

    def _on_accept(self) -> None:
        selected: Set[ET._Element] = set()
        for i in range(self.tree.topLevelItemCount()):
            self._collect_roots(self.tree.topLevelItem(i), selected)
        if not selected:
            QMessageBox.warning(
                self, "Import Units", "Select at least one unit to import."
            )
            return
        self._selected_roots = selected
        super().accept()

    def selected_root_units(self) -> Set[ET._Element]:
        return set(self._selected_roots)


def merge_obs_units(
    target_model: ObsModel,
    source_model: ObsModel,
    root_units: Iterable[ET._Element],
    *,
    override_side: Optional[str] = None,
    fallback_side: Optional[str] = None,
) -> Dict[str, Any]:

    if target_model is None:
        raise ValueError("Target model is not initialized")
    if source_model is None:
        raise ValueError("Source model is not initialized")
    normalized_override = (override_side or "").strip() or None
    normalized_fallback = (fallback_side or "").strip() or None
    stats: Dict[str, Any] = {
        "units_added": 0,
        "entities_added": 0,
        "unit_lvc_conflicts": 0,
        "entity_lvc_conflicts": 0,
        "unit_superiors_cleared": 0,
        "unit_superiors_updated": 0,
        "sides_forced": 0,
        "sides_filled": 0,
        "ecc_added": 0,
        "ecc_skipped": 0,
        "ucl_added": 0,
        "ucl_skipped": 0,
        "sides_added": 0,
        "unit_ids_assigned": 0,
        "unit_timestamps_assigned": 0,
        "unit_sides_assigned": 0,
        "entities_skipped_orphaned": 0,
        "new_units": [],
    }
    root_list: List[ET._Element] = []
    seen_roots: Set[int] = set()
    for unit in root_units:
        if unit is None:
            continue
        key = id(unit)
        if key in seen_roots:
            continue
        seen_roots.add(key)
        root_list.append(unit)
    if not root_list:
        return stats

    # Merge class lists first so that references remain valid.
    def _merge_classes(tag: str, key_fn) -> Tuple[int, int]:
        src_parent = jfind(source_model.classdata, f"j:{tag}")
        if src_parent is None:
            return 0, 0
        tgt_parent = jfind(target_model.classdata, f"j:{tag}")
        if tgt_parent is None:
            tgt_parent = ET.SubElement(target_model.classdata, jtag(tag))
        existing_keys: Set[str] = set()
        for elem in list(tgt_parent):
            key = key_fn(elem)
            if key:
                existing_keys.add(key)
        added = skipped = 0
        for elem in list(src_parent):
            key = key_fn(elem)
            if not key:
                continue
            if key in existing_keys:
                skipped += 1
                continue
            tgt_parent.append(clone_element(elem))
            existing_keys.add(key)
            added += 1
        return added, skipped

    def _ecc_key(elem: ET._Element) -> Optional[str]:
        return (text(jfind(elem, "j:Name")) or "").strip().upper() or None

    def _ucl_key(elem: ET._Element) -> Optional[str]:
        return (text(jfind(elem, "j:AggregateName")) or "").strip().upper() or None

    ecc_added, ecc_skipped = _merge_classes("EntityCompositionClassList", _ecc_key)
    ucl_added, ucl_skipped = _merge_classes("UnitClassList", _ucl_key)
    stats["ecc_added"] = ecc_added
    stats["ecc_skipped"] = ecc_skipped
    stats["ucl_added"] = ucl_added
    stats["ucl_skipped"] = ucl_skipped
    used_lvc_ids: Set[str] = set()
    used_xml_ids: Set[str] = set()
    existing_unit_lvc: Dict[str, ET._Element] = {}
    for unit in _iter_local(target_model.unit_list, "Unit"):
        lvc = (text(jfind(unit, "j:LvcId")) or "").strip()
        if lvc:
            used_lvc_ids.add(lvc)
            existing_unit_lvc[lvc] = unit
        unit_id = unit.get("id")
        if unit_id:
            used_xml_ids.add(unit_id)
    for entity in _iter_local(target_model.entcomp_list, "EntityComposition"):
        lvc = (text(jfind(entity, "j:LvcId")) or "").strip()
        if lvc:
            used_lvc_ids.add(lvc)
        entity_id = entity.get("id")
        if entity_id:
            used_xml_ids.add(entity_id)
    lvc_to_unit_src, parent_to_children_src = collect_unit_maps(source_model)
    ordered_units: List[ET._Element] = []
    visited_units: Set[int] = set()

    def visit(unit: ET._Element) -> None:
        key = id(unit)
        if key in visited_units:
            return
        visited_units.add(key)
        ordered_units.append(unit)
        lvc = (text(jfind(unit, "j:LvcId")) or "").strip()
        children = []
        if lvc:
            children = parent_to_children_src.get(lvc, [])
        for child in sorted(children, key=unit_sort_key):
            visit(child)

    for unit in sorted(root_list, key=unit_sort_key):
        visit(unit)
    selected_unit_lvc: Set[str] = set()
    for unit in ordered_units:
        lvc = (text(jfind(unit, "j:LvcId")) or "").strip()
        if lvc:
            selected_unit_lvc.add(lvc)
    entities_to_copy: List[ET._Element] = []
    if selected_unit_lvc:
        for entity in _iter_local(source_model.entcomp_list, "EntityComposition"):
            sup = (text(jfind(entity, "j:UnitSuperior")) or "").strip()
            if sup and sup in selected_unit_lvc:
                entities_to_copy.append(entity)
    unit_lvc_map: Dict[str, str] = {}
    new_unit_elements: List[ET._Element] = []
    imported_sides: Set[str] = set()
    override_upper = (
        normalized_override.strip().upper() if normalized_override else None
    )
    fallback_upper = (
        normalized_fallback.strip().upper() if normalized_fallback else None
    )
    for unit in ordered_units:
        new_unit = clone_element(unit)
        orig_lvc = (text(jfind(unit, "j:LvcId")) or "").strip()
        lvc_elem = jfind(new_unit, "j:LvcId")
        new_lvc = orig_lvc
        if not new_lvc or new_lvc in used_lvc_ids:
            if orig_lvc and orig_lvc in used_lvc_ids:
                stats["unit_lvc_conflicts"] += 1
            new_lvc = _gen_unique_lvcid("U", used_lvc_ids)
            if lvc_elem is None:
                lvc_elem = ET.SubElement(new_unit, jtag("LvcId"))
            lvc_elem.text = new_lvc
        else:
            used_lvc_ids.add(new_lvc)
        if orig_lvc:
            unit_lvc_map[orig_lvc] = new_lvc
        else:
            unit_lvc_map[new_lvc] = new_lvc
        unit_id = new_unit.get("id")
        if not _is_valid_unit_xml_id(unit_id) or unit_id in used_xml_ids:
            unit_id = _gen_unit_xml_id(used_xml_ids)
            new_unit.set("id", unit_id)
        used_xml_ids.add(unit_id)
        stamp = new_unit.get("dateTimeStamp")
        if not _is_valid_timestamp(stamp):
            new_unit.set("dateTimeStamp", _now_iso_seconds_no_tz())
        sup_el = jfind(new_unit, "j:UnitSuperior")
        sup_val = (
            (sup_el.text or "").strip() if sup_el is not None and sup_el.text else ""
        )
        if sup_val:
            if sup_val in unit_lvc_map:
                new_sup = unit_lvc_map[sup_val]
                if new_sup != sup_val:
                    stats["unit_superiors_updated"] += 1
                if sup_el is None:
                    sup_el = ET.SubElement(new_unit, jtag("UnitSuperior"))
                sup_el.text = new_sup
            elif sup_val in existing_unit_lvc:
                new_sup = sup_val
                if sup_el is None:
                    sup_el = ET.SubElement(new_unit, jtag("UnitSuperior"))
                sup_el.text = new_sup
            else:
                stats["unit_superiors_cleared"] += 1
                if sup_el is not None:
                    sup_el.text = ""
        # Apply side policies
        side_el = jfind(new_unit, "j:SideSuperior")
        current_side = (
            (side_el.text or "").strip() if side_el is not None and side_el.text else ""
        )
        assigned_side = current_side
        if normalized_override:
            assigned_side = normalized_override
            if side_el is None:
                side_el = ET.SubElement(new_unit, jtag("SideSuperior"))
            if side_el.text != normalized_override:
                side_el.text = normalized_override
                stats["sides_forced"] += 1
        elif not current_side and normalized_fallback:
            assigned_side = normalized_fallback
            if side_el is None:
                side_el = ET.SubElement(new_unit, jtag("SideSuperior"))
            side_el.text = normalized_fallback
            stats["sides_filled"] += 1
        if assigned_side:
            imported_sides.add(assigned_side)
        target_model.unit_list.append(new_unit)
        stats["units_added"] += 1
        new_unit_elements.append(new_unit)
        existing_unit_lvc[new_lvc] = new_unit
    entity_added = 0
    entity_lvc_map: Dict[str, str] = {}
    for entity in entities_to_copy:
        new_entity = clone_element(entity)
        orig_lvc = (text(jfind(entity, "j:LvcId")) or "").strip()
        lvc_elem = jfind(new_entity, "j:LvcId")
        new_lvc = orig_lvc
        if not new_lvc or new_lvc in used_lvc_ids:
            if orig_lvc and orig_lvc in used_lvc_ids:
                stats["entity_lvc_conflicts"] += 1
            new_lvc = _gen_unique_lvcid("E", used_lvc_ids)
            if lvc_elem is None:
                lvc_elem = ET.SubElement(new_entity, jtag("LvcId"))
            lvc_elem.text = new_lvc
        else:
            used_lvc_ids.add(new_lvc)
        if orig_lvc:
            entity_lvc_map[orig_lvc] = new_lvc
        entity_id = new_entity.get("id")
        if not entity_id or entity_id in used_xml_ids:
            entity_id = _gen_unit_xml_id(used_xml_ids)
            new_entity.set("id", entity_id)
        used_xml_ids.add(entity_id)
        sup_el = jfind(new_entity, "j:UnitSuperior")
        sup_val = (
            (sup_el.text or "").strip() if sup_el is not None and sup_el.text else ""
        )
        if sup_val:
            if sup_val in unit_lvc_map:
                new_sup = unit_lvc_map[sup_val]
            elif sup_val in existing_unit_lvc:
                new_sup = sup_val
            else:
                new_sup = ""
            if not new_sup:
                stats["entities_skipped_orphaned"] += 1
                continue
            if sup_el is None:
                sup_el = ET.SubElement(new_entity, jtag("UnitSuperior"))
            sup_el.text = new_sup
        target_model.entcomp_list.append(new_entity)
        entity_added += 1
    stats["entities_added"] = entity_added
    if imported_sides:
        stats["sides_added"] = ensure_side_entries(target_model, imported_sides)
    if new_unit_elements:
        ids_assigned, stamps_assigned, sides_assigned = ensure_unit_metadata(
            target_model,
            normalized_override or normalized_fallback,
            set(new_unit_elements),
        )
        stats["unit_ids_assigned"] = ids_assigned
        stats["unit_timestamps_assigned"] = stamps_assigned
        stats["unit_sides_assigned"] = sides_assigned
    stats["new_units"] = new_unit_elements
    return stats


class ObsMoveUnitDialog(QDialog):

    def __init__(
        self,
        parent: Optional[QWidget],
        model: ObsModel,
        *,
        blocked_units: Optional[Set[ET._Element]] = None,
        allow_side_selection: bool = True,
    ):
        super().__init__(parent)
        self._model = model
        self._allow_side_selection = allow_side_selection
        self._result: Optional[Tuple[str, Any]] = None
        self._blocked: Set[int] = {id(unit) for unit in (blocked_units or set())}
        self.setWindowTitle("Select New Parent")
        self.resize(640, 520)
        layout = QVBoxLayout(self)
        layout.addWidget(QLabel("Choose the destination for the selected item."))
        self.tree = QTreeWidget()
        self.tree.setAlternatingRowColors(True)
        self.tree.setColumnCount(3)
        self.tree.setHeaderLabels(["Unit / Side", "LvcId", "2525C"])
        self.tree.setSelectionMode(QTreeWidget.SelectionMode.SingleSelection)
        layout.addWidget(self.tree)
        self._status = QLabel("")
        layout.addWidget(self._status)
        buttons = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        buttons.accepted.connect(self._on_accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)
        if model is not None:
            self._build_tree(blocked_units or set())
            self.tree.expandToDepth(1)
            self.tree.currentItemChanged.connect(self._on_current_item_changed)

    def _collect_blocked_descendants(self, blocked_units: Set[ET._Element]) -> None:
        if self._model is None or not blocked_units:
            return
        _, parent_to_children = collect_unit_maps(self._model)
        for unit in blocked_units:
            for desc in _gather_unit_descendants(
                unit, parent_to_children, include_root=True
            ):
                self._blocked.add(id(desc))

    def _build_tree(self, blocked_units: Set[ET._Element]) -> None:
        self.tree.clear()
        self._collect_blocked_descendants(blocked_units)
        model = self._model
        if model is None:
            return
        id_to_unit, parent_to_children = collect_unit_maps(model)
        side_labels: Dict[str, str] = {}

        def unit_side(unit: ET._Element) -> str:
            raw = (text(jfind(unit, "j:SideSuperior")) or "").strip()
            if raw:
                side_labels.setdefault(raw.upper(), raw)
                return raw.upper()
            derived = _resolve_unit_side(unit, id_to_unit)
            if derived:
                side_labels.setdefault(derived.upper(), derived)
                return derived.upper()
            side_labels.setdefault("UNKNOWN", "UNKNOWN")
            return "UNKNOWN"

        roots_by_side: Dict[str, List[ET._Element]] = {}
        for unit in list(model.unit_list):
            lvc = (text(jfind(unit, "j:LvcId")) or "").strip()
            superior = (text(jfind(unit, "j:UnitSuperior")) or "").strip()
            if superior and superior in id_to_unit:
                continue
            key = unit_side(unit)
            roots_by_side.setdefault(key, []).append(unit)
        if not roots_by_side:
            roots_by_side["UNKNOWN"] = []
            side_labels.setdefault("UNKNOWN", "UNKNOWN")
        for side_key in sorted(roots_by_side.keys(), key=side_sort_key):
            display = side_labels.get(side_key, side_key)
            shown = display
            if side_key == "UNKNOWN" and display.upper() == "UNKNOWN":
                shown = "UNKNOWN / Unassigned"
            side_item = QTreeWidgetItem([shown, "", ""])
            side_item.setData(0, Qt.ItemDataRole.UserRole, ("side", side_key, display))
            flags = side_item.flags() | Qt.ItemFlag.ItemIsEnabled
            if self._allow_side_selection:
                flags |= Qt.ItemFlag.ItemIsSelectable
            else:
                flags &= ~Qt.ItemFlag.ItemIsSelectable
            side_item.setFlags(flags)
            self.tree.addTopLevelItem(side_item)
            for unit in sorted(roots_by_side.get(side_key, []), key=unit_sort_key):
                self._add_unit_item(side_item, unit, parent_to_children)

    def _add_unit_item(
        self,
        parent_item: QTreeWidgetItem,
        unit: ET._Element,
        parent_to_children: Dict[str, List[ET._Element]],
    ) -> None:
        label = format_unit_label(unit)
        lvc = (text(jfind(unit, "j:LvcId")) or "").strip()
        code = (text(jfind(unit, "j:MilStd2525CCode")) or "").strip()
        item = QTreeWidgetItem([label, lvc, code])
        item.setData(0, Qt.ItemDataRole.UserRole, ("unit", unit))
        flags = item.flags() | Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsSelectable
        if id(unit) in self._blocked:
            flags &= ~Qt.ItemFlag.ItemIsEnabled
            flags &= ~Qt.ItemFlag.ItemIsSelectable
        item.setFlags(flags)
        parent_item.addChild(item)
        unit_lvc = lvc
        children = parent_to_children.get(unit_lvc, []) if unit_lvc else []
        for child in sorted(children, key=unit_sort_key):
            self._add_unit_item(item, child, parent_to_children)

    def _on_current_item_changed(
        self, current: Optional[QTreeWidgetItem], previous: Optional[QTreeWidgetItem]
    ) -> None:
        if current is None:
            self._status.setText("")
            return
        data = _normalize_user_data(current.data(0, Qt.ItemDataRole.UserRole))
        if not isinstance(data, tuple):
            self._status.setText("")
            return
        kind = data[0]
        if kind == "unit":
            unit = data[1]
            name = (text(jfind(unit, "j:Name")) or "").strip()
            lvc = (text(jfind(unit, "j:LvcId")) or "").strip()
            self._status.setText(f"Destination unit: {name} [{lvc}]")
        elif kind == "side":
            self._status.setText(f"Destination side: {data[2]}")
        else:
            self._status.setText("")

    def _on_accept(self) -> None:
        current = self.tree.currentItem()
        if current is None:
            QMessageBox.information(self, "Move", "Select a destination.")
            return
        data = _normalize_user_data(current.data(0, Qt.ItemDataRole.UserRole))
        if not isinstance(data, tuple):
            QMessageBox.information(self, "Move", "Select a valid destination.")
            return
        kind = data[0]
        if kind == "unit":
            unit = data[1]
            if id(unit) in self._blocked:
                QMessageBox.information(
                    self, "Move", "Cannot select the original unit or its descendants."
                )
                return
            self._result = data
        elif kind == "side":
            if not self._allow_side_selection:
                QMessageBox.information(self, "Move", "Select a unit destination.")
                return
            self._result = data
        else:
            QMessageBox.information(self, "Move", "Select a valid destination.")
            return
        super().accept()

    def selected_target(self) -> Optional[Tuple[str, Any]]:
        return self._result


@dataclass
class _EntityClassRowMeta:

    element: Optional[ET._Element]
    alias_element: Optional[ET._Element]
    alias_federate: Optional[str]
    is_new: bool


@dataclass
class _UnitClassRowMeta:

    element: Optional[ET._Element]
    alias_element: Optional[ET._Element]
    alias_federate: Optional[str]
    is_new: bool


@dataclass(frozen=True)
class MilStd2525Entry:

    identifier: str
    scheme: str
    battle_dimension: str
    function_id: str
    modifier1: str
    modifier2: str
    size_fragment: str
    country_fragment: str
    order_fragment: str
    title: str
    description: str
    search_text: str

    def function_block(self) -> str:
        return f"{self.function_id}{self.modifier1}{self.modifier2}"


_MILSTD2525_CACHE: Optional[List[MilStd2525Entry]] = None
_SIDC_LOOKUP_CACHE: Optional[_SidcAttributeLookup] = None


def _clean_sidc_fragment(value: Any, length: int, default: str = "-") -> str:

    text = ""
    if isinstance(value, str):
        text = value.strip()
    if not text or text == "nan":
        text = ""
    text = text.replace("*", "-").upper()
    cleaned = "".join(ch if ch.isalnum() or ch == "-" else "-" for ch in text)
    if not cleaned:
        cleaned = default * length
    if len(cleaned) < length:
        cleaned += "-" * (length - len(cleaned))
    return cleaned[:length]


def _load_sidc_attribute_lookup() -> _SidcAttributeLookup:

    global _SIDC_LOOKUP_CACHE
    if _SIDC_LOOKUP_CACHE is not None:
        return _SIDC_LOOKUP_CACHE
    csv_path = _resolve_symbols_path("2525c-tablei_ii.csv")
    if not csv_path:
        raise FileNotFoundError(
            "Could not locate Symbols/2525c-tablei_ii.csv needed for MIL-STD-2525C lookup."
        )
    try:
        df = pd.read_csv(csv_path)
    except Exception as exc:
        raise ValueError(f"Failed to load {csv_path}: {exc}") from exc

    def collect(label: str, modifier: str) -> List[Tuple[str, str]]:
        options: List[Tuple[str, str]] = []
        if label not in df.columns or modifier not in df.columns:
            return options
        for _, row in df.iterrows():
            raw_name = row.get(label)
            raw_code = row.get(modifier)
            if pd.isna(raw_name) or pd.isna(raw_code):
                continue
            name = str(raw_name).strip()
            code = str(raw_code).strip().upper()
            if name and code:
                options.append((name, code))
        return options

    lookup = _SidcAttributeLookup(
        affiliations=collect("Affiliation", "Affiliation Modifier"),
        statuses=collect("Status", "Status Modifier"),
        sizes=collect("Size", "Size Modifier"),
        countries=collect("Country", "Country Modifier"),
        orders_of_battle=collect("OrderOfBattle", "OrderOfBattle Modifier"),
    )
    _SIDC_LOOKUP_CACHE = lookup
    return lookup


def _load_milstd2525_library() -> List[MilStd2525Entry]:

    global _MILSTD2525_CACHE
    if _MILSTD2525_CACHE is not None:
        return _MILSTD2525_CACHE
    csv_path = _resolve_symbols_path("2525c-tableiii.csv")
    if not csv_path:
        raise FileNotFoundError(
            "Could not locate Symbols/2525c-tableiii.csv needed for MIL-STD-2525C lookup."
        )
    try:
        df = pd.read_csv(csv_path)
    except Exception as exc:
        raise ValueError(f"Failed to load {csv_path}: {exc}") from exc
    entries: List[MilStd2525Entry] = []
    for _, row in df.iterrows():
        identifier = str(row.get("Unnamed: 0") or "").strip()
        if not identifier:
            continue
        scheme = _clean_sidc_fragment(row.get("Scheme"), 1, "S")
        battle_dimension = _clean_sidc_fragment(row.get("BattleDimension"), 1, "G")
        function_id = _clean_sidc_fragment(row.get("FunctionID"), 2, "-")
        modifier1 = _clean_sidc_fragment(row.get("Modifier1"), 2, "-")
        modifier2 = _clean_sidc_fragment(row.get("Modifier2"), 2, "-")
        size_fragment = _clean_sidc_fragment(row.get("Size"), 2, "-")
        country_fragment = _clean_sidc_fragment(row.get("Country"), 2, "-")
        order_fragment = _clean_sidc_fragment(row.get("OrderOfBattle"), 1, "-")
        text_bits = [
            str(val).strip()
            for val in row.values.tolist()[11:]
            if isinstance(val, str) and val.strip()
        ]
        description = " ".join(text_bits)
        title = description or identifier.split(".")[-1]
        search_text = " ".join([identifier, description]).upper()
        entries.append(
            MilStd2525Entry(
                identifier=identifier,
                scheme=scheme,
                battle_dimension=battle_dimension,
                function_id=function_id,
                modifier1=modifier1,
                modifier2=modifier2,
                size_fragment=size_fragment,
                country_fragment=country_fragment,
                order_fragment=order_fragment,
                title=title.strip(),
                description=description.strip(),
                search_text=search_text,
            )
        )
    _MILSTD2525_CACHE = entries
    return entries


class MilStd2525PickerDialog(QDialog):

    def __init__(
        self,
        parent: Optional[QWidget],
        entries: List[MilStd2525Entry],
        lookup: _SidcAttributeLookup,
        *,
        initial_code: Optional[str] = None,
    ):
        super().__init__(parent)
        self.setWindowTitle("Select MIL-STD-2525C Code")
        self.resize(800, 600)
        self._entries = entries
        self._lookup = lookup
        self._entry_by_identifier: Dict[str, MilStd2525Entry] = {
            e.identifier: e for e in entries
        }
        self._entry_by_function: Dict[str, MilStd2525Entry] = {}
        for entry in entries:
            self._entry_by_function.setdefault(entry.function_block(), entry)
        self._selected_entry: Optional[MilStd2525Entry] = None
        self._selected_code: Optional[str] = None

        layout = QVBoxLayout(self)
        intro = QLabel(
            "Filter the MIL-STD-2525C hierarchy, choose a row, then adjust "
            "affiliation/status/echelon/country/order to build the 15-character SIDC."
        )
        intro.setWordWrap(True)
        layout.addWidget(intro)

        filter_row = QHBoxLayout()
        filter_row.addWidget(QLabel("Scheme"))
        self.scheme_filter = QComboBox()
        filter_row.addWidget(self.scheme_filter)
        filter_row.addWidget(QLabel("Battle Dimension"))
        self.dimension_filter = QComboBox()
        filter_row.addWidget(self.dimension_filter)
        filter_row.addWidget(QLabel("Search"))
        self.search_edit = QLineEdit()
        self.search_edit.setPlaceholderText("Type to search identifiers or descriptions")
        filter_row.addWidget(self.search_edit, 1)
        layout.addLayout(filter_row)

        self.tree = QTreeWidget()
        self.tree.setColumnCount(2)
        self.tree.setHeaderLabels(["Hierarchy", "Description"])
        self.tree.setUniformRowHeights(True)
        self.tree.itemSelectionChanged.connect(self._handle_selection_changed)
        self.tree.itemDoubleClicked.connect(self._handle_item_double_clicked)

        self.count_label = QLabel("")

        component_box = QGroupBox("Code Components")
        comp_grid = QGridLayout(component_box)
        self.affiliation_component = QComboBox()
        self.status_component = QComboBox()
        self.size_component = QComboBox()
        self.country_component = QComboBox()
        self.order_component = QComboBox()
        comp_grid.addWidget(QLabel("Affiliation"), 0, 0)
        comp_grid.addWidget(self.affiliation_component, 0, 1)
        comp_grid.addWidget(QLabel("Status"), 0, 2)
        comp_grid.addWidget(self.status_component, 0, 3)
        comp_grid.addWidget(QLabel("Echelon / HQ / Task"), 1, 0)
        comp_grid.addWidget(self.size_component, 1, 1)
        comp_grid.addWidget(QLabel("Country"), 1, 2)
        comp_grid.addWidget(self.country_component, 1, 3)
        comp_grid.addWidget(QLabel("Order of Battle"), 2, 0)
        comp_grid.addWidget(self.order_component, 2, 1)
        self.detail_label = QLabel("Select an entry to preview the SIDC.")
        self.detail_label.setWordWrap(True)

        self.preview_toggle = QCheckBox("Show symbol preview")
        self.preview_toggle.setChecked(True)

        preview_box = QGroupBox("Symbol Preview")
        self.preview_box = preview_box
        preview_layout = QVBoxLayout(preview_box)
        self.preview_notice = QLabel("")
        self.preview_notice.setWordWrap(True)
        preview_layout.addWidget(self.preview_notice)
        if QWebEngineView is not None:
            self.preview_view = QWebEngineView()
            self.preview_view.setContextMenuPolicy(Qt.ContextMenuPolicy.NoContextMenu)
            self.preview_view.setMinimumHeight(125)
            preview_layout.addWidget(self.preview_view, 1)
        else:
            self.preview_view = None
            self.preview_notice.setText(
                "Symbol preview unavailable: install PySide6-QtWebEngine."
            )
        function_panel = QWidget()
        function_layout = QVBoxLayout(function_panel)
        function_layout.setContentsMargins(0, 0, 0, 0)
        function_layout.addWidget(self.tree, 1)
        function_layout.addWidget(self.count_label)
        function_layout.addWidget(component_box)
        function_layout.addWidget(self.detail_label)
        function_layout.addWidget(self.preview_toggle, 0, Qt.AlignmentFlag.AlignLeft)

        splitter = QSplitter(Qt.Orientation.Vertical)
        splitter.addWidget(function_panel)
        splitter.addWidget(preview_box)
        splitter.setStretchFactor(0, 3)
        splitter.setStretchFactor(1, 2)
        splitter.setHandleWidth(10)
        splitter.setStyleSheet(
            """
QSplitter::handle {
    background-color: palette(midlight);
    border: 1px solid palette(dark);
}
QSplitter::handle:vertical {
    height: 10px;
}
"""
        )
        self._splitter = splitter
        layout.addWidget(splitter, 1)

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        self._ok_button = buttons.button(QDialogButtonBox.Ok)
        self._ok_button.setEnabled(False)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

        self.preview_toggle.toggled.connect(self._handle_preview_toggle)
        self.preview_box.setVisible(self.preview_toggle.isChecked())

        self._preview_enabled = False
        self._preview_ready = False
        self._preview_display_active = False
        self._current_preview_code: str = ""

        self._populate_filter_combos()
        self._populate_component_combos()
        self.scheme_filter.currentIndexChanged.connect(self._refilter_entries)
        self.dimension_filter.currentIndexChanged.connect(self._refilter_entries)
        self.search_edit.textChanged.connect(self._refilter_entries)
        for combo in [
            self.affiliation_component,
            self.status_component,
            self.size_component,
            self.country_component,
            self.order_component,
        ]:
            combo.currentIndexChanged.connect(self._update_preview)
        self._setup_symbol_preview()
        self._handle_preview_toggle(self.preview_toggle.isChecked())
        self._refilter_entries()
        if initial_code:
            self._apply_initial_code(initial_code)

    def selected_entry(self) -> Optional[MilStd2525Entry]:
        return self._selected_entry

    def selected_code(self) -> Optional[str]:
        return self._selected_code

    def _populate_filter_combos(self) -> None:
        def fill(combo: QComboBox, values: Iterable[str], mapping: Mapping[str, str]) -> None:
            combo.blockSignals(True)
            combo.clear()
            combo.addItem("Any", None)
            for value in sorted({v for v in values if v}):
                label = mapping.get(value.upper(), "Unknown")
                combo.addItem(f"{value} - {label}", value)
            combo.setCurrentIndex(0)
            combo.blockSignals(False)

        fill(self.scheme_filter, (e.scheme for e in self._entries), _SIDC_SCHEME_NAMES)
        fill(
            self.dimension_filter,
            (e.battle_dimension for e in self._entries),
            _SIDC_DIMENSION_NAMES,
        )

    def _populate_component_combos(self) -> None:
        def fill(combo: QComboBox, pairs: List[Tuple[str, str]], *, default: str, label: str) -> None:
            combo.blockSignals(True)
            combo.clear()
            default_display = f"{label} ({default})" if default not in ("--", "-") else f"{label}"
            combo.addItem(default_display, default)
            seen: Set[str] = set()
            for name, code in pairs:
                if code in seen:
                    continue
                seen.add(code)
                combo.addItem(f"{name} ({code})", code)
            combo.setCurrentIndex(0)
            combo.blockSignals(False)

        fill(
            self.affiliation_component,
            self._lookup.affiliations,
            default="F",
            label="Friendly",
        )
        fill(
            self.status_component,
            self._lookup.statuses,
            default="P",
            label="Present/Known",
        )
        fill(
            self.size_component,
            self._lookup.sizes,
            default="--",
            label="Unspecified (--)",
        )
        fill(
            self.country_component,
            self._lookup.countries,
            default="--",
            label="No Country (--)",
        )
        fill(
            self.order_component,
            self._lookup.orders_of_battle,
            default="-",
            label="No O/B (-)",
        )

    def _setup_symbol_preview(self) -> None:
        self._preview_enabled = False
        self._preview_ready = False
        if self.preview_view is None:
            if not self.preview_notice.text().strip():
                self.preview_notice.setText(
                    "Symbol preview unavailable: install PySide6-QtWebEngine."
                )
            self.preview_toggle.setChecked(False)
            self.preview_toggle.setEnabled(False)
            self.preview_box.setVisible(False)
            return
        js_path = _resolve_symbols_path("milsymbol.js")
        if not js_path:
            self.preview_notice.setText(
                "Symbol preview unavailable: Symbols/milsymbol.js not found."
            )
            self.preview_toggle.setChecked(False)
            self.preview_toggle.setEnabled(False)
            self.preview_box.setVisible(False)
            return
        self._preview_enabled = True
        self.preview_notice.setText("Symbol preview powered by Symbols/milsymbol.js.")
        base_dir = js_path.parent.resolve()
        base_url = QUrl.fromLocalFile(str(base_dir) + os.sep)
        self.preview_view.loadFinished.connect(self._handle_preview_load_finished)
        self.preview_view.setHtml(_MILSYMBOL_PREVIEW_HTML, base_url)

    def _handle_preview_toggle(self, checked: bool) -> None:
        active = bool(checked and self._preview_enabled)
        self._preview_display_active = active
        self.preview_box.setVisible(active)
        if active:
            self._maybe_update_symbol_preview()

    def _handle_preview_load_finished(self, ok: bool) -> None:
        self._preview_ready = bool(ok)
        if not ok:
            self.preview_notice.setText("Failed to load symbol preview page.")
            return
        self._maybe_update_symbol_preview()

    def _set_preview_code(self, code: Optional[str]) -> None:
        self._current_preview_code = (code or "").strip()
        self._maybe_update_symbol_preview()

    def _maybe_update_symbol_preview(self) -> None:
        if (
            not self._preview_enabled
            or self.preview_view is None
            or not self._preview_display_active
            or not self._preview_ready
        ):
            return
        payload = json.dumps(self._current_preview_code)
        self.preview_view.page().runJavaScript(f"window.renderSymbol({payload});")

    def _combo_value(self, combo: QComboBox) -> Optional[str]:
        value = combo.currentData()
        if isinstance(value, str) and value:
            return value
        return None

    def _refilter_entries(self) -> None:
        scheme = self._combo_value(self.scheme_filter)
        dimension = self._combo_value(self.dimension_filter)
        query = self.search_edit.text().strip().upper()
        matches: List[MilStd2525Entry] = []
        for entry in self._entries:
            if scheme and entry.scheme != scheme:
                continue
            if dimension and entry.battle_dimension != dimension:
                continue
            if query and query not in entry.search_text:
                continue
            matches.append(entry)
        self._populate_tree(matches)

    def _populate_tree(self, entries: List[MilStd2525Entry]) -> None:
        self.tree.setSortingEnabled(False)
        self.tree.clear()
        for entry in entries:
            description = entry.description or entry.title
            item = QTreeWidgetItem([entry.identifier, description])
            item.setData(0, Qt.ItemDataRole.UserRole, entry.identifier)
            self.tree.addTopLevelItem(item)
        self.tree.sortItems(0, Qt.SortOrder.AscendingOrder)
        self.tree.setSortingEnabled(True)
        self.count_label.setText(f"{len(entries)} matching symbol(s).")
        if entries:
            self.tree.setCurrentItem(self.tree.topLevelItem(0))
        else:
            self._selected_entry = None
            self._selected_code = None
            self._ok_button.setEnabled(False)
            self.detail_label.setText("No matches. Adjust the filters to continue.")
            self._set_preview_code(None)

    def _handle_selection_changed(self) -> None:
        item = self.tree.currentItem()
        if not item:
            self._selected_entry = None
            self._selected_code = None
            self._ok_button.setEnabled(False)
            self.detail_label.setText("Select an entry to preview the SIDC.")
            self._set_preview_code(None)
            return
        identifier = item.data(0, Qt.ItemDataRole.UserRole)
        entry = self._entry_by_identifier.get(identifier)
        self._selected_entry = entry
        if entry:
            self._apply_entry_defaults(entry)
        self._update_preview()

    def _apply_entry_defaults(self, entry: MilStd2525Entry) -> None:
        if entry.size_fragment and entry.size_fragment != "--":
            self._set_combo_to_value(self.size_component, entry.size_fragment)
        if entry.country_fragment and entry.country_fragment != "--":
            self._set_combo_to_value(self.country_component, entry.country_fragment)
        if entry.order_fragment and entry.order_fragment != "-":
            self._set_combo_to_value(self.order_component, entry.order_fragment)

    def _handle_item_double_clicked(self, item: QTreeWidgetItem, _column: int) -> None:
        if item is not None:
            self.accept()

    def _apply_initial_code(self, code: str) -> None:
        cleaned = "".join(ch for ch in str(code).strip().upper() if not ch.isspace())
        if len(cleaned) != 15:
            return
        function_block = cleaned[4:10]
        entry = self._entry_by_function.get(function_block)
        if entry:
            self._set_combo_to_value(self.scheme_filter, entry.scheme)
            self._set_combo_to_value(self.dimension_filter, entry.battle_dimension)
            self._refilter_entries()
            for idx in range(self.tree.topLevelItemCount()):
                item = self.tree.topLevelItem(idx)
                if item.data(0, Qt.ItemDataRole.UserRole) == entry.identifier:
                    self.tree.setCurrentItem(item)
                    self.tree.scrollToItem(item)
                    break
        self._set_combo_to_value(self.affiliation_component, cleaned[1])
        self._set_combo_to_value(self.status_component, cleaned[3])
        self._set_combo_to_value(self.size_component, cleaned[10:12])
        self._set_combo_to_value(self.country_component, cleaned[12:14])
        self._set_combo_to_value(self.order_component, cleaned[14])

    def _set_combo_to_value(self, combo: QComboBox, value: str) -> None:
        if value is None:
            combo.setCurrentIndex(0)
            return
        idx = combo.findData(value)
        if idx >= 0:
            combo.setCurrentIndex(idx)

    def _component_fragment(self, combo: QComboBox, *, length: int, default: str) -> str:
        value = combo.currentData()
        if isinstance(value, str) and value:
            return _clean_sidc_fragment(value, length, default)
        return default * length

    def _compose_sidc(self) -> Optional[str]:
        entry = self._selected_entry
        if not entry:
            return None
        scheme = _clean_sidc_fragment(entry.scheme, 1, "S")
        dimension = _clean_sidc_fragment(entry.battle_dimension, 1, "G")
        affiliation = self._component_fragment(self.affiliation_component, length=1, default="F")
        status = self._component_fragment(self.status_component, length=1, default="P")
        size_fragment = self._component_fragment(self.size_component, length=2, default="-")
        country_fragment = self._component_fragment(self.country_component, length=2, default="-")
        order_fragment = self._component_fragment(self.order_component, length=1, default="-")
        code = (
            scheme
            + affiliation
            + dimension
            + status
            + entry.function_block()
            + size_fragment
            + country_fragment
            + order_fragment
        )
        return code

    def _update_preview(self) -> None:
        entry = self._selected_entry
        if not entry:
            self._selected_code = None
            self._ok_button.setEnabled(False)
            self.detail_label.setText("Select an entry to preview the SIDC.")
            self._set_preview_code(None)
            return
        code = self._compose_sidc()
        if not code:
            self._selected_code = None
            self._ok_button.setEnabled(False)
            self.detail_label.setText("Unable to compose a SIDC for the selected entry.")
            self._set_preview_code(None)
            return
        self._selected_code = code
        self._ok_button.setEnabled(True)
        descriptor = entry.description or entry.title
        self.detail_label.setText(f"{code}\n{entry.title}\n{descriptor}")
        self._set_preview_code(code)


class EntityClassesTab(QWidget):

    def __init__(self, parent):
        super().__init__(parent)
        self.model: Optional[ObsModel] = None
        self._row_meta: Dict[str, _EntityClassRowMeta] = {}
        self._new_row_seq = 0
        self._federate_label_all = "All Federates"
        self._refreshing_federate_combo = False
        self._build()

    def _build(self):
        outer = QVBoxLayout(self)
        self.info_label = QLabel(
            "Load or create a JTDS OBS model to view EntityCompositionClass entries."
        )
        self.info_label.setWordWrap(True)
        outer.addWidget(self.info_label)
        filter_row = QHBoxLayout()
        filter_row.addWidget(QLabel("Federate filter:"))
        self.federate_combo = QComboBox()
        self.federate_combo.setEditable(True)
        self.federate_combo.lineEdit().setPlaceholderText(self._federate_label_all)
        self.federate_combo.setInsertPolicy(QComboBox.InsertPolicy.NoInsert)
        self.federate_combo.addItem(self._federate_label_all)
        self.federate_combo.currentTextChanged.connect(self._on_federate_filter_changed)
        filter_row.addWidget(self.federate_combo)
        filter_row.addStretch()
        outer.addLayout(filter_row)
        controls = QHBoxLayout()
        self.btn_refresh = QPushButton("Reload")
        self.btn_refresh.clicked.connect(self._reload_table)
        self.btn_add = QPushButton("Add Row")
        self.btn_add.clicked.connect(self._on_add_row)
        self.btn_save = QPushButton("Save Changes")
        self.btn_save.clicked.connect(self._on_save)
        self.btn_import = QPushButton("Import CSV...")
        self.btn_import.clicked.connect(self._on_import)
        self.btn_export = QPushButton("Export CSV...")
        self.btn_export.clicked.connect(self._on_export)
        self.sort_combo = QComboBox()
        self.sort_combo.addItems(["Name (A-Z)", "DisCode (A-Z)"])
        self.btn_sort = QPushButton("Sort")
        self.btn_sort.clicked.connect(self._on_sort)
        controls.addWidget(self.btn_refresh)
        controls.addWidget(self.btn_add)
        controls.addWidget(self.btn_save)
        controls.addWidget(self.btn_import)
        controls.addWidget(self.btn_export)
        controls.addWidget(QLabel("Sort by:"))
        controls.addWidget(self.sort_combo)
        controls.addWidget(self.btn_sort)
        controls.addStretch()
        outer.addLayout(controls)
        self.table = QTableWidget(0, 4)
        self.table.setHorizontalHeaderLabels(["Name", "Federate", "TypeName", "DisCode"])
        self.table.verticalHeader().setVisible(False)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectItems)
        self.table.setEditTriggers(
            QAbstractItemView.EditTrigger.DoubleClicked
            | QAbstractItemView.EditTrigger.SelectedClicked
        )
        header = self.table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        outer.addWidget(self.table)
        self.status_label = QLabel("No model loaded.")
        outer.addWidget(self.status_label)
        self._set_enabled(False)

    def _set_enabled(self, enabled: bool) -> None:
        for widget in [
            self.btn_refresh,
            self.btn_add,
            self.btn_save,
            self.btn_import,
            self.btn_export,
            self.sort_combo,
            self.btn_sort,
            self.table,
            self.federate_combo,
        ]:
            widget.setEnabled(enabled)

    def _current_federate_filter(self) -> Optional[str]:
        text = self.federate_combo.currentText().strip()
        if not text or text.lower() == self._federate_label_all.lower():
            return None
        return text

    def _on_federate_filter_changed(self, _text: str) -> None:
        if self._refreshing_federate_combo:
            return
        self._reload_table()

    def _collect_federate_names(self) -> List[str]:
        names: Set[str] = set()
        if not self.model:
            return []
        ecc_parent = jfind(self.model.classdata, jtag("EntityCompositionClassList"))
        if ecc_parent is None:
            return []
        for alias in _iter_local(ecc_parent, "Alias"):
            fed = text(jfind(alias, "j:Federate"))
            if fed:
                names.add(fed.strip())
        return sorted(names, key=lambda s: s.upper())

    def _refresh_federate_combo(self) -> None:
        options = self._collect_federate_names()
        current_text = self.federate_combo.currentText().strip()
        self._refreshing_federate_combo = True
        block = self.federate_combo.blockSignals(True)
        self.federate_combo.clear()
        self.federate_combo.addItem(self._federate_label_all)
        for name in options:
            self.federate_combo.addItem(name)
        if current_text and current_text.lower() != self._federate_label_all.lower():
            idx = self.federate_combo.findText(
                current_text, Qt.MatchFlag.MatchFixedString
            )
            if idx >= 0:
                self.federate_combo.setCurrentIndex(idx)
            else:
                self.federate_combo.setEditText(current_text)
        else:
            self.federate_combo.setCurrentIndex(0)
        self.federate_combo.blockSignals(block)
        self._refreshing_federate_combo = False

    def set_model(self, model: Optional[ObsModel]):
        self.model = model
        has_model = model is not None
        self._set_enabled(has_model)
        self._refresh_federate_combo()
        if has_model:
            self._reload_table()
        else:
            self._clear_table()
            self.info_label.setText(
                "Load or create a JTDS OBS model to manage EntityCompositionClasses."
            )
            self.status_label.setText("No model loaded.")

    def _clear_table(self) -> None:
        self.table.setSortingEnabled(False)
        self.table.setRowCount(0)
        self._row_meta.clear()
        self.table.setSortingEnabled(True)

    def _reload_table(self) -> None:
        if not self.model:
            return
        self.table.setSortingEnabled(False)
        self.table.setRowCount(0)
        self._row_meta.clear()
        ecc_parent = jfind(self.model.classdata, jtag("EntityCompositionClassList"))
        if ecc_parent is None:
            self.info_label.setText(
                "This model does not have an EntityCompositionClassList yet. Saving new rows will create it."
            )
            self.status_label.setText("0 entity classes loaded.")
            self.table.setSortingEnabled(True)
            return
        federate_filter = self._current_federate_filter()
        rows: List[
            Tuple[
                str,
                str,
                str,
                str,
                ET._Element,
                Optional[ET._Element],
                Optional[str],
            ]
        ] = []
        seen: Set[str] = set()
        for ecc in list(ecc_parent):
            if _local(ecc.tag) != "EntityCompositionClass":
                continue
            name = self._extract_name(ecc)
            if not name:
                continue
            key = name.upper()
            if key in seen:
                continue
            seen.add(key)
            discode = self._extract_discode(ecc)
            if federate_filter:
                alias = self._find_alias_for_federate(ecc, federate_filter)
                federate_val, typename_val = self._alias_fields(alias)
                rows.append(
                    (
                        name,
                        federate_val,
                        typename_val,
                        discode,
                        ecc,
                        alias,
                        federate_filter,
                    )
                )
            else:
                federate_val, typename_val, alias = self._extract_alias_values(ecc)
                rows.append(
                    (
                        name,
                        federate_val,
                        typename_val,
                        discode,
                        ecc,
                        alias,
                        federate_val or None,
                    )
                )
        rows.sort(key=lambda item: item[0].upper())
        for name, federate, typename, discode, ecc, alias_elem, alias_fed in rows:
            self._append_row(
                name,
                federate,
                typename,
                discode,
                ecc,
                is_new=False,
                alias_element=alias_elem,
                alias_federate=alias_fed,
            )
        self.table.setSortingEnabled(True)
        self.info_label.setText(
            "Edit Federate, TypeName, and DisCode values. Add rows to create new EntityCompositionClasses."
        )
        self.status_label.setText(f"{len(rows)} entity classes loaded.")

    def _next_row_key(self, element: Optional[ET._Element]) -> str:
        if element is not None:
            return f"ecc-{id(element)}"
        self._new_row_seq += 1
        return f"new-{self._new_row_seq}"

    def _append_row(
        self,
        name: str,
        federate: str,
        typename: str,
        discode: str,
        element: Optional[ET._Element],
        *,
        is_new: bool,
        alias_element: Optional[ET._Element] = None,
        alias_federate: Optional[str] = None,
    ) -> None:
        row = self.table.rowCount()
        self.table.insertRow(row)
        key = self._next_row_key(element)
        name_item = QTableWidgetItem(name)
        name_item.setData(Qt.ItemDataRole.UserRole, key)
        if not is_new:
            name_item.setFlags(name_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
        self.table.setItem(row, 0, name_item)
        for col, value in enumerate([federate, typename, discode], start=1):
            item = QTableWidgetItem(value or "")
            self.table.setItem(row, col, item)
        self._row_meta[key] = _EntityClassRowMeta(
            element=element,
            alias_element=alias_element,
            alias_federate=alias_federate,
            is_new=is_new,
        )

    def _table_rows(self) -> List[Tuple[str, str, str, str]]:
        rows: List[Tuple[str, str, str, str]] = []
        federate_filter = self._current_federate_filter()
        for row in range(self.table.rowCount()):
            name_item = self.table.item(row, 0)
            fed_item = self.table.item(row, 1)
            type_item = self.table.item(row, 2)
            discode_item = self.table.item(row, 3)
            name = (name_item.text() if name_item else "").strip()
            federate = (fed_item.text() if fed_item else "").strip()
            if federate_filter:
                federate = federate_filter
            typename = (type_item.text() if type_item else "").strip()
            discode = (discode_item.text() if discode_item else "").strip()
            rows.append((name, federate, typename, discode))
        return rows

    def _name_to_row(self) -> Dict[str, int]:
        mapping: Dict[str, int] = {}
        for row in range(self.table.rowCount()):
            name_item = self.table.item(row, 0)
            if not name_item:
                continue
            name = (name_item.text() or "").strip()
            if not name:
                continue
            mapping[name.upper()] = row
        return mapping

    def _table_discode_map(self) -> Dict[str, str]:
        mapping: Dict[str, str] = {}
        for name, _, _, discode in self._table_rows():
            if not name or not discode:
                continue
            try:
                _, normalized = self._normalize_discode_value(discode)
            except ValueError:
                continue
            mapping[normalized] = name.upper()
        return mapping

    def _set_cell_text(self, row: int, column: int, value: str) -> None:
        item = self.table.item(row, column)
        if item is None:
            item = QTableWidgetItem(value)
            self.table.setItem(row, column, item)
        else:
            item.setText(value)

    def _normalize_discode_value(self, raw: str) -> Tuple[Dict[str, str], str]:
        attrs = _parse_discode_query(raw)
        normalized = ".".join(attrs[key] for key in _DISCODE_ATTRS)
        return attrs, normalized

    def _extract_name(self, ecc: ET._Element) -> str:
        for tag in ("Name", "ClassName", "TypeName"):
            value = text(jfind(ecc, f"j:{tag}"))
            if value:
                return value.strip()
        for child in list(ecc):
            if _local(child.tag) in {"Name", "ClassName", "TypeName"} and child.text:
                return child.text.strip()
        return ""

    def _first_alias(self, ecc: ET._Element) -> Optional[ET._Element]:
        alias_list = jfind(ecc, "j:AliasList")
        if alias_list is not None:
            for child in list(alias_list):
                if _local(child.tag) == "Alias":
                    return child
        return next((child for child in _iter_local(ecc, "Alias")), None)

    def _alias_fields(self, alias: Optional[ET._Element]) -> Tuple[str, str]:
        federate = ""
        typename = ""
        if alias is None:
            return federate, typename
        fed_elem = next(
            (child for child in list(alias) if _local(child.tag) == "Federate" and child.text),
            None,
        )
        if fed_elem is not None:
            federate = fed_elem.text.strip()
        type_elem = next(
            (child for child in list(alias) if _local(child.tag) == "TypeName" and child.text),
            None,
        )
        if type_elem is not None:
            typename = type_elem.text.strip()
        return federate, typename

    def _alias_map(self, ecc: ET._Element) -> Dict[str, ET._Element]:
        mapping: Dict[str, ET._Element] = {}
        alias_list = jfind(ecc, "j:AliasList")
        candidates: List[ET._Element] = []
        if alias_list is not None:
            for child in list(alias_list):
                if _local(child.tag) == "Alias":
                    candidates.append(child)
        else:
            candidates.extend(_iter_local(ecc, "Alias"))
        for alias in candidates:
            federate = (text(jfind(alias, "j:Federate")) or "").strip()
            if federate:
                mapping[federate.upper()] = alias
        return mapping

    def _find_alias_for_federate(
        self, ecc: ET._Element, federate: Optional[str]
    ) -> Optional[ET._Element]:
        if not federate:
            return None
        return self._alias_map(ecc).get(federate.strip().upper())

    def _extract_alias_values(
        self, ecc: ET._Element
    ) -> Tuple[str, str, Optional[ET._Element]]:
        alias = self._first_alias(ecc)
        federate, typename = self._alias_fields(alias)
        return federate, typename, alias

    def _find_discode_element(self, ecc: ET._Element) -> Optional[ET._Element]:
        for child in list(ecc):
            if _local(child.tag) == "DisCode":
                return child
        return next((child for child in _iter_local(ecc, "DisCode")), None)

    def _extract_discode(self, ecc: ET._Element) -> str:
        discode_elem = self._find_discode_element(ecc)
        if discode_elem is None:
            return ""
        parts: List[str] = []
        for key in _DISCODE_ATTRS:
            attr = discode_elem.get(key)
            if attr is None:
                return ""
            parts.append(_normalize_discode_component(attr))
        return ".".join(parts)

    def _on_add_row(self) -> None:
        if not self.model:
            return
        self._append_row(
            "",
            "",
            "",
            "",
            None,
            is_new=True,
            alias_federate=self._current_federate_filter(),
        )
        self.status_label.setText(
            f"{self.table.rowCount()} row(s) in table. New rows require a Name before saving."
        )

    def _on_sort(self) -> None:
        column = 0 if self.sort_combo.currentIndex() == 0 else 3
        self.table.sortItems(column)

    def _on_export(self) -> None:
        rows = self._table_rows()
        if not rows:
            QMessageBox.information(
                self, "Export Entity Classes", "No rows to export."
            )
            return
        path, _ = QFileDialog.getSaveFileName(
            self,
            "Export Entity Classes",
            "EntityClasses.csv",
            "CSV Files (*.csv);;All Files (*)",
        )
        if not path:
            return
        if not path.lower().endswith(".csv"):
            path += ".csv"
        try:
            with open(path, "w", newline="", encoding="utf-8") as handle:
                writer = csv.writer(handle)
                writer.writerow(["Name", "Federate", "TypeName", "DisCode"])
                for record in rows:
                    writer.writerow(record)
        except Exception as exc:
            QMessageBox.warning(
                self, "Export Entity Classes", f"Failed to export: {exc}"
            )
            return
        QMessageBox.information(
            self,
            "Export Entity Classes",
            f"Exported {len(rows)} row(s) to {path}",
        )

    def _resolve_field_name(self, fieldnames: Optional[List[str]], target: str) -> Optional[str]:
        if not fieldnames:
            return None
        normalized_target = target.replace(" ", "").lower()
        for name in fieldnames:
            normalized = name.replace(" ", "").lower()
            if normalized == normalized_target:
                return name
        return None

    def _on_import(self) -> None:
        if not self.model:
            QMessageBox.information(
                self,
                "Import Entity Classes",
                "Load or create a JTDS OBS model before importing.",
            )
            return
        filter_value = self._current_federate_filter()
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Import Entity Classes",
            "",
            "CSV Files (*.csv);;All Files (*)",
        )
        if not path:
            return
        try:
            with open(path, "r", newline="", encoding="utf-8-sig") as handle:
                reader = csv.DictReader(handle)
                fieldnames = reader.fieldnames or []
                name_field = self._resolve_field_name(fieldnames, "Name")
                if not name_field:
                    raise ValueError("CSV is missing a 'Name' column.")
                fed_field = self._resolve_field_name(fieldnames, "Federate")
                type_field = self._resolve_field_name(fieldnames, "TypeName")
                disc_field = self._resolve_field_name(fieldnames, "DisCode")
                incoming: List[Dict[str, str]] = []
                seen_names: Set[str] = set()
                discode_conflicts: Dict[str, str] = {}
                issues: List[str] = []
                line_num = 1
                existing_discodes = self._table_discode_map()
                for row in reader:
                    line_num += 1
                    name = (row.get(name_field) or "").strip()
                    if not name:
                        issues.append(f"Row {line_num}: Missing Name; skipped.")
                        continue
                    upper = name.upper()
                    if upper in seen_names:
                        issues.append(
                            f"Row {line_num}: Duplicate Name '{name}' within file; skipped."
                        )
                        continue
                    federate = (
                        (row.get(fed_field) or "").strip() if fed_field else ""
                    )
                    typename = (
                        (row.get(type_field) or "").strip() if type_field else ""
                    )
                    discode_raw = (
                        (row.get(disc_field) or "").strip() if disc_field else ""
                    )
                    normalized_disc = ""
                    if discode_raw:
                        try:
                            _, normalized_disc = self._normalize_discode_value(discode_raw)
                        except ValueError as exc:
                            issues.append(f"Row {line_num}: {exc}; skipped.")
                            continue
                        owner_existing = existing_discodes.get(normalized_disc)
                        if owner_existing and owner_existing != upper:
                            issues.append(
                                f"Row {line_num}: DisCode already used by '{owner_existing}'."
                            )
                            continue
                        owner_incoming = discode_conflicts.get(normalized_disc)
                        if owner_incoming and owner_incoming != upper:
                            issues.append(
                                f"Row {line_num}: DisCode duplicates '{owner_incoming}'."
                            )
                            continue
                        discode_conflicts[normalized_disc] = upper
                    incoming.append(
                        {
                            "name": name,
                            "name_upper": upper,
                            "federate": federate,
                            "typename": typename,
                            "discode": normalized_disc,
                        }
                    )
                    seen_names.add(upper)
        except ValueError as exc:
            QMessageBox.warning(self, "Import Entity Classes", str(exc))
            return
        except Exception as exc:
            QMessageBox.warning(
                self, "Import Entity Classes", f"Failed to import: {exc}"
            )
            return
        if not incoming:
            msg = "No valid rows were imported."
            if issues:
                preview = "\n".join(issues[:5])
                if len(issues) > 5:
                    preview += "\n..."
                msg += f"\nIssues:\n{preview}"
            QMessageBox.information(self, "Import Entity Classes", msg)
            return
        name_to_row = self._name_to_row()
        updated = 0
        created = 0
        sorting_prev = self.table.isSortingEnabled()
        self.table.setSortingEnabled(False)
        try:
            for entry in incoming:
                idx = name_to_row.get(entry["name_upper"])
                if idx is not None:
                    self._set_cell_text(idx, 1, entry["federate"])
                    self._set_cell_text(idx, 2, entry["typename"])
                    self._set_cell_text(idx, 3, entry["discode"])
                    updated += 1
                else:
                    self._append_row(
                        entry["name"],
                        entry["federate"],
                        entry["typename"],
                        entry["discode"],
                        None,
                        is_new=True,
                        alias_federate=filter_value,
                    )
                    created += 1
        finally:
            self.table.setSortingEnabled(sorting_prev)
        total_rows = self.table.rowCount()
        self.status_label.setText(
            f"Table now has {total_rows} row(s). Imported {created} new, updated {updated}."
        )
        summary = (
            f"Import complete. Added {created} row(s), updated {updated} row(s)."
        )
        if issues:
            preview = "\n".join(issues[:5])
            if len(issues) > 5:
                preview += "\n..."
            summary += f"\nWarnings:\n{preview}"
        QMessageBox.information(self, "Import Entity Classes", summary)
        self._refresh_federate_combo()

    def _ensure_child_text(self, parent: ET._Element, tag: str, value: str) -> ET._Element:
        node = jfind(parent, f"j:{tag}")
        if node is None:
            node = ET.SubElement(parent, jtag(tag))
        node.text = value
        return node

    def _update_alias(
        self,
        ecc: ET._Element,
        federate: str,
        typename: str,
        *,
        alias_element: Optional[ET._Element],
        target_federate: Optional[str],
    ) -> None:
        alias_list = jfind(ecc, "j:AliasList")
        if alias_list is None:
            alias_list = ET.SubElement(ecc, jtag("AliasList"))
        desired_federate = (federate or "").strip()
        target_key = (target_federate or "").strip()
        alias: Optional[ET._Element] = None
        if target_key:
            alias = self._find_alias_for_federate(ecc, target_key)
        if alias is None and alias_element is not None:
            existing_fed, _ = self._alias_fields(alias_element)
            if not target_key or existing_fed.strip().upper() == target_key.upper():
                alias = alias_element
        if alias is None and not target_key and desired_federate:
            alias = self._find_alias_for_federate(ecc, desired_federate)
        if alias is None and not target_key:
            alias = jfind(alias_list, "j:Alias")
        if alias is None:
            alias = ET.SubElement(alias_list, jtag("Alias"))
        fed = jfind(alias, "j:Federate")
        if fed is None:
            fed = ET.SubElement(alias, jtag("Federate"))
        fed.text = desired_federate or (target_key or None)
        tname = jfind(alias, "j:TypeName")
        if tname is None:
            tname = ET.SubElement(alias, jtag("TypeName"))
        tname.text = typename or None

    def _update_discode(
        self, ecc: ET._Element, attrs: Optional[Dict[str, str]]
    ) -> None:
        discode_elem = self._find_discode_element(ecc)
        if attrs is None:
            if discode_elem is not None and discode_elem in list(ecc):
                ecc.remove(discode_elem)
            return
        if discode_elem is None:
            discode_elem = ET.SubElement(ecc, jtag("DisCode"))
        for key in _DISCODE_ATTRS:
            discode_elem.set(key, attrs.get(key, "0"))

    def _collect_existing_name_keys(self) -> Set[str]:
        keys: Set[str] = set()
        for row in range(self.table.rowCount()):
            item = self.table.item(row, 0)
            if not item:
                continue
            key = _normalize_user_data(item.data(Qt.ItemDataRole.UserRole))
            if key is None:
                continue
            meta = self._row_meta.get(key)
            if not meta or meta.is_new:
                continue
            name = (item.text() or "").strip()
            if name:
                keys.add(name.upper())
        return keys

    def _on_save(self) -> None:
        if not self.model:
            return
        ecc_parent = jfind(self.model.classdata, jtag("EntityCompositionClassList"))
        if ecc_parent is None:
            ecc_parent = ET.SubElement(
                self.model.classdata, jtag("EntityCompositionClassList")
            )
        errors: List[str] = []
        created = 0
        updated = 0
        existing_name_keys = self._collect_existing_name_keys()
        discode_owner: Dict[str, Tuple[str, str]] = {}
        for row in range(self.table.rowCount()):
            name_item = self.table.item(row, 0)
            if name_item is None:
                continue
            key = _normalize_user_data(name_item.data(Qt.ItemDataRole.UserRole))
            meta = self._row_meta.get(key)
            if not meta:
                continue
            name = (name_item.text() or "").strip()
            fed_item = self.table.item(row, 1)
            federate = (fed_item.text() if fed_item else "").strip()
            filter_value = self._current_federate_filter()
            if filter_value:
                federate = filter_value
            filter_value = self._current_federate_filter()
            if filter_value:
                federate = filter_value
            type_item = self.table.item(row, 2)
            typename = (type_item.text() if type_item else "").strip()
            discode_item = self.table.item(row, 3)
            discode_raw = (discode_item.text() if discode_item else "").strip()
            discode_attrs: Optional[Dict[str, str]] = None
            normalized_disc: Optional[str] = None
            if discode_raw:
                try:
                    discode_attrs, normalized_disc = self._normalize_discode_value(
                        discode_raw
                    )
                except ValueError as exc:
                    errors.append(f"Row {row + 1}: {exc}")
                    continue
            if normalized_disc:
                display_name = name or f"Row {row + 1}"
                upper_name = display_name.upper()
                owner = discode_owner.get(normalized_disc)
                if owner and owner[0] != upper_name:
                    errors.append(
                        f"Row {row + 1}: DisCode duplicates '{owner[1]}' entries."
                    )
                    continue
                discode_owner[normalized_disc] = (upper_name, display_name)
            if meta.is_new:
                if not name:
                    errors.append(f"Row {row + 1}: Name is required for new entries.")
                    continue
                key_upper = name.upper()
                if key_upper in existing_name_keys:
                    errors.append(
                        f"Row {row + 1}: Name '{name}' already exists in the current list."
                    )
                    continue
                existing_name_keys.add(key_upper)
                ecc = ET.SubElement(ecc_parent, jtag("EntityCompositionClass"))
                self._ensure_child_text(ecc, "Name", name)
                self._ensure_child_text(ecc, "ClassName", name)
                meta.element = ecc
                meta.is_new = False
                created += 1
            else:
                ecc = meta.element
                if ecc is None:
                    continue
                if name:
                    self._ensure_child_text(ecc, "Name", name)
                    self._ensure_child_text(ecc, "ClassName", name)
                updated += 1
            if meta.element is None:
                continue
            alias_target = filter_value or meta.alias_federate or federate
            self._update_alias(
                meta.element,
                federate,
                typename,
                alias_element=meta.alias_element,
                target_federate=alias_target,
            )
            self._update_discode(meta.element, discode_attrs)
        if errors:
            preview = "\n".join(errors[:5])
            if len(errors) > 5:
                preview += "\n..."
            QMessageBox.warning(self, "Save Entity Classes", preview)
            return
        QMessageBox.information(
            self,
            "Save Entity Classes",
            f"EntityCompositionClass changes saved. Updated {updated}, created {created}.",
        )
        self._reload_table()
        self._refresh_federate_combo()


class UnitClassesTab(QWidget):

    def __init__(self, parent):
        super().__init__(parent)
        self.model: Optional[ObsModel] = None
        self._row_meta: Dict[str, _UnitClassRowMeta] = {}
        self._new_row_seq = 0
        self._federate_label_all = "All Federates"
        self._refreshing_federate_combo = False
        (
            self._warsim_typename_choices,
            self._warsim_typename_lookup,
        ) = self._load_warsim_typename_options()
        self._table_item_change_block = 0
        self._build()

    def _build(self):
        outer = QVBoxLayout(self)
        self.info_label = QLabel(
            "Load or create a JTDS OBS model to view UnitClass entries."
        )
        self.info_label.setWordWrap(True)
        outer.addWidget(self.info_label)
        filter_row = QHBoxLayout()
        filter_row.addWidget(QLabel("Federate filter:"))
        self.federate_combo = QComboBox()
        self.federate_combo.setEditable(True)
        self.federate_combo.lineEdit().setPlaceholderText(self._federate_label_all)
        self.federate_combo.setInsertPolicy(QComboBox.InsertPolicy.NoInsert)
        self.federate_combo.addItem(self._federate_label_all)
        self.federate_combo.currentTextChanged.connect(self._on_federate_filter_changed)
        filter_row.addWidget(self.federate_combo)
        filter_row.addStretch()
        outer.addLayout(filter_row)
        controls = QHBoxLayout()
        self.btn_refresh = QPushButton("Reload")
        self.btn_refresh.clicked.connect(self._reload_table)
        self.btn_add = QPushButton("Add Row")
        self.btn_add.clicked.connect(self._on_add_row)
        self.btn_pick_code = QPushButton("Pick 2525C...")
        self.btn_pick_code.clicked.connect(self._on_pick_2525_code)
        self.btn_save = QPushButton("Save Changes")
        self.btn_save.clicked.connect(self._on_save)
        self.btn_import = QPushButton("Import CSV...")
        self.btn_import.clicked.connect(self._on_import)
        self.btn_export = QPushButton("Export CSV...")
        self.btn_export.clicked.connect(self._on_export)
        self.sort_combo = QComboBox()
        self.sort_combo.addItems(["AggregateName (A-Z)", "2525C Code (A-Z)"])
        self.btn_sort = QPushButton("Sort")
        self.btn_sort.clicked.connect(self._on_sort)
        controls.addWidget(self.btn_refresh)
        controls.addWidget(self.btn_add)
        controls.addWidget(self.btn_pick_code)
        controls.addWidget(self.btn_save)
        controls.addWidget(self.btn_import)
        controls.addWidget(self.btn_export)
        controls.addWidget(QLabel("Sort by:"))
        controls.addWidget(self.sort_combo)
        controls.addWidget(self.btn_sort)
        controls.addStretch()
        outer.addLayout(controls)
        self.table = QTableWidget(0, 5)
        self.table.setHorizontalHeaderLabels(
            [
                "AggregateName",
                "Federate",
                "TypeName",
                "2525C Code",
                "UnitDisEnumeration",
            ]
        )
        self.table.verticalHeader().setVisible(False)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectItems)
        self.table.setEditTriggers(
            QAbstractItemView.EditTrigger.DoubleClicked
            | QAbstractItemView.EditTrigger.SelectedClicked
        )
        header = self.table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.table.itemChanged.connect(self._on_table_item_changed)
        outer.addWidget(self.table)
        self.status_label = QLabel("No model loaded.")
        outer.addWidget(self.status_label)
        self._set_enabled(False)

    def _set_enabled(self, enabled: bool) -> None:
        for widget in [
            self.btn_refresh,
            self.btn_add,
            self.btn_pick_code,
            self.btn_save,
            self.btn_import,
            self.btn_export,
            self.sort_combo,
            self.btn_sort,
            self.table,
            self.federate_combo,
        ]:
            widget.setEnabled(enabled)

    @contextmanager
    def _suspend_table_item_signals(self):
        self._table_item_change_block += 1
        try:
            yield
        finally:
            self._table_item_change_block -= 1

    def _on_table_item_changed(self, item: Optional[QTableWidgetItem]) -> None:
        if not item or self._table_item_change_block > 0:
            return
        if item.column() == 1:
            self._update_row_typename_editor(item.row())

    def _load_warsim_typename_options(self) -> Tuple[List[str], Dict[str, str]]:
        options: List[str] = []
        lookup: Dict[str, str] = {}
        path = WARSIM_TYPENAME_MASTER
        try:
            with open(path, "r", encoding="utf-8-sig", newline="") as handle:
                reader = csv.reader(handle)
                for row in reader:
                    if not row:
                        continue
                    value = (row[0] or "").strip()
                    if not value:
                        continue
                    lower = value.lower()
                    if lower in lookup:
                        continue
                    options.append(value)
                    lookup[lower] = value
        except FileNotFoundError:
            return [], {}
        except Exception as exc:
            print(f"Warning: Failed to load WARSIM TypeName master: {exc}")
            return [], {}
        return options, lookup

    def _is_warsim_federate(self, federate: str) -> bool:
        return federate.strip().upper() == "WARSIM"

    def _canonical_warsim_typename(self, value: str) -> Optional[str]:
        trimmed = value.strip()
        if not trimmed:
            return ""
        return self._warsim_typename_lookup.get(trimmed.lower())

    def _is_warsim_typename_widget(self, widget: Optional[QWidget]) -> bool:
        return bool(
            isinstance(widget, QComboBox) and widget.property("warsim_typename_combo")
        )

    def _create_warsim_combobox(self) -> QComboBox:
        combo = QComboBox()
        combo.addItem("")
        combo.addItems(self._warsim_typename_choices)
        combo.setProperty("warsim_typename_combo", True)
        combo.currentTextChanged.connect(self._on_warsim_typename_changed)
        return combo

    def _set_combo_value(self, combo: QComboBox, value: str) -> None:
        combo.blockSignals(True)
        if not value:
            combo.setCurrentIndex(0)
        else:
            idx = combo.findText(value, Qt.MatchFlag.MatchFixedString)
            combo.setCurrentIndex(idx if idx >= 0 else 0)
        combo.blockSignals(False)

    def _update_row_typename_editor(self, row: int) -> None:
        if (
            row < 0
            or row >= self.table.rowCount()
            or not self._warsim_typename_choices
        ):
            return
        federate = self._cell_text(row, 1)
        widget = self.table.cellWidget(row, 2)
        item = self.table.item(row, 2)
        raw_item_text = (item.text() if item else "").strip()
        current_value = (
            self._cell_text(row, 2)
            if self._is_warsim_typename_widget(widget)
            else raw_item_text
        )
        if self._is_warsim_federate(federate):
            if not self._is_warsim_typename_widget(widget):
                combo = self._create_warsim_combobox()
                with self._suspend_table_item_signals():
                    self.table.setCellWidget(row, 2, combo)
                widget = combo
            combo_widget = widget if isinstance(widget, QComboBox) else None
            if combo_widget is None:
                return
            canonical = self._canonical_warsim_typename(current_value)
            if canonical is None:
                canonical = ""
            self._set_combo_value(combo_widget, canonical)
            with self._suspend_table_item_signals():
                target_item = self.table.item(row, 2)
                if target_item is None:
                    target_item = QTableWidgetItem(canonical)
                    self.table.setItem(row, 2, target_item)
                else:
                    target_item.setText(canonical)
        else:
            if self._is_warsim_typename_widget(widget):
                combo_widget = widget if isinstance(widget, QComboBox) else None
                text_value = (
                    combo_widget.currentText().strip()
                    if combo_widget
                    else current_value
                )
                with self._suspend_table_item_signals():
                    self.table.removeCellWidget(row, 2)
                    target_item = self.table.item(row, 2)
                    if target_item is None:
                        target_item = QTableWidgetItem(text_value)
                        self.table.setItem(row, 2, target_item)
                    else:
                        target_item.setText(text_value)

    def _cell_text(self, row: int, column: int) -> str:
        widget = self.table.cellWidget(row, column)
        if self._is_warsim_typename_widget(widget):
            combo = widget if isinstance(widget, QComboBox) else None
            if combo:
                return combo.currentText().strip()
        item = self.table.item(row, column)
        return (item.text() if item else "").strip()

    def _on_warsim_typename_changed(self, value: str) -> None:
        combo = self.sender()
        if not isinstance(combo, QComboBox) or not combo.property("warsim_typename_combo"):
            return
        view_pos = combo.mapTo(self.table.viewport(), combo.rect().center())
        index = self.table.indexAt(view_pos)
        row = index.row()
        if row < 0:
            return
        with self._suspend_table_item_signals():
            item = self.table.item(row, 2)
            if item is None:
                item = QTableWidgetItem(value)
                self.table.setItem(row, 2, item)
            else:
                item.setText(value)

    def _current_federate_filter(self) -> Optional[str]:
        text = self.federate_combo.currentText().strip()
        if not text or text.lower() == self._federate_label_all.lower():
            return None
        return text

    def _on_federate_filter_changed(self, _text: str) -> None:
        if self._refreshing_federate_combo:
            return
        self._reload_table()

    def _collect_federate_names(self) -> List[str]:
        names: Set[str] = set()
        if not self.model:
            return []
        ucl_parent = jfind(self.model.classdata, jtag("UnitClassList"))
        if ucl_parent is None:
            return []
        for alias in _iter_local(ucl_parent, "Alias"):
            fed = text(jfind(alias, "j:Federate"))
            if fed:
                names.add(fed.strip())
        return sorted(names, key=lambda s: s.upper())

    def _refresh_federate_combo(self) -> None:
        options = self._collect_federate_names()
        current_text = self.federate_combo.currentText().strip()
        self._refreshing_federate_combo = True
        block = self.federate_combo.blockSignals(True)
        self.federate_combo.clear()
        self.federate_combo.addItem(self._federate_label_all)
        for name in options:
            self.federate_combo.addItem(name)
        if current_text and current_text.lower() != self._federate_label_all.lower():
            idx = self.federate_combo.findText(
                current_text, Qt.MatchFlag.MatchFixedString
            )
            if idx >= 0:
                self.federate_combo.setCurrentIndex(idx)
            else:
                self.federate_combo.setEditText(current_text)
        else:
            self.federate_combo.setCurrentIndex(0)
        self.federate_combo.blockSignals(block)
        self._refreshing_federate_combo = False

    def set_model(self, model: Optional[ObsModel]):
        self.model = model
        has_model = model is not None
        self._set_enabled(has_model)
        self._refresh_federate_combo()
        if has_model:
            self._reload_table()
        else:
            self._clear_table()
            self.info_label.setText(
                "Load or create a JTDS OBS model to manage UnitClasses."
            )
            self.status_label.setText("No model loaded.")

    def _clear_table(self) -> None:
        self.table.setSortingEnabled(False)
        self.table.setRowCount(0)
        self._row_meta.clear()
        self.table.setSortingEnabled(True)

    def _reload_table(self) -> None:
        if not self.model:
            return
        self.table.setSortingEnabled(False)
        self.table.setRowCount(0)
        self._row_meta.clear()
        ucl_parent = jfind(self.model.classdata, jtag("UnitClassList"))
        if ucl_parent is None:
            self.info_label.setText(
                "This model does not have a UnitClassList yet. Saving new rows will create it."
            )
            self.status_label.setText("0 unit classes loaded.")
            self.table.setSortingEnabled(True)
            return
        federate_filter = self._current_federate_filter()
        rows: List[
            Tuple[
                str,
                str,
                str,
                str,
                str,
                ET._Element,
                Optional[ET._Element],
                Optional[str],
            ]
        ] = []
        seen: Set[str] = set()
        for ucl in list(ucl_parent):
            if _local(ucl.tag) != "UnitClass":
                continue
            agg = self._extract_aggregate_name(ucl)
            if not agg:
                continue
            key = agg.upper()
            if key in seen:
                continue
            seen.add(key)
            code = (text(jfind(ucl, "j:MilStd2525CCode")) or "").strip()
            ude = self._extract_unit_dis_enum(ucl)
            if federate_filter:
                alias = self._find_alias_for_federate(ucl, federate_filter)
                federate_val, typename_val = self._alias_fields(alias)
                rows.append(
                    (
                        agg,
                        federate_val,
                        typename_val,
                        code,
                        ude,
                        ucl,
                        alias,
                        federate_filter,
                    )
                )
            else:
                federate_val, typename_val, alias = self._extract_alias_values(ucl)
                rows.append(
                    (
                        agg,
                        federate_val,
                        typename_val,
                        code,
                        ude,
                        ucl,
                        alias,
                        federate_val or None,
                    )
                )
        rows.sort(key=lambda item: item[0].upper())
        for agg, federate, typename, code, ude, ucl, alias_elem, alias_fed in rows:
            self._append_row(
                agg,
                federate,
                typename,
                code,
                ude,
                ucl,
                is_new=False,
                alias_element=alias_elem,
                alias_federate=alias_fed,
            )
        self.table.setSortingEnabled(True)
        self.info_label.setText(
            "Edit Federate, TypeName, 2525C code, and UnitDisEnumeration values. Add rows to create new UnitClasses."
        )
        self.status_label.setText(f"{len(rows)} unit classes loaded.")

    def _next_row_key(self, element: Optional[ET._Element]) -> str:
        if element is not None:
            return f"ucl-{id(element)}"
        self._new_row_seq += 1
        return f"new-{self._new_row_seq}"

    def _append_row(
        self,
        name: str,
        federate: str,
        typename: str,
        code2525: str,
        unit_enum: str,
        element: Optional[ET._Element],
        *,
        is_new: bool,
        alias_element: Optional[ET._Element] = None,
        alias_federate: Optional[str] = None,
    ) -> None:
        row = self.table.rowCount()
        self.table.insertRow(row)
        key = self._next_row_key(element)
        with self._suspend_table_item_signals():
            name_item = QTableWidgetItem(name)
            name_item.setData(Qt.ItemDataRole.UserRole, key)
            if not is_new:
                name_item.setFlags(name_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.table.setItem(row, 0, name_item)
            for col, value in enumerate(
                [federate, typename, code2525, unit_enum], start=1
            ):
                item = QTableWidgetItem(value or "")
                self.table.setItem(row, col, item)
        self._row_meta[key] = _UnitClassRowMeta(
            element=element,
            alias_element=alias_element,
            alias_federate=alias_federate,
            is_new=is_new,
        )
        self._update_row_typename_editor(row)

    def _table_rows(self) -> List[Tuple[str, str, str, str, str]]:
        rows: List[Tuple[str, str, str, str, str]] = []
        federate_filter = self._current_federate_filter()
        for row in range(self.table.rowCount()):
            agg_item = self.table.item(row, 0)
            fed_item = self.table.item(row, 1)
            code_item = self.table.item(row, 3)
            enum_item = self.table.item(row, 4)
            agg = (agg_item.text() if agg_item else "").strip()
            federate = (fed_item.text() if fed_item else "").strip()
            if federate_filter:
                federate = federate_filter
            typename = self._cell_text(row, 2)
            code2525 = (code_item.text() if code_item else "").strip()
            unit_enum = (enum_item.text() if enum_item else "").strip()
            rows.append((agg, federate, typename, code2525, unit_enum))
        return rows

    def _name_to_row(self) -> Dict[str, int]:
        mapping: Dict[str, int] = {}
        for row in range(self.table.rowCount()):
            item = self.table.item(row, 0)
            if not item:
                continue
            name = (item.text() or "").strip()
            if not name:
                continue
            mapping[name.upper()] = row
        return mapping

    def _set_cell_text(self, row: int, column: int, value: str) -> None:
        refresh_editor = column == 1
        with self._suspend_table_item_signals():
            widget = self.table.cellWidget(row, column)
            if column == 2 and self._is_warsim_typename_widget(widget):
                combo = widget if isinstance(widget, QComboBox) else None
                canonical = self._canonical_warsim_typename(value)
                if canonical is None:
                    canonical = ""
                if combo is not None:
                    self._set_combo_value(combo, canonical)
                value = canonical
            item = self.table.item(row, column)
            if item is None:
                item = QTableWidgetItem(value)
                self.table.setItem(row, column, item)
            else:
                item.setText(value)
        if refresh_editor:
            self._update_row_typename_editor(row)

    def _extract_aggregate_name(self, ucl: ET._Element) -> str:
        agg = text(jfind(ucl, "j:AggregateName"))
        if agg:
            return agg.strip()
        name = text(jfind(ucl, "j:Name"))
        return name.strip() if name else ""

    def _first_alias(self, node: ET._Element) -> Optional[ET._Element]:
        alias_list = jfind(node, "j:AliasList")
        if alias_list is not None:
            for child in list(alias_list):
                if _local(child.tag) == "Alias":
                    return child
        return next((child for child in _iter_local(node, "Alias")), None)

    def _alias_fields(self, alias: Optional[ET._Element]) -> Tuple[str, str]:
        federate = ""
        typename = ""
        if alias is None:
            return federate, typename
        fed_elem = next(
            (child for child in list(alias) if _local(child.tag) == "Federate" and child.text),
            None,
        )
        if fed_elem is not None:
            federate = fed_elem.text.strip()
        type_elem = next(
            (child for child in list(alias) if _local(child.tag) == "TypeName" and child.text),
            None,
        )
        if type_elem is not None:
            typename = type_elem.text.strip()
        return federate, typename

    def _alias_map(self, node: ET._Element) -> Dict[str, ET._Element]:
        mapping: Dict[str, ET._Element] = {}
        alias_list = jfind(node, "j:AliasList")
        candidates: List[ET._Element] = []
        if alias_list is not None:
            for child in list(alias_list):
                if _local(child.tag) == "Alias":
                    candidates.append(child)
        else:
            candidates.extend(_iter_local(node, "Alias"))
        for alias in candidates:
            federate = (text(jfind(alias, "j:Federate")) or "").strip()
            if federate:
                mapping[federate.upper()] = alias
        return mapping

    def _find_alias_for_federate(
        self, node: ET._Element, federate: Optional[str]
    ) -> Optional[ET._Element]:
        if not federate:
            return None
        return self._alias_map(node).get(federate.strip().upper())

    def _extract_alias_values(
        self, node: ET._Element
    ) -> Tuple[str, str, Optional[ET._Element]]:
        alias = self._first_alias(node)
        federate, typename = self._alias_fields(alias)
        return federate, typename, alias

    def _extract_unit_dis_enum(self, ucl: ET._Element) -> str:
        ude = jfind(ucl, "j:UnitDisEnumeration")
        if ude is None:
            for child in list(ucl):
                if _local(child.tag) == "UnitDisEnumeration":
                    ude = child
                    break
        if ude is None:
            return ""
        attrs: Dict[str, str] = {}
        for key in _UNIT_DIS_ENUM_ATTRS:
            raw = ude.get(key, "")
            if raw:
                attrs[key] = raw
        if not any(attrs.values()):
            for key, legacy in zip(_UNIT_DIS_ENUM_ATTRS, _DISCODE_ATTRS):
                raw = ude.get(legacy, "")
                if raw:
                    attrs[key] = raw
        if not any(attrs.values()):
            return ""
        return _format_unit_dis_enum(attrs)

    def _on_add_row(self) -> None:
        if not self.model:
            return
        self._append_row(
            "",
            "",
            "",
            "",
            "",
            None,
            is_new=True,
            alias_federate=self._current_federate_filter(),
        )
        self.status_label.setText(
            f"{self.table.rowCount()} row(s) in table. New rows require an AggregateName before saving."
        )

    def _on_pick_2525_code(self) -> None:
        if not self.model:
            QMessageBox.information(
                self, "Pick 2525C", "Load or create a JTDS OBS model first."
            )
            return
        row = self.table.currentRow()
        if row < 0:
            QMessageBox.information(
                self, "Pick 2525C", "Select a row before choosing a code."
            )
            return
        try:
            entries = _load_milstd2525_library()
            lookup = _load_sidc_attribute_lookup()
        except (FileNotFoundError, ValueError) as exc:
            QMessageBox.warning(self, "Pick 2525C", str(exc))
            return
        if not entries:
            QMessageBox.warning(
                self,
                "Pick 2525C",
                "No MIL-STD-2525C entries were available in the Symbols tables.",
            )
            return
        code_item = self.table.item(row, 3)
        current_code = (code_item.text() if code_item else "").strip()
        dlg = MilStd2525PickerDialog(
            self, entries, lookup, initial_code=current_code or None
        )
        if dlg.exec() == QDialog.DialogCode.Accepted:
            code_value = dlg.selected_code()
            if not code_value:
                return
            entry = dlg.selected_entry()
            self._set_cell_text(row, 3, code_value)
            type_text = self._cell_text(row, 2)
            if not type_text and entry:
                self._set_cell_text(row, 2, entry.title)
            self.status_label.setText(
                f"Row {row + 1}: Applied MIL-STD-2525C code {code_value}."
            )

    def _on_sort(self) -> None:
        column = 0 if self.sort_combo.currentIndex() == 0 else 3
        self.table.sortItems(column)

    def _on_export(self) -> None:
        rows = self._table_rows()
        if not rows:
            QMessageBox.information(
                self, "Export Unit Classes", "No rows to export."
            )
            return
        path, _ = QFileDialog.getSaveFileName(
            self,
            "Export Unit Classes",
            "UnitClasses.csv",
            "CSV Files (*.csv);;All Files (*)",
        )
        if not path:
            return
        if not path.lower().endswith(".csv"):
            path += ".csv"
        try:
            with open(path, "w", newline="", encoding="utf-8") as handle:
                writer = csv.writer(handle)
                writer.writerow(
                    ["AggregateName", "Federate", "TypeName", "2525C", "UnitDisEnumeration"]
                )
                for record in rows:
                    writer.writerow(record)
        except Exception as exc:
            QMessageBox.warning(self, "Export Unit Classes", f"Failed to export: {exc}")
            return
        QMessageBox.information(
            self,
            "Export Unit Classes",
            f"Exported {len(rows)} row(s) to {path}",
        )

    def _resolve_field_name(self, fieldnames: Optional[List[str]], target: str) -> Optional[str]:
        if not fieldnames:
            return None
        normalized_target = target.replace(" ", "").lower()
        for name in fieldnames:
            normalized = name.replace(" ", "").lower()
            if normalized == normalized_target:
                return name
        return None

    def _on_import(self) -> None:
        if not self.model:
            QMessageBox.information(
                self,
                "Import Unit Classes",
                "Load or create a JTDS OBS model before importing.",
            )
            return
        filter_value = self._current_federate_filter()
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Import Unit Classes",
            "",
            "CSV Files (*.csv);;All Files (*)",
        )
        if not path:
            return
        try:
            with open(path, "r", newline="", encoding="utf-8-sig") as handle:
                reader = csv.DictReader(handle)
                fieldnames = reader.fieldnames or []
                name_field = self._resolve_field_name(fieldnames, "AggregateName")
                if not name_field:
                    raise ValueError("CSV is missing an 'AggregateName' column.")
                fed_field = self._resolve_field_name(fieldnames, "Federate")
                type_field = self._resolve_field_name(fieldnames, "TypeName")
                code_field = self._resolve_field_name(fieldnames, "2525C")
                enum_field = self._resolve_field_name(fieldnames, "UnitDisEnumeration")
                incoming: List[Dict[str, str]] = []
                seen_names: Set[str] = set()
                issues: List[str] = []
                line_num = 1
                for row in reader:
                    line_num += 1
                    agg = (row.get(name_field) or "").strip()
                    if not agg:
                        issues.append(f"Row {line_num}: Missing AggregateName; skipped.")
                        continue
                    upper = agg.upper()
                    if upper in seen_names:
                        issues.append(
                            f"Row {line_num}: Duplicate AggregateName '{agg}' within file; skipped."
                        )
                        continue
                    federate = (
                        (row.get(fed_field) or "").strip() if fed_field else ""
                    )
                    if filter_value:
                        federate = filter_value
                    typename = (
                        (row.get(type_field) or "").strip() if type_field else ""
                    )
                    code2525 = (
                        (row.get(code_field) or "").strip() if code_field else ""
                    )
                    enum_raw = (
                        (row.get(enum_field) or "").strip() if enum_field else ""
                    )
                    enum_text = ""
                    if enum_raw:
                        try:
                            enum_attrs = _parse_unit_dis_enum_to_attrs(enum_raw)
                            enum_text = _format_unit_dis_enum(enum_attrs)
                        except ValueError:
                            issues.append(
                                f"Row {line_num}: Invalid UnitDisEnumeration '{enum_raw}'; skipped."
                            )
                            continue
                    incoming.append(
                        {
                            "name": agg,
                            "name_upper": upper,
                            "federate": federate,
                            "typename": typename,
                            "code2525": code2525,
                            "enum_raw": enum_text,
                        }
                    )
                    seen_names.add(upper)
        except ValueError as exc:
            QMessageBox.warning(self, "Import Unit Classes", str(exc))
            return
        except Exception as exc:
            QMessageBox.warning(
                self, "Import Unit Classes", f"Failed to import: {exc}"
            )
            return
        if not incoming:
            msg = "No valid rows were imported."
            if issues:
                preview = "\n".join(issues[:5])
                if len(issues) > 5:
                    preview += "\n..."
                msg += f"\nIssues:\n{preview}"
            QMessageBox.information(self, "Import Unit Classes", msg)
            return
        name_to_row = self._name_to_row()
        updated = 0
        created = 0
        sorting_prev = self.table.isSortingEnabled()
        self.table.setSortingEnabled(False)
        try:
            for entry in incoming:
                idx = name_to_row.get(entry["name_upper"])
                if idx is not None:
                    self._set_cell_text(idx, 1, entry["federate"])
                    self._set_cell_text(idx, 2, entry["typename"])
                    self._set_cell_text(idx, 3, entry["code2525"])
                    self._set_cell_text(idx, 4, entry["enum_raw"])
                    updated += 1
                else:
                    self._append_row(
                        entry["name"],
                        entry["federate"],
                        entry["typename"],
                        entry["code2525"],
                        entry["enum_raw"],
                        None,
                        is_new=True,
                        alias_federate=filter_value,
                    )
                    created += 1
        finally:
            self.table.setSortingEnabled(sorting_prev)
        total_rows = self.table.rowCount()
        self.status_label.setText(
            f"Table now has {total_rows} row(s). Imported {created} new, updated {updated}."
        )
        summary = (
            f"Import complete. Added {created} row(s), updated {updated} row(s)."
        )
        if issues:
            preview = "\n".join(issues[:5])
            if len(issues) > 5:
                preview += "\n..."
            summary += f"\nWarnings:\n{preview}"
        QMessageBox.information(self, "Import Unit Classes", summary)
        self._refresh_federate_combo()

    def _ensure_child_text(self, parent: ET._Element, tag: str, value: str) -> ET._Element:
        node = jfind(parent, f"j:{tag}")
        if node is None:
            node = ET.SubElement(parent, jtag(tag))
        node.text = value
        return node

    def _update_alias(
        self,
        ucl: ET._Element,
        federate: str,
        typename: str,
        *,
        alias_element: Optional[ET._Element],
        target_federate: Optional[str],
    ) -> None:
        alias_list = jfind(ucl, "j:AliasList")
        if alias_list is None:
            alias_list = ET.SubElement(ucl, jtag("AliasList"))
        desired_federate = (federate or "").strip()
        target_key = (target_federate or "").strip()
        alias: Optional[ET._Element] = None
        if target_key:
            alias = self._find_alias_for_federate(ucl, target_key)
        if alias is None and alias_element is not None:
            existing_fed, _ = self._alias_fields(alias_element)
            if not target_key or existing_fed.strip().upper() == target_key.upper():
                alias = alias_element
        if alias is None and not target_key and desired_federate:
            alias = self._find_alias_for_federate(ucl, desired_federate)
        if alias is None and not target_key:
            alias = jfind(alias_list, "j:Alias")
        if alias is None:
            alias = ET.SubElement(alias_list, jtag("Alias"))
        fed = jfind(alias, "j:Federate")
        if fed is None:
            fed = ET.SubElement(alias, jtag("Federate"))
        fed.text = desired_federate or (target_key or None)
        tname = jfind(alias, "j:TypeName")
        if tname is None:
            tname = ET.SubElement(alias, jtag("TypeName"))
        tname.text = typename or None

    def _update_unit_dis_enum(
        self, ucl: ET._Element, attrs: Optional[Dict[str, str]]
    ) -> None:
        ude = jfind(ucl, "j:UnitDisEnumeration")
        if attrs is None or not attrs:
            if ude is not None and ude in list(ucl):
                ucl.remove(ude)
            return
        if ude is None:
            ude = ET.SubElement(ucl, jtag("UnitDisEnumeration"))
        for legacy in set(_DISCODE_ATTRS) - set(_UNIT_DIS_ENUM_ATTRS):
            if legacy in ude.attrib:
                del ude.attrib[legacy]
        for key in _UNIT_DIS_ENUM_ATTRS:
            value = attrs.get(key, "") if attrs else ""
            if value:
                ude.set(key, value)
            elif key in ude.attrib:
                del ude.attrib[key]

    def _collect_existing_name_keys(self) -> Set[str]:
        keys: Set[str] = set()
        for row in range(self.table.rowCount()):
            item = self.table.item(row, 0)
            if not item:
                continue
            key = _normalize_user_data(item.data(Qt.ItemDataRole.UserRole))
            if key is None:
                continue
            meta = self._row_meta.get(key)
            if not meta or meta.is_new:
                continue
            name = (item.text() or "").strip()
            if name:
                keys.add(name.upper())
        return keys

    def _on_save(self) -> None:
        if not self.model:
            return
        ucl_parent = jfind(self.model.classdata, jtag("UnitClassList"))
        if ucl_parent is None:
            ucl_parent = ET.SubElement(self.model.classdata, jtag("UnitClassList"))
        errors: List[str] = []
        created = 0
        updated = 0
        existing_name_keys = self._collect_existing_name_keys()
        filter_value = self._current_federate_filter()
        for row in range(self.table.rowCount()):
            name_item = self.table.item(row, 0)
            if name_item is None:
                continue
            key = _normalize_user_data(name_item.data(Qt.ItemDataRole.UserRole))
            meta = self._row_meta.get(key)
            if not meta:
                continue
            aggregate = (name_item.text() or "").strip()
            fed_item = self.table.item(row, 1)
            federate = (fed_item.text() if fed_item else "").strip()
            typename = self._cell_text(row, 2)
            code_item = self.table.item(row, 3)
            code2525 = (code_item.text() if code_item else "").strip()
            enum_item = self.table.item(row, 4)
            enum_raw = (enum_item.text() if enum_item else "").strip()
            enum_attrs: Optional[Dict[str, str]] = None
            if enum_raw:
                try:
                    enum_attrs = _parse_unit_dis_enum_to_attrs(enum_raw)
                except ValueError:
                    errors.append(
                        f"Row {row + 1}: UnitDisEnumeration must be 'kind.domain.country.echelon.type.subtype.modifier'."
                    )
                    continue
            if meta.is_new:
                if not aggregate:
                    errors.append(
                        f"Row {row + 1}: AggregateName is required for new entries."
                    )
                    continue
                key_upper = aggregate.upper()
                if key_upper in existing_name_keys:
                    errors.append(
                        f"Row {row + 1}: AggregateName '{aggregate}' already exists."
                    )
                    continue
                existing_name_keys.add(key_upper)
                ucl = ET.SubElement(ucl_parent, jtag("UnitClass"))
                self._ensure_child_text(ucl, "AggregateName", aggregate)
                meta.element = ucl
                meta.is_new = False
                created += 1
            else:
                ucl = meta.element
                if ucl is None:
                    continue
                if aggregate:
                    self._ensure_child_text(ucl, "AggregateName", aggregate)
                updated += 1
            if meta.element is None:
                continue
            if code2525:
                self._ensure_child_text(meta.element, "MilStd2525CCode", code2525)
            alias_target = filter_value or meta.alias_federate or federate
            self._update_alias(
                meta.element,
                federate,
                typename,
                alias_element=meta.alias_element,
                target_federate=alias_target,
            )
            self._update_unit_dis_enum(meta.element, enum_attrs)
        if errors:
            preview = "\n".join(errors[:5])
            if len(errors) > 5:
                preview += "\n..."
            QMessageBox.warning(self, "Save Unit Classes", preview)
            return
        QMessageBox.information(
            self,
            "Save Unit Classes",
            f"UnitClass changes saved. Updated {updated}, created {created}.",
        )
        self._reload_table()
        self._refresh_federate_combo()

class GenerateTab(QWidget):

    def __init__(self, parent):
        super().__init__(parent)
        self.model: Optional[ObsModel] = None
        self._preview_dialog: Optional[ModelPreviewDialog] = None
        self._build()

    def _build(self):
        lay = QVBoxLayout(self)
        opt_box = QGroupBox("Generation Options")
        form = QFormLayout(opt_box)
        self.cmb_consol = QComboBox()
        self.cmb_consol.addItems(["None"] + CONSOLIDATION_LEVELS)
        form.addRow("Consolidate lower echelon into parent:", self.cmb_consol)
        self.btn_autofill = QPushButton("Auto-fill LvcId (unique)")
        self.btn_autofill.clicked.connect(self.on_autofill)
        self.btn_consolidate = QPushButton("Run Consolidation")
        self.btn_consolidate.clicked.connect(self.on_consolidate)
        self.btn_fix_flags = QPushButton("Fix Empty Flags")
        self.btn_fix_flags.clicked.connect(self.on_fix_flags)
        self.btn_import_obs = QPushButton("Import Units from OBS...")
        self.btn_import_obs.clicked.connect(self.on_import_obs)
        self.btn_preview = QPushButton("Preview Model")
        self.btn_preview.clicked.connect(self.on_preview)
        self.btn_save = QPushButton("Save as New JTDS OBS XML.")
        self.btn_save.clicked.connect(self.on_save)
        btn_row = QHBoxLayout()
        btn_row.addWidget(self.btn_autofill)
        btn_row.addWidget(self.btn_consolidate)
        btn_row.addWidget(self.btn_fix_flags)
        btn_row.addWidget(self.btn_import_obs)
        btn_row.addWidget(self.btn_preview)
        btn_row.addWidget(self.btn_save)
        self.log = QTextEdit()
        self.log.setReadOnly(True)
        self.log.setMinimumHeight(180)
        lay.addWidget(opt_box)
        lay.addLayout(btn_row)
        side_grp = QGroupBox("Consolidate Sides")
        self.side_checks_layout = QHBoxLayout(side_grp)
        self.side_checks: Dict[str, QCheckBox] = {}
        for label in ["BLUFOR", "OPFOR", "NEUTRAL", "CIVILIAN", "UNKNOWN"]:
            chk = QCheckBox(label)
            chk.setChecked(True)
            self.side_checks[label] = chk
            self.side_checks_layout.addWidget(chk)
        self.side_checks_layout.addStretch()
        lay.addWidget(side_grp)
        lay.addWidget(self.log)
        grp = QGroupBox("Build Class Lists")
        gl = QGridLayout(grp)
        gl.addWidget(QLabel("EntityCompositionClass master (.xlsx):"), 0, 0)
        self.ecc_edit = QLineEdit()
        btn_ecc = QPushButton("Browse...")
        btn_ecc.clicked.connect(self.on_browse_ecc)
        gl.addWidget(self.ecc_edit, 0, 1)
        gl.addWidget(btn_ecc, 0, 2)
        gl.addWidget(QLabel("UnitClass master (.xlsx):"), 1, 0)
        self.ucl_edit = QLineEdit()
        btn_ucl = QPushButton("Browse...")
        btn_ucl.clicked.connect(self.on_browse_ucl)
        gl.addWidget(self.ucl_edit, 1, 1)
        gl.addWidget(btn_ucl, 1, 2)
        self.btn_build = QPushButton("Build Classes")
        self.btn_build.clicked.connect(self.on_build_classes)
        gl.addWidget(self.btn_build, 2, 0, 1, 3)
        lay.addWidget(grp)
        self._enable(False)

    def on_browse_ecc(self):
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Open EntityCompositionClass master",
            "",
            "Excel (*.xlsx);;All Files (*)",
        )
        if path:
            self.ecc_edit.setText(path)

    def on_browse_ucl(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Open UnitClass master", "", "Excel (*.xlsx);;All Files (*)"
        )
        if path:
            self.ucl_edit.setText(path)

    def on_build_classes(self):
        if not self.model:
            QMessageBox.information(self, "No model", "Load or create a model first.")
            return
        ecc_path = self.ecc_edit.text().strip() if hasattr(self, "ecc_edit") else ""
        ucl_path = self.ucl_edit.text().strip() if hasattr(self, "ucl_edit") else ""
        try:
            cnt_ecc = (
                build_entitycomposition_classlist_from_xlsx(self.model, ecc_path)
                if ecc_path
                else 0
            )
            cnt_ucl = (
                build_unitclass_list_from_xlsx(self.model, ucl_path) if ucl_path else 0
            )
        except Exception as e:
            QMessageBox.warning(self, "Build Classes", f"Failed: {e}")
            return
        self.log.append(
            f"Built EntityCompositionClassList: {cnt_ecc} items; UnitClassList: {cnt_ucl} items."
        )

    def _enable(self, ok: bool):
        for w in [
            self.btn_autofill,
            self.btn_consolidate,
            self.btn_fix_flags,
            self.btn_import_obs,
            self.btn_preview,
            self.btn_save,
            self.cmb_consol,
        ]:
            w.setEnabled(ok)
        for chk in getattr(self, "side_checks", {}).values():
            chk.setEnabled(ok)

    def _refresh_side_checks(self) -> None:
        if not getattr(self, "side_checks_layout", None):
            return
        existing = set(getattr(self, "side_checks", {}).keys())
        if not existing and self.side_checks_layout.count() == 0:
            return
        stretch_index = (
            self.side_checks_layout.count() - 1
            if self.side_checks_layout.count()
            else None
        )
        extra_sides: Set[str] = set()
        if self.model:
            for unit in list(self.model.unit_list):
                side = (text(jfind(unit, "j:SideSuperior")) or "").strip()
                if side:
                    extra_sides.add(side.upper())
        for side in sorted(extra_sides):
            if side in existing:
                continue
            chk = QCheckBox(side)
            chk.setChecked(True)
            chk.setEnabled(self.btn_autofill.isEnabled())
            self.side_checks[side] = chk
            existing.add(side)
            if stretch_index is not None and stretch_index >= 0:
                self.side_checks_layout.insertWidget(stretch_index, chk)
                stretch_index += 1
            else:
                self.side_checks_layout.addWidget(chk)
        if self.side_checks_layout.count() == len(self.side_checks):
            self.side_checks_layout.addStretch()

    def on_import_obs(self) -> None:
        if self.model is None:
            choice = QMessageBox.question(
                self,
                "No model loaded",
                "No JTDS OBS model is currently loaded. Create a new blank model to import into?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No,
            )
            if choice != QMessageBox.StandardButton.Yes:
                return
            self.set_model(_new_blank_model())
            if self.model is None:
                return
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Import Units from OBS",
            "",
            "OBS XML or ZIP (*.xml *.zip);;All Files (*)",
        )
        if not path:
            return
        try:
            source_model = load_obs_xml(path)
        except Exception as exc:
            traceback.print_exc()
            QMessageBox.critical(
                self, "Import Units", f"Failed to load source OBS: {exc}"
            )
            return
        picker = ObsUnitSelectionDialog(self, source_model)
        if picker.exec() != QDialog.DialogCode.Accepted:
            return
        selected_units = picker.selected_root_units()
        if not selected_units:
            QMessageBox.information(
                self, "Import Units", "No units were selected to import."
            )
            return
        lvc_to_unit_src, parent_to_children_src = collect_unit_maps(source_model)
        subtree_units: List[ET._Element] = []
        visited: Set[int] = set()

        def collect(unit: ET._Element) -> None:
            key = id(unit)
            if key in visited:
                return
            visited.add(key)
            subtree_units.append(unit)
            lvc = (text(jfind(unit, "j:LvcId")) or "").strip()
            children = parent_to_children_src.get(lvc, []) if lvc else []
            for child in sorted(children, key=unit_sort_key):
                collect(child)

        for unit in selected_units:
            collect(unit)
        source_sides: Set[str] = set()
        missing_side = False
        for unit in subtree_units:
            side_val = (text(jfind(unit, "j:SideSuperior")) or "").strip()
            if side_val:
                source_sides.add(side_val)
            else:
                missing_side = True
        available_sides: Set[str] = set(source_sides)
        if self.model is not None:
            for unit in _iter_local(self.model.unit_list, "Unit"):
                side_val = (text(jfind(unit, "j:SideSuperior")) or "").strip()
                if side_val:
                    available_sides.add(side_val)
        available_sides.update(DEFAULT_SIDE_ORDER)
        choices = ["Keep source values"]
        sorted_sides = sorted(
            {s for s in available_sides if s}, key=lambda s: s.upper()
        )
        for side in sorted_sides:
            choices.append(f"Force side {side}")
        override_side: Optional[str] = None
        fallback_side: Optional[str] = None
        if choices:
            selection, ok = QInputDialog.getItem(
                self,
                "Side Assignment",
                "SideSuperior handling for imported units:",
                choices,
                0,
                False,
            )
            if not ok:
                return
            if selection.startswith("Force side "):
                override_side = selection[len("Force side ") :].strip()
        if override_side is None and missing_side and len(source_sides) == 1:
            fallback_side = next(iter(source_sides))
        try:
            stats = merge_obs_units(
                self.model,
                source_model,
                selected_units,
                override_side=override_side,
                fallback_side=fallback_side,
            )
        except Exception as exc:
            traceback.print_exc()
            QMessageBox.critical(self, "Import Units", f"Failed to merge models: {exc}")
            return
        self.set_model(self.model)
        source_name = Path(path).name
        lines = [
            f"Imported {stats['units_added']} unit(s) and {stats['entities_added']} entity composition(s) from {source_name}.",
        ]
        if override_side:
            lines.append(f" - Forced SideSuperior='{override_side}' on imported units.")
        elif fallback_side:
            lines.append(
                f" - Missing SideSuperior values defaulted to '{fallback_side}'."
            )
        elif missing_side and not fallback_side:
            lines.append(
                " - Some imported units still lack SideSuperior; review as needed."
            )
        if stats.get("unit_lvc_conflicts"):
            lines.append(
                f" - Resolved {stats['unit_lvc_conflicts']} duplicate unit LvcId value(s)."
            )
        if stats.get("entity_lvc_conflicts"):
            lines.append(
                f" - Resolved {stats['entity_lvc_conflicts']} duplicate entity composition LvcId value(s)."
            )
        if stats.get("unit_superiors_cleared"):
            lines.append(
                f" - Cleared {stats['unit_superiors_cleared']} missing UnitSuperior reference(s)."
            )
        if stats.get("entities_skipped_orphaned"):
            lines.append(
                f" - Skipped {stats['entities_skipped_orphaned']} entity composition(s) without a parent unit in scope."
            )
        if stats.get("ecc_added") or stats.get("ucl_added"):
            lines.append(
                f" - Class list merge: EntityCompositionClass +{stats['ecc_added']}, UnitClass +{stats['ucl_added']}."
            )
        if stats.get("sides_added"):
            lines.append(
                f" - Added {stats['sides_added']} side definition(s) to Scenario/SideList."
            )
        if (
            stats.get("unit_ids_assigned")
            or stats.get("unit_timestamps_assigned")
            or stats.get("unit_sides_assigned")
        ):
            lines.append(
                f" - Unit metadata normalized: ids={stats['unit_ids_assigned']}, timestamps={stats['unit_timestamps_assigned']}, sides set={stats['unit_sides_assigned']}."
            )
        self.log.append("\n".join(lines))
        QMessageBox.information(self, "Import Units", lines[0])

    def set_model(self, model: Optional[ObsModel]):
        self.model = model
        self._enable(model is not None)
        self._refresh_side_checks()

    def on_autofill(self):
        if not self.model:
            return
        count = autofill_unique_lvcids(self.model)
        self.log.append(f"Auto-filled/Fixed LvcId fields: {count}")

    def on_consolidate(self):
        if not self.model:
            return
        mode = self.cmb_consol.currentText()
        selected_sides = {
            side
            for side, chk in getattr(self, "side_checks", {}).items()
            if chk.isChecked()
        }
        if not selected_sides:
            QMessageBox.information(
                self,
                "Consolidate",
                "Select at least one side before running consolidation.",
            )
            return
        stats = consolidate_units(self.model, mode, selected_sides)
        sides_text = ", ".join(sorted(selected_sides))
        msg = f"Consolidation '{mode}' (sides: {sides_text}): removed units {stats['removed_units']}, moved ECs {stats['moved_entity_compositions']}, reparented units {stats['relinked_units']}"
        skipped = stats.get("skipped_units", 0)
        if skipped:
            msg += f", skipped {skipped} unit(s) without a surviving parent"
        self.log.append(msg)

    def on_fix_flags(self):
        if not self.model:
            return
        try:
            stats = fix_empty_flags(self.model)
        except Exception as e:
            traceback.print_exc()
            QMessageBox.warning(self, "Fix Empty Flags", f"Failed: {e}")
            return
        total = stats.get("created", 0)
        per_side = stats.get("per_side", {})
        skipped = stats.get("skipped", 0)
        parts = [f"{side}={count}" for side, count in per_side.items() if count]
        if total:
            msg = f"Fix Empty Flags: created {total} EntityCompositions"
            if parts:
                msg += f" ({', '.join(parts)})"
            if skipped:
                msg += f"; skipped {skipped} units without side mapping"
        else:
            msg = "Fix Empty Flags: no missing UnitSuperior references found"
            if skipped:
                msg += f" (skipped {skipped} units without side mapping)"
        self.log.append(msg)

    def on_preview(self):
        if not self.model:
            QMessageBox.information(self, "No model", "Load or create a model first.")
            return
        if self._preview_dialog is None:
            self._preview_dialog = ModelPreviewDialog(self)
        self._preview_dialog.set_model(self.model)
        self._preview_dialog.exec()

    def on_save(self):
        if not self.model:
            return
        out, _ = QFileDialog.getSaveFileName(
            self, "Save JTDS OBS XML", "JTDS_Scrubbed.xml", "OBS XML (*.xml)"
        )
        if not out:
            return
        try:
            save_xml(self.model, out)
        except Exception as e:
            QMessageBox.critical(self, "Save Error", str(e))
        else:
            QMessageBox.information(self, "Saved", f"Saved to: {out}")


# === DRAGON Enhanced Import Helpers ===


def _pick_col(cols, *cands):

    up = {str(c).strip().upper(): c for c in cols}
    for cand in cands:
        if isinstance(cand, (list, tuple)):
            for x in cand:
                if x and x.upper() in up:
                    return up[x.upper()]
        elif cand and cand.upper() in up:
            return up[cand.upper()]
    # loose contains
    for c in cols:
        cu = str(c).upper()
        for cand in cands:
            if isinstance(cand, str) and cand and cand.upper() in cu:
                return c
            if isinstance(cand, (list, tuple)):
                for x in cand:
                    if x and str(x).upper() in cu:
                        return c
    return None


def _norm(s):

    return str(s).strip() if s is not None else ""


def _parse_den_to_attrs(s):

    import re as _re

    if not s:
        return {}
    parts = _re.split(r"[.\s]+", str(s).strip())
    keys = ["kind", "domain", "country", "category", "subcategory", "specific", "extra"]
    out = {}
    for i, k in enumerate(keys):
        if i < len(parts) and parts[i].isdigit():
            out[k] = parts[i]
    return out


@dataclass
class DragonUnitNode:

    uid: str
    name: str
    template: str
    parent_uid: Optional[str]
    unit_class: str
    ms2525: str
    order: int
    children: List["DragonUnitNode"]

    def display_name(self) -> str:
        base = self.name or self.uid
        if self.unit_class:
            return f"{base} [{self.unit_class}]"
        return base


@dataclass
class DragonWorkbookTree:

    roots: List[DragonUnitNode]
    node_map: Dict[str, DragonUnitNode]

    def total_units(self) -> int:
        return len(self.node_map)

    def gather_subtree_uids(self, root_uids: Set[str]) -> Set[str]:
        acc: Set[str] = set()
        for uid in root_uids:
            node = self.node_map.get(uid.upper())
            if node is not None:
                _collect_dragon_uids(node, acc)
        return acc


def _collect_dragon_uids(node: DragonUnitNode, acc: Set[str]) -> None:

    key = node.uid.strip().upper()
    if not key or key in acc:
        pass
    else:
        acc.add(key)
    for child in node.children:
        _collect_dragon_uids(child, acc)


def parse_dragon_unit_hierarchy(xlsx_path: str) -> DragonWorkbookTree:

    try:
        xls = pd.ExcelFile(xlsx_path)
    except Exception as exc:
        raise ValueError(f"Failed to open DRAGON workbook: {exc}") from exc
    if "UNIT INFO" not in xls.sheet_names:
        raise ValueError("UNIT INFO sheet not found in DRAGON workbook.")
    df = xls.parse("UNIT INFO")
    df.columns = [str(c).strip() for c in df.columns]
    col_uid = _pick_col(df.columns, "UID", "UNIT UID", "ID")
    if not col_uid:
        raise ValueError("UID column not found in UNIT INFO sheet.")
    col_name = _pick_col(
        df.columns, "UNIT", "UNIT NAME", "NAME", "UNIT TITLE", "DISPLAY NAME"
    )
    col_puid = _pick_col(df.columns, "PARENT UID", "PARENT", "PARENT ID")
    col_templ = _pick_col(
        df.columns, "TEMPLATE NAME", "TEMPLATE", "SHEET", "STRUCTURE SHEET"
    )
    col_uclass = _pick_col(df.columns, "UNIT CLASS", "AGGREGATE NAME", "CLASS")
    col_2525 = _pick_col(df.columns, "2525C", "MILSTD2525C", "MILSTD 2525C")
    nodes: Dict[str, DragonUnitNode] = {}
    order_keys: List[str] = []
    for idx, row in df.iterrows():
        uid_raw = _norm(row.get(col_uid)) if col_uid else ""
        if not uid_raw:
            continue
        uid = uid_raw.strip()
        key = uid.upper()
        if key in nodes:
            continue
        name = _norm(row.get(col_name)) if col_name else ""
        template = _norm(row.get(col_templ)) if col_templ else ""
        parent_uid = _norm(row.get(col_puid)) if col_puid else ""
        unit_class = _norm(row.get(col_uclass)) if col_uclass else ""
        ms2525 = _norm(row.get(col_2525)) if col_2525 else ""
        node = DragonUnitNode(
            uid=uid,
            name=name or uid,
            template=template,
            parent_uid=parent_uid or None,
            unit_class=unit_class,
            ms2525=ms2525,
            order=idx,
            children=[],
        )
        nodes[key] = node
        order_keys.append(key)
    if not nodes:
        raise ValueError("No units with UID values were found in UNIT INFO sheet.")
    roots: List[DragonUnitNode] = []
    for key in order_keys:
        node = nodes[key]
        p_key = node.parent_uid.strip().upper() if node.parent_uid else ""
        if p_key and p_key in nodes:
            nodes[p_key].children.append(node)
        else:
            roots.append(node)
    return DragonWorkbookTree(roots=roots, node_map=nodes)


def di_import_dragon_enhanced(
    model: "ObsModel" | None,
    xlsx_path: str,
    unitlist_map: Optional[Dict[str, Dict[str, str]]] = None,
    default_side: Optional[str] = None,
    selected_uids: Optional[Set[str]] = None,
    *,
    progress_cb: Optional[Callable[[str, float], None]] = None,
    cancel_cb: Optional[Callable[[], bool]] = None,
) -> Tuple[ObsModel, int, int]:

    if cli_import_dragon_enhanced is None:
        raise RuntimeError("dragon_importer_new.import_dragon_enhanced is unavailable")
    if model is None:
        model = _new_blank_model()

    def _emit(message: str, fraction: float) -> None:
        if not progress_cb:
            return
        try:
            progress_cb(message, max(0.0, min(fraction, 1.0)))
        except Exception:
            pass

    def _check_cancel() -> None:
        if cancel_cb and cancel_cb():
            raise DragonImportCancelled()

    before_units = set(_iter_local(model.unit_list, "Unit"))
    _emit("Starting DRAGON workbook import...", 0.1)

    def _import_progress(message: str, inner_fraction: float) -> None:
        scaled = 0.2 + max(0.0, min(inner_fraction, 1.0)) * 0.6
        _emit(message, min(scaled, 0.85))
        _check_cancel()

    cli_import_dragon_enhanced(
        model,
        xlsx_path,
        allowed_uids=selected_uids,
        progress_cb=_import_progress,
        cancel_cb=cancel_cb,
    )
    _emit("Normalizing imported data...", 0.87)
    if cli_normalize_to_jtds_schema is not None:
        _check_cancel()
        cli_normalize_to_jtds_schema(model)
    _emit("Auto-filling LvcId fields...", 0.9)
    autofill_unique_lvcids(model)
    if unitlist_map:
        _emit("Applying reference UnitDisEnumeration map...", 0.94)
        for unit in _iter_local(model.unit_list, "Unit"):
            _check_cancel()
            code = (text(jfind(unit, "j:MilStd2525CCode")) or "").strip()
            if not code:
                continue
            attrs = unitlist_map.get(code) or unitlist_map.get(code.upper())
            if not attrs:
                continue
            ude = jfind(unit, "j:UnitDisEnumeration")
            if ude is None:
                ude = ET.SubElement(unit, jtag("UnitDisEnumeration"))
            for k, v in attrs.items():
                if ude.get(k) in (None, "") and v not in (None, ""):
                    ude.set(k, str(v))
    after_units = set(_iter_local(model.unit_list, "Unit"))
    new_units = after_units - before_units
    _emit("Finalizing import metadata...", 0.97)
    _check_cancel()
    unit_meta = ensure_unit_metadata(model, default_side, new_units)
    units = len(list(model.unit_list))
    ecs = len(list(model.entcomp_list))
    try:
        existing = getattr(model, "_dragon_counts", {})
        existing.update({"units": units, "ecs": ecs, "unit_meta": unit_meta})
        setattr(model, "_dragon_counts", existing)
    except Exception:
        pass
    _emit("Import complete", 1.0)
    return model, units, ecs


class DragonUnitSelectionDialog(QDialog):

    def __init__(
        self, parent, tree: DragonWorkbookTree, preselected: Optional[Set[str]] = None
    ):
        super().__init__(parent)
        self.setWindowTitle("Select DRAGON Units")
        self.resize(540, 520)
        self._tree = tree
        self._selected_root_uids: Set[str] = set()
        layout = QVBoxLayout(self)
        self.tree_widget = QTreeWidget()
        self.tree_widget.setColumnCount(3)
        self.tree_widget.setHeaderLabels(["Unit", "UID", "Template"])
        self.tree_widget.setAlternatingRowColors(True)
        layout.addWidget(self.tree_widget)
        prechecked = {uid.upper() for uid in (preselected or tree.node_map.keys())}
        self._populate_tree(tree.roots, prechecked)
        self.tree_widget.expandToDepth(1)
        self.tree_widget.itemChanged.connect(self._on_item_changed)
        ctrl_row = QHBoxLayout()
        btn_all = QPushButton("Select All")
        btn_none = QPushButton("Clear")
        btn_expand = QPushButton("Expand")
        btn_collapse = QPushButton("Collapse")
        btn_all.clicked.connect(lambda: self._set_all(Qt.CheckState.Checked))
        btn_none.clicked.connect(lambda: self._set_all(Qt.CheckState.Unchecked))
        btn_expand.clicked.connect(self.tree_widget.expandAll)
        btn_collapse.clicked.connect(self.tree_widget.collapseAll)
        for btn in (btn_all, btn_none, btn_expand, btn_collapse):
            ctrl_row.addWidget(btn)
        ctrl_row.addStretch()
        layout.addLayout(ctrl_row)
        buttons = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        buttons.accepted.connect(self._on_accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def _populate_tree(self, nodes: List[DragonUnitNode], prechecked: Set[str]) -> None:
        self.tree_widget.blockSignals(True)

        def add_node(
            node: DragonUnitNode, parent_item: Optional[QTreeWidgetItem]
        ) -> None:
            texts = [node.display_name(), node.uid, node.template or ""]
            item = QTreeWidgetItem(texts)
            item.setData(0, Qt.ItemDataRole.UserRole, node.uid.strip().upper())
            state = (
                Qt.CheckState.Checked
                if node.uid.strip().upper() in prechecked
                else Qt.CheckState.Unchecked
            )
            item.setCheckState(0, state)
            if parent_item is not None:
                parent_item.addChild(item)
            else:
                self.tree_widget.addTopLevelItem(item)
            for child in node.children:
                add_node(child, item)

        for node in nodes:
            add_node(node, None)
        self.tree_widget.blockSignals(False)
        self._refresh_parent_states()

    def _refresh_parent_states(self) -> None:
        def refresh(item: QTreeWidgetItem) -> Qt.CheckState:
            child_states = []
            for idx in range(item.childCount()):
                child_states.append(refresh(item.child(idx)))
            if not child_states:
                return item.checkState(0)
            if all(state == Qt.CheckState.Checked for state in child_states):
                state = Qt.CheckState.Checked
            elif all(state == Qt.CheckState.Unchecked for state in child_states):
                state = Qt.CheckState.Unchecked
            else:
                state = Qt.CheckState.PartiallyChecked
            self.tree_widget.blockSignals(True)
            item.setCheckState(0, state)
            self.tree_widget.blockSignals(False)
            return state

        for i in range(self.tree_widget.topLevelItemCount()):
            refresh(self.tree_widget.topLevelItem(i))

    def _set_all(self, state: Qt.CheckState) -> None:
        self.tree_widget.blockSignals(True)

        def set_item(item: QTreeWidgetItem) -> None:
            item.setCheckState(0, state)
            for idx in range(item.childCount()):
                set_item(item.child(idx))

        for i in range(self.tree_widget.topLevelItemCount()):
            set_item(self.tree_widget.topLevelItem(i))
        self.tree_widget.blockSignals(False)
        self._refresh_parent_states()

    def _on_item_changed(self, item: QTreeWidgetItem, column: int) -> None:
        if column != 0:
            return
        state = item.checkState(0)
        self.tree_widget.blockSignals(True)
        for idx in range(item.childCount()):
            item.child(idx).setCheckState(0, state)
        self.tree_widget.blockSignals(False)
        self._update_parent_state(item)

    def _update_parent_state(self, item: QTreeWidgetItem) -> None:
        parent = item.parent()
        while parent is not None:
            checked = 0
            unchecked = 0
            partial = 0
            for idx in range(parent.childCount()):
                ch_state = parent.child(idx).checkState(0)
                if ch_state == Qt.CheckState.Checked:
                    checked += 1
                elif ch_state == Qt.CheckState.Unchecked:
                    unchecked += 1
                else:
                    partial += 1
            self.tree_widget.blockSignals(True)
            if partial > 0 or (checked > 0 and unchecked > 0):
                parent.setCheckState(0, Qt.CheckState.PartiallyChecked)
            elif checked == parent.childCount():
                parent.setCheckState(0, Qt.CheckState.Checked)
            else:
                parent.setCheckState(0, Qt.CheckState.Unchecked)
            self.tree_widget.blockSignals(False)
            parent = parent.parent()

    def _collect_roots(self, item: QTreeWidgetItem, out: Set[str]) -> None:
        state = item.checkState(0)
        uid = _normalize_user_data(item.data(0, Qt.ItemDataRole.UserRole))
        if state == Qt.CheckState.Checked and uid:
            out.add(str(uid))
            return
        if state == Qt.CheckState.Unchecked:
            return
        for idx in range(item.childCount()):
            self._collect_roots(item.child(idx), out)

    def _on_accept(self) -> None:
        roots: Set[str] = set()
        for i in range(self.tree_widget.topLevelItemCount()):
            self._collect_roots(self.tree_widget.topLevelItem(i), roots)
        if not roots:
            QMessageBox.warning(
                self, "Select DRAGON Units", "Select at least one unit to import."
            )
            return
        self._selected_root_uids = roots
        super().accept()

    def selected_root_uids(self) -> Set[str]:
        return set(self._selected_root_uids)


class DragonImportWorker(QObject):

    progress = Signal(str, float)
    finished = Signal(object, int, int, bool)
    failed = Signal(str)
    cancelled = Signal()

    def __init__(
        self,
        model: ObsModel,
        dragon_path: str,
        ref_xml_path: Optional[str],
        selected_side: str,
        selected_uids: Optional[Set[str]],
    ):
        super().__init__()
        self._model = model
        self._dragon_path = dragon_path
        self._ref_xml_path = ref_xml_path
        self._selected_side = selected_side
        self._selected_uids = selected_uids
        self._cancel_requested = False

    def request_cancel(self) -> None:
        self._cancel_requested = True

    def _is_cancelled(self) -> bool:
        return self._cancel_requested

    @Slot()
    def run(self) -> None:
        try:
            self.progress.emit("Preparing import...", 0.0)
            ref_map: Optional[Dict[str, Dict[str, str]]] = None
            ref_applied = False
            if self._ref_xml_path:
                self.progress.emit("Loading reference unitlist.xml ...", 0.05)
                ref_map = load_unitlist_map(self._ref_xml_path)
                ref_applied = True

            def progress_cb(message: str, fraction: float) -> None:
                self.progress.emit(message, fraction)

            def cancel_cb() -> bool:
                return self._is_cancelled()

            base_model = self._model if self._model is not None else _new_blank_model()
            working_model = clone_obs_model(base_model)

            model, units, ecs = di_import_dragon_enhanced(
                working_model,
                self._dragon_path,
                ref_map,
                self._selected_side,
                self._selected_uids,
                progress_cb=progress_cb,
                cancel_cb=cancel_cb,
            )
            self.finished.emit(model, units, ecs, ref_applied)
        except DragonImportCancelled:
            self.cancelled.emit()
        except Exception as exc:
            traceback.print_exc()
            self.failed.emit(str(exc))


class ImportDragonTab(QWidget):

    modelImported = Signal(object)

    def __init__(self, parent=None):
        super().__init__(parent)
        self.model: Optional[ObsModel] = None
        self._dragon_tree: Optional[DragonWorkbookTree] = None
        self._dragon_tree_path: Optional[str] = None
        self._dragon_selected_roots: Optional[Set[str]] = None
        self._import_thread: Optional[QThread] = None
        self._import_worker: Optional[DragonImportWorker] = None
        self._pending_selected_uids: Optional[Set[str]] = None
        self._pending_side: Optional[str] = None
        self._build_ui()

    def _build_ui(self):
        outer = QVBoxLayout(self)
        row1 = QHBoxLayout()
        self.path_edit = QLineEdit()
        self.path_edit.setPlaceholderText("Select DRAGON workbook (.xlsx)...")
        self.path_edit.textChanged.connect(self._on_dragon_path_changed)
        btn1 = QPushButton("Browse...")
        btn1.clicked.connect(self.on_browse_dragon)
        row1.addWidget(self.path_edit)
        row1.addWidget(btn1)
        outer.addLayout(row1)
        row2 = QHBoxLayout()
        self.ref_edit = QLineEdit()
        self.ref_edit.setPlaceholderText("Select reference unitlist.xml (optional)")
        btn2 = QPushButton("Browse XML...")
        btn2.clicked.connect(self.on_browse_refxml)
        row2.addWidget(self.ref_edit)
        row2.addWidget(btn2)
        outer.addLayout(row2)
        self.log = QTextEdit()
        self.log.setReadOnly(True)
        self.log.setMinimumHeight(220)
        outer.addWidget(self.log)
        progress_row = QHBoxLayout()
        self.progress_bar = QProgressBar()
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setValue(0)
        self.progress_bar.setFormat("Idle")
        progress_row.addWidget(self.progress_bar, 1)
        self.btn_cancel_import = QPushButton("Cancel Import")
        self.btn_cancel_import.setEnabled(False)
        self.btn_cancel_import.clicked.connect(self._on_cancel_import)
        progress_row.addWidget(self.btn_cancel_import)
        outer.addLayout(progress_row)
        side_row = QHBoxLayout()
        side_row.addWidget(QLabel("Imported Side:"))
        self.side_group = QButtonGroup(self)
        self.side_buttons: Dict[str, QRadioButton] = {}
        for label in ("BLUFOR", "NEUTRAL", "OPFOR"):
            rb = QRadioButton(label)
            self.side_group.addButton(rb)
            self.side_buttons[label] = rb
            side_row.addWidget(rb)
            if label == "BLUFOR":
                rb.setChecked(True)
        side_row.addStretch()
        outer.addLayout(side_row)
        select_row = QHBoxLayout()
        self.btn_select_units = QPushButton("Select Units...")
        self.btn_select_units.setEnabled(False)
        self.btn_select_units.clicked.connect(self.on_select_units)
        select_row.addWidget(self.btn_select_units)
        self.selection_label = QLabel("Units: ALL")
        select_row.addWidget(self.selection_label)
        select_row.addStretch()
        outer.addLayout(select_row)
        grp = QGroupBox("Build Class Lists")
        gl = QGridLayout(grp)
        gl.addWidget(QLabel("EntityCompositionClass master (.xlsx):"), 0, 0)
        self.ecc_edit = QLineEdit()
        btn_ecc = QPushButton("Browse...")
        btn_ecc.clicked.connect(self.on_browse_ecc)
        gl.addWidget(self.ecc_edit, 0, 1)
        gl.addWidget(btn_ecc, 0, 2)
        gl.addWidget(QLabel("UnitClass master (.xlsx):"), 1, 0)
        self.ucl_edit = QLineEdit()
        btn_ucl = QPushButton("Browse...")
        btn_ucl.clicked.connect(self.on_browse_ucl)
        gl.addWidget(self.ucl_edit, 1, 1)
        gl.addWidget(btn_ucl, 1, 2)
        self.btn_build_classes = QPushButton("Build Classes")
        self.btn_build_classes.clicked.connect(self.on_build_classes)
        gl.addWidget(self.btn_build_classes, 2, 0, 1, 3)
        self.btn_append_ecc = QPushButton(
            "Append DRAGON Entity Classes to Master"
        )
        self.btn_append_ecc.clicked.connect(self.on_append_entityclasses)
        self.btn_append_ecc.setEnabled(
            cli_append_entityclasses_to_master is not None
        )
        gl.addWidget(self.btn_append_ecc, 3, 0, 1, 3)
        self.btn_append_ucl = QPushButton("Append DRAGON Unit Classes to Master")
        self.btn_append_ucl.clicked.connect(self.on_append_unitclasses)
        self.btn_append_ucl.setEnabled(cli_append_unitclasses_to_master is not None)
        gl.addWidget(self.btn_append_ucl, 4, 0, 1, 3)
        outer.addWidget(grp)
        self.btn_import = QPushButton("Import into Current Model (or New)")
        self.btn_import.clicked.connect(self.on_import)
        outer.addWidget(self.btn_import, alignment=Qt.AlignmentFlag.AlignRight)

    def _is_import_running(self) -> bool:
        return self._import_thread is not None

    def _set_import_running(self, running: bool) -> None:
        if running:
            self.progress_bar.setRange(0, 100)
            self.progress_bar.setValue(0)
            self.progress_bar.setFormat("Importing DRAGON workbook ...")
        else:
            self.progress_bar.setRange(0, 100)
            self.progress_bar.setValue(0)
            self.progress_bar.setFormat("Idle")
        enable = not running
        self.path_edit.setEnabled(enable)
        self.ref_edit.setEnabled(enable)
        self.ecc_edit.setEnabled(enable)
        self.ucl_edit.setEnabled(enable)
        self.btn_import.setEnabled(enable)
        self.btn_build_classes.setEnabled(enable)
        if hasattr(self, "btn_append_ecc") and self.btn_append_ecc is not None:
            self.btn_append_ecc.setEnabled(
                enable and cli_append_entityclasses_to_master is not None
            )
        if hasattr(self, "btn_append_ucl") and self.btn_append_ucl is not None:
            self.btn_append_ucl.setEnabled(
                enable and cli_append_unitclasses_to_master is not None
            )
        can_select_units = enable and bool(self.path_edit.text().strip())
        self.btn_select_units.setEnabled(can_select_units)
        self.btn_cancel_import.setEnabled(running)

    def _handle_import_progress(self, message: str, fraction: float) -> None:
        pct = max(0, min(int(fraction * 100), 100))
        self.progress_bar.setValue(pct)
        if message:
            self.progress_bar.setFormat(f"{message} ({pct}%)")
            self.log.append(message)

    def _on_cancel_import(self) -> None:
        if self._import_worker:
            self._import_worker.request_cancel()
            self.btn_cancel_import.setEnabled(False)
            self.progress_bar.setFormat("Canceling import...")
            self.log.append("Cancel requested. Waiting for current import to stop...")

    def _cleanup_import_state(self) -> None:
        self._import_thread = None
        self._import_worker = None
        self._pending_selected_uids = None
        self._pending_side = None
        self._set_import_running(False)

    def _handle_import_finished(
        self, model: ObsModel, units: int, ecs: int, ref_applied: bool
    ) -> None:
        self.set_model(model)
        selected_uids = self._pending_selected_uids
        selected_side = self._pending_side or ""
        self.log.append("Imported DRAGON workbook OK.")
        if selected_uids is not None:
            total = len(selected_uids)
            roots = len(self._dragon_selected_roots or [])
            self.log.append(
                f" - Imported selection: {roots} root unit(s), {total} workbook unit rows."
            )
        assigned = 0
        for unit in _iter_local(self.model.unit_list, "Unit"):
            side_el = jfind(unit, "j:SideSuperior")
            current = side_el.text.strip() if side_el is not None and side_el.text else ""
            if not current and selected_side:
                if side_el is None:
                    side_el = ET.SubElement(unit, jtag("SideSuperior"))
                side_el.text = selected_side
                assigned += 1
        if assigned:
            self.log.append(f" - Applied side {selected_side} to {assigned} units.")
        self.log.append(f" - Units total: {units}")
        self.log.append(f" - EntityCompositions total: {ecs}")
        if ref_applied:
            self.log.append(" - Applied reference UnitDisEnumeration map.")
        meta = (
            getattr(self.model, "_dragon_counts", {}).get("unit_meta")
            if self.model is not None
            else None
        )
        if meta:
            ids_fixed, stamps_fixed, sides_fixed = meta
            self.log.append(
                f" - Unit metadata normalized: ids={ids_fixed}, timestamps={stamps_fixed}, sides set={sides_fixed}"
            )
        self.modelImported.emit(self.model)
        self._cleanup_import_state()

    def _handle_import_failed(self, message: str) -> None:
        self.log.append(f"Import DRAGON failed: {message}")
        QMessageBox.warning(self, "Import DRAGON", f"Failed: {message}")
        self._cleanup_import_state()

    def _handle_import_cancelled(self) -> None:
        self.log.append("Import DRAGON was cancelled.")
        QMessageBox.information(self, "Import DRAGON", "Import cancelled by user.")
        self._cleanup_import_state()

    def set_model(self, model: Optional[ObsModel]):
        self.model = model

    def _on_dragon_path_changed(self, text: str) -> None:
        path = (text or "").strip()
        self.btn_select_units.setEnabled(bool(path))
        if not path:
            self._dragon_tree = None
            self._dragon_tree_path = None
            self._dragon_selected_roots = None
            self.selection_label.setText("Units: ALL")
        elif self._dragon_tree_path and path != self._dragon_tree_path:
            self._dragon_tree = None
            self._dragon_tree_path = None
            self._dragon_selected_roots = None
            self.selection_label.setText("Units: ALL")

    def _ensure_dragon_tree(self, path: str, *, reset_selection: bool = False) -> bool:
        if not path:
            return False
        if self._dragon_tree is not None and self._dragon_tree_path == path:
            return True
        try:
            tree = parse_dragon_unit_hierarchy(path)
        except Exception as exc:
            traceback.print_exc()
            QMessageBox.warning(
                self, "Parse DRAGON Workbook", f"Failed to parse workbook: {exc}"
            )
            return False
        self._dragon_tree = tree
        self._dragon_tree_path = path
        if reset_selection:
            self._dragon_selected_roots = None
            self.selection_label.setText("Units: ALL")
        self.log.append(
            f"Parsed DRAGON workbook: {tree.total_units()} units available."
        )
        return True

    def _resolve_master_path(self, raw_value: str, *, subfolder: str, filename: str) -> Path:
        raw = (raw_value or "").strip()
        if raw:
            resolved = Path(raw).expanduser()
            if resolved.is_dir():
                return resolved / filename
            return resolved
        candidates = _resource_search_paths()
        for base in candidates:
            candidate = base / subfolder / filename
            if candidate.exists():
                return candidate
        base = candidates[0] if candidates else Path.cwd()
        return base / subfolder / filename

    def _resolve_entityclass_master_path(self) -> Path:
        return self._resolve_master_path(
            self.ecc_edit.text(),
            subfolder="EntityCompositionClass-Masters",
            filename="entitycompositionclass-master.xlsx",
        )

    def _resolve_unitclass_master_path(self) -> Path:
        return self._resolve_master_path(
            self.ucl_edit.text(),
            subfolder="UnitClass-Masters",
            filename="unitclass-master.xlsx",
        )

    def on_browse_dragon(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Open DRAGON Workbook", "", "Excel Workbook (*.xlsx);;All Files (*)"
        )
        if path:
            self.path_edit.setText(path)
            self._ensure_dragon_tree(path, reset_selection=True)

    def on_browse_refxml(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Open reference UnitList XML", "", "OBS XML (*.xml);;All Files (*)"
        )
        if path:
            self.ref_edit.setText(path)

    def on_browse_ecc(self):
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Open EntityCompositionClass master",
            "",
            "Excel (*.xlsx);;All Files (*)",
        )
        if path:
            self.ecc_edit.setText(path)

    def on_browse_ucl(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Open UnitClass master", "", "Excel (*.xlsx);;All Files (*)"
        )
        if path:
            self.ucl_edit.setText(path)

    def on_append_entityclasses(self) -> None:
        if cli_append_entityclasses_to_master is None:
            QMessageBox.warning(
                self,
                "Append EntityCompositionClass master",
                "DRAGON helper is unavailable in this build.",
            )
            return
        dragon_path = self.path_edit.text().strip()
        if not dragon_path:
            QMessageBox.information(
                self,
                "Append EntityCompositionClass master",
                "Select a DRAGON workbook first.",
            )
            return
        master_path = self._resolve_entityclass_master_path()
        if not (self.ecc_edit.text() or "").strip():
            self.ecc_edit.setText(str(master_path))
        try:
            added = cli_append_entityclasses_to_master(
                dragon_path, str(master_path)
            )
        except FileNotFoundError as exc:
            QMessageBox.warning(self, "Append EntityCompositionClass master", str(exc))
            return
        except ValueError as exc:
            QMessageBox.warning(self, "Append EntityCompositionClass master", str(exc))
            return
        except Exception as exc:
            traceback.print_exc()
            QMessageBox.critical(
                self,
                "Append EntityCompositionClass master",
                f"Failed to update EntityCompositionClass master: {exc}",
            )
            return
        if added:
            msg = (
                f"Appended {added} new EntityCompositionClass row(s) to {master_path}"
            )
        else:
            msg = (
                "No new EntityCompositionClass rows were added; "
                f"{master_path} already contains workbook classes."
            )
        self.log.append(msg)
        QMessageBox.information(
            self, "Append EntityCompositionClass master", msg
        )

    def on_append_unitclasses(self) -> None:
        if cli_append_unitclasses_to_master is None:
            QMessageBox.warning(
                self,
                "Append UnitClass master",
                "DRAGON helper is unavailable in this build.",
            )
            return
        dragon_path = self.path_edit.text().strip()
        if not dragon_path:
            QMessageBox.information(
                self, "Append UnitClass master", "Select a DRAGON workbook first."
            )
            return
        master_path = self._resolve_unitclass_master_path()
        if not (self.ucl_edit.text() or "").strip():
            self.ucl_edit.setText(str(master_path))
        try:
            added = cli_append_unitclasses_to_master(dragon_path, str(master_path))
        except FileNotFoundError as exc:
            QMessageBox.warning(self, "Append UnitClass master", str(exc))
            return
        except ValueError as exc:
            QMessageBox.warning(self, "Append UnitClass master", str(exc))
            return
        except Exception as exc:
            traceback.print_exc()
            QMessageBox.critical(
                self,
                "Append UnitClass master",
                f"Failed to update UnitClass master: {exc}",
            )
            return
        if added:
            msg = f"Appended {added} new UnitClass row(s) to {master_path}"
        else:
            msg = f"No new UnitClass rows were added; {master_path} already contains workbook classes."
        self.log.append(msg)
        QMessageBox.information(self, "Append UnitClass master", msg)

    def on_select_units(self):
        path = self.path_edit.text().strip()
        if not path:
            QMessageBox.information(
                self, "No DRAGON file", "Select a DRAGON workbook first."
            )
            return
        if not self._ensure_dragon_tree(path, reset_selection=False):
            return
        tree = self._dragon_tree
        if tree is None:
            return
        total_root_uids = {node.uid.strip().upper() for node in tree.roots}
        preselected = self._dragon_selected_roots or total_root_uids
        dlg = DragonUnitSelectionDialog(self, tree, preselected)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            chosen = dlg.selected_root_uids()
            if not chosen or chosen == total_root_uids:
                self._dragon_selected_roots = None
                self.selection_label.setText("Units: ALL")
                self.log.append("Unit selection: ALL workbook units will be imported.")
            else:
                self._dragon_selected_roots = chosen
                names = []
                for uid in chosen:
                    node = tree.node_map.get(uid)
                    if node:
                        names.append(node.name or node.uid)
                preview = ", ".join(names[:3])
                if len(names) > 3:
                    preview += ", ..."
                self.selection_label.setText(
                    f"Units: {len(chosen)} root(s) ({preview})"
                )
                self.log.append(f"Unit selection: {len(chosen)} root(s) chosen.")

    def on_build_classes(self):
        if not self.model:
            QMessageBox.information(self, "No model", "Load or create a model first.")
            return
        ecc = self.ecc_edit.text().strip()
        ucl = self.ucl_edit.text().strip()
        try:
            c1 = (
                build_entitycomposition_classlist_from_xlsx(self.model, ecc)
                if ecc
                else 0
            )
            c2 = build_unitclass_list_from_xlsx(self.model, ucl) if ucl else 0
        except Exception as e:
            QMessageBox.warning(self, "Build Classes", f"Failed: {e}")
            return
        self.log.append(f"Built EntityCompositionClassList: {c1}  UnitClassList: {c2}")
        self.modelImported.emit(self.model)

    def on_import(self):
        if self._is_import_running():
            QMessageBox.information(
                self, "Import in progress", "Please wait for the current import to finish."
            )
            return
        path = self.path_edit.text().strip()
        if not path:
            QMessageBox.information(
                self, "No DRAGON file", "Please select a DRAGON .xlsx workbook."
            )
            return
        if self.model is None:
            self.model = _new_blank_model()
        selected_uids: Optional[Set[str]] = None
        if self._dragon_selected_roots:
            if not self._ensure_dragon_tree(path, reset_selection=False):
                return
            tree = self._dragon_tree
            if tree is None:
                return
            selected_uids = tree.gather_subtree_uids(self._dragon_selected_roots)
            if not selected_uids:
                QMessageBox.warning(
                    self, "Import DRAGON", "No units were selected to import."
                )
                return
        side_btn = self.side_group.checkedButton() if hasattr(self, "side_group") else None
        if side_btn is None:
            QMessageBox.information(
                self,
                "Select Side",
                "Please choose BLUFOR, NEUTRAL, or OPFOR for imported units.",
            )
            return
        selected_side = side_btn.text().strip().upper()
        ref_path = self.ref_edit.text().strip() or None
        self._pending_selected_uids = selected_uids
        self._pending_side = selected_side
        self._start_import_worker(path, selected_side, selected_uids, ref_path)

    def _start_import_worker(
        self,
        path: str,
        selected_side: str,
        selected_uids: Optional[Set[str]],
        ref_path: Optional[str],
    ) -> None:
        if self.model is None:
            self.model = _new_blank_model()
        worker = DragonImportWorker(
            self.model,
            path,
            ref_path,
            selected_side,
            selected_uids,
        )
        thread = QThread(self)
        worker.moveToThread(thread)
        thread.started.connect(worker.run)
        worker.progress.connect(self._handle_import_progress)
        worker.finished.connect(self._handle_import_finished)
        worker.failed.connect(self._handle_import_failed)
        worker.cancelled.connect(self._handle_import_cancelled)
        worker.finished.connect(thread.quit)
        worker.failed.connect(thread.quit)
        worker.cancelled.connect(thread.quit)
        thread.finished.connect(thread.deleteLater)
        worker.finished.connect(worker.deleteLater)
        worker.failed.connect(worker.deleteLater)
        worker.cancelled.connect(worker.deleteLater)
        thread.start()
        self._import_thread = thread
        self._import_worker = worker
        self._set_import_running(True)


class LicenseGateDialog(QDialog):

    def __init__(self, parent: Optional[QWidget] = None, *, initial_error: Optional[str] = None):
        super().__init__(parent)
        self.setWindowTitle("OBS Scrubber License Required")
        self._payload: Optional[LicensePayload] = None
        lay = QVBoxLayout(self)
        intro = QLabel(
            "A valid OBS Scrubber license file (.obsl) is required before the tool can run."
        )
        intro.setWordWrap(True)
        lay.addWidget(intro)
        path_row = QHBoxLayout()
        self.path_edit = QLineEdit(self._default_path())
        browse = QPushButton("Browse")
        browse.clicked.connect(self._choose_file)
        path_row.addWidget(self.path_edit, 1)
        path_row.addWidget(browse)
        lay.addLayout(path_row)
        self.status_label = QLabel(initial_error or "Select a license file and click Validate.")
        self.status_label.setWordWrap(True)
        lay.addWidget(self.status_label)
        button_row = QHBoxLayout()
        button_row.addStretch(1)
        self.validate_btn = QPushButton("Validate")
        self.validate_btn.clicked.connect(self._validate)
        self.continue_btn = QPushButton("Continue")
        self.continue_btn.clicked.connect(self.accept)
        self.continue_btn.setEnabled(False)
        cancel_btn = QPushButton("Exit")
        cancel_btn.clicked.connect(self.reject)
        button_row.addWidget(self.validate_btn)
        button_row.addWidget(self.continue_btn)
        button_row.addWidget(cancel_btn)
        lay.addLayout(button_row)
        if initial_error:
            self.status_label.setText(initial_error)

    @property
    def payload(self) -> Optional[LicensePayload]:
        return self._payload

    @staticmethod
    def _default_path() -> str:
        cached = cached_license_path()
        if cached:
            return str(cached)
        return str(Path.cwd() / DEFAULT_LICENSE_FILENAME)

    def _choose_file(self) -> None:
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Select License File",
            self.path_edit.text() or self._default_path(),
            "OBS License (*.obsl);;All Files (*)",
        )
        if path:
            self.path_edit.setText(path)
            self._payload = None
            self.continue_btn.setEnabled(False)

    def _validate(self) -> None:
        candidate = self.path_edit.text().strip()
        if not candidate:
            QMessageBox.warning(self, "License File", "Please select a license file.")
            return
        try:
            payload = load_license_file(candidate)
        except FileNotFoundError:
            QMessageBox.warning(self, "License File", "Selected file does not exist.")
            return
        except LicenseError as exc:
            QMessageBox.critical(self, "License Error", str(exc))
            self.continue_btn.setEnabled(False)
            self._payload = None
            return
        self._payload = payload
        remember_license_path(candidate)
        info = f"{describe_license(payload)}\nSource: {candidate}"
        self.status_label.setText(info)
        self.continue_btn.setEnabled(True)

    def accept(self) -> None:
        if self._payload is None:
            self._validate()
            if self._payload is None:
                return
        super().accept()


class MainWindow(QWidget):

    def __init__(self, license_payload: LicensePayload):
        super().__init__()
        self._license_payload = license_payload
        self.setWindowTitle(
            f"OBS Scrubber Light v{__version__} - JTDS OBS XML Analyzer/Scrubber/Generator/Importer"
        )
        self.resize(1080, 720)
        self.tabs = QTabWidget()
        self.tab_an = AnalyzerTab(self)
        self.tab_sc = ScrubberTab(self)
        self.tab_ge = GenerateTab(self)
        self.tab_ecc = EntityClassesTab(self)
        self.tab_ucl = UnitClassesTab(self)
        self.tab_imp = ImportDragonTab(self)
        self.tabs.addTab(self.tab_an, "1) Analyzer")
        self.tabs.addTab(self.tab_sc, "2) Scrubber")
        self.tabs.addTab(self.tab_ge, "3) Generate")
        self.tabs.addTab(self.tab_ecc, "4) Entity Classes")
        self.tabs.addTab(self.tab_ucl, "5) Unit Classes")
        self.tabs.addTab(self.tab_imp, "6) Import DRAGON")
        lay = QVBoxLayout(self)
        self.license_label = QLabel(self._license_summary())
        self.license_label.setAlignment(Qt.AlignmentFlag.AlignRight)
        lay.addWidget(self.license_label)
        lay.addWidget(self.tabs)
        self.tab_an.modelLoaded.connect(self.set_model)
        self.tab_imp.modelImported.connect(self.set_model)

    def set_model(self, model: Optional[ObsModel]):
        self.tab_an.set_model(model)
        self.tab_sc.set_model(model)
        self.tab_ge.set_model(model)
        self.tab_ecc.set_model(model)
        self.tab_ucl.set_model(model)
        self.tab_imp.set_model(model)

    def _license_summary(self) -> str:
        path_hint = (
            Path(self._license_payload.path).name
            if self._license_payload and self._license_payload.path
            else DEFAULT_LICENSE_FILENAME
        )
        return f"License: {describe_license(self._license_payload)} ({path_hint})"


def _obtain_license(parent: Optional[QWidget] = None) -> Optional[LicensePayload]:
    try:
        return load_first_valid_license()
    except LicenseError as exc:
        dlg = LicenseGateDialog(parent=parent, initial_error=str(exc))
        if dlg.exec() == QDialog.DialogCode.Accepted and dlg.payload:
            return dlg.payload
        return None


def main() -> int:

    app = QApplication(sys.argv)
    license_payload = _obtain_license()
    if license_payload is None:
        QMessageBox.critical(
            None,
            "License Required",
            "A valid OBS Scrubber license (.obsl) is required. The application will now exit.",
        )
        return 1
    global LICENSE_INFO
    LICENSE_INFO = license_payload
    w = MainWindow(license_payload)
    w.show()
    return app.exec()


if __name__ == "__main__":

    sys.exit(main())
